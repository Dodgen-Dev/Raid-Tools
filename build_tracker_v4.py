"""
Raid Tracker v3 — Reads from pandas DataFrames
Reads raid_dataframes.xlsx + roster.json → two xlsx files:
  1. Raid Tracker.xlsx  — clean display (Summary, Raids, Raid Performance, etc.)
  2. Raid Data Tables.xlsx — raw data tables for analysis
Formulas use XLOOKUP — works in Google Sheets natively.
"""
import json, os, re, sys
from datetime import datetime, timezone
from collections import defaultdict, Counter, Counter

# Force UTF-8 output regardless of Windows console encoding
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import pandas as pd

# ── Load data from xlsx (single source of truth) ──
if getattr(sys, "frozen", False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.json")
RAID_DATA = os.environ.get("RAID_DATA", "raid_dataframes.xlsx")
ROSTER_FILE = os.environ.get("ROSTER_FILE", "roster.json")


def _read_json(path):
    """Read a JSON file, handling both UTF-8 and Windows cp1252 encoding."""
    raw = open(path, "rb").read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("cp1252")
        try:
            with open(path, "w", encoding="utf-8") as fix:
                fix.write(text)
        except OSError:
            pass
    return json.loads(text)


# Read output_dir from config.json (falls back to env var or default)
_config = {}
if os.path.isfile(CONFIG_FILE):
    _config = _read_json(CONFIG_FILE)

OUTPUT_DIR = os.environ.get("OUTPUT_DIR",
    _config.get("output_dir",
                r"G:\My Drive\Detention Raid\Raid Metrics"))

sys.path.insert(0, SCRIPT_DIR)
try:
    from raid_pull import load_raid_dataframes
    dfs = load_raid_dataframes(RAID_DATA)
except Exception:
    # Fallback: inline loader if raid_pull isn't available
    raw_sheets = pd.read_excel(RAID_DATA, sheet_name=None)
    schema = raw_sheets.pop("_dtypes", None)
    if schema is not None:
        for name, df in raw_sheets.items():
            sheet_schema = schema[schema["sheet"] == name]
            for _, row in sheet_schema.iterrows():
                col, dtype = row["column"], row["dtype"]
                if col not in df.columns: continue
                try:
                    if dtype == "bool": df[col] = df[col].astype(bool)
                    elif dtype == "int64": df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
                    elif dtype == "float64": df[col] = pd.to_numeric(df[col], errors="coerce")
                    elif dtype == "object": df[col] = df[col].where(df[col].notna(), None)
                except: pass
    dfs = raw_sheets

print(f"Loaded {len(dfs)} sheets from {RAID_DATA}")

roster_locked = False
roster_auto_created = False

if os.path.isfile(ROSTER_FILE):
    roster = _read_json(ROSTER_FILE)
    roster_locked = roster.get("meta", {}).get("locked", False)
    print(f"Loaded roster.json ({'LOCKED' if roster_locked else 'unlocked'})")
else:
    # Auto-create roster from player data
    print("  No roster.json found — will auto-create from raid data.")
    roster = {"meta": {"last_updated": "", "source": "auto-created",
                       "locked": False}, "players": {}, "unlinked": []}
    roster_auto_created = True

# Build set of all rostered character names (for filtering when locked)
rostered_chars = set()
for pdata in roster.get("players", {}).values():
    for c in pdata.get("mains", []):
        rostered_chars.add(c)
    for c in pdata.get("alts", []):
        rostered_chars.add(c)
for c in roster.get("unlinked", []):
    rostered_chars.add(c)
rostered_players = set(roster.get("players", {}).keys())

# ── Read editable roster fields from existing xlsx (if it exists) ──
# User edits on Roster sheet are THE TRUTH — they override all auto-detected values.
# Editable columns: B=Player, E=Role, F=Main/Alt, G=Notes
EXISTING_XLSX = os.environ.get("EXISTING_XLSX", "")
roster_overrides = {}  # character → {player, role, main_alt, notes}

if EXISTING_XLSX and os.path.exists(EXISTING_XLSX):
    try:
        from openpyxl import load_workbook
        prev_wb = load_workbook(EXISTING_XLSX, data_only=True, read_only=True)
        if "Roster" in prev_wb.sheetnames:
            prev_ws = prev_wb["Roster"]
            for row in prev_ws.iter_rows(min_row=2, values_only=False):
                vals = [c.value for c in row]
                # Layout: A=#, B=Player, C=Character, D=Realm, E=Role,
                #         F=Main/Alt, G=Notes, H+=auto
                if len(vals) < 7 or not vals[2]: continue
                char_name = str(vals[2]).strip()
                if char_name in ("Character", "MAINS", "ALTS", ""): continue

                player_ov = str(vals[1]).strip() if vals[1] else ""
                if player_ov in ("Player", "#", ""): player_ov = ""
                role_ov = str(vals[4]).strip() if vals[4] else ""
                if role_ov in ("Role", ""): role_ov = ""
                mainalt_ov = str(vals[5]).strip() if vals[5] else ""
                if mainalt_ov in ("Main/Alt", ""): mainalt_ov = ""
                notes = str(vals[6]).strip() if vals[6] else ""
                if notes in ("Notes",): notes = ""

                if player_ov or role_ov or mainalt_ov or notes:
                    roster_overrides[char_name] = {
                        "player": player_ov,
                        "role": role_ov,
                        "main_alt": mainalt_ov,
                        "notes": notes,
                    }
            print(f"  Loaded {len(roster_overrides)} roster overrides from {EXISTING_XLSX}")
            for char, ov in sorted(roster_overrides.items()):
                parts = []
                if ov["player"]: parts.append(f"player={ov['player']}")
                if ov["role"]: parts.append(f"role={ov['role']}")
                if ov["main_alt"]: parts.append(f"main_alt={ov['main_alt']}")
                if ov["notes"]: parts.append(f"notes={ov['notes']}")
                print(f"    {char}: {', '.join(parts)}")
        prev_wb.close()
    except Exception as e:
        print(f"  WARN: Could not read existing xlsx: {e}")
else:
    if EXISTING_XLSX:
        print(f"  WARN: EXISTING_XLSX={EXISTING_XLSX} not found, starting fresh")
    else:
        print("  No EXISTING_XLSX set — roster overrides skipped")

# ── Reverse lookup: character → (player, main/alt) ──
char_to_player = {}
char_to_mainalt = {}
for player_name, pdata in roster["players"].items():
    for c in pdata.get("mains", []):
        char_to_player[c] = player_name
        char_to_mainalt[c] = "Main"
    for c in pdata.get("alts", []):
        char_to_player[c] = player_name
        char_to_mainalt[c] = "Alt"
for c in roster.get("unlinked", []):
    char_to_player[c] = c
    char_to_mainalt[c] = "Alt"

# ── Apply xlsx overrides for Player and Main/Alt (these are THE TRUTH) ──
for char_name, ov in roster_overrides.items():
    if ov.get("player"):
        char_to_player[char_name] = ov["player"]
    if ov.get("main_alt"):
        char_to_mainalt[char_name] = ov["main_alt"]


SKIP_SLOTS = {"SHIRT", "TABARD"}
ENCHANTABLE = {"CHEST", "LEGS", "FEET", "WRIST", "BACK", "FINGER_1", "FINGER_2", "MAIN_HAND"}
MAX_TIER = 5
TIER_SLOTS = {"HEAD", "SHOULDER", "CHEST", "HANDS", "LEGS"}

# ══════════════════════════════════════════════════════════════════
#  EXTRACT FLAT TABLES FROM DATAFRAMES
# ══════════════════════════════════════════════════════════════════

df_perf = dfs.get("player_performance", pd.DataFrame())
df_fights = dfs.get("fights", pd.DataFrame())
df_deaths = dfs.get("deaths", pd.DataFrame())
df_players = dfs.get("players", pd.DataFrame())
df_equipment = dfs.get("player_equipment", pd.DataFrame())
df_fight_summary = dfs.get("player_fight_summary", pd.DataFrame())
df_meta = dfs.get("meta", pd.DataFrame())
df_guild_rankings = dfs.get("guild_rankings", pd.DataFrame())
df_pdt = dfs.get("player_damage_taken", pd.DataFrame())
df_consumables = dfs.get("consumables", pd.DataFrame())
df_mtd = dfs.get("mechanic_target_damage", pd.DataFrame())
df_fight_roster = dfs.get("fight_roster", pd.DataFrame())
df_actors = dfs.get("actors", pd.DataFrame())
df_debuffs = dfs.get("debuffs", pd.DataFrame())
df_interrupts = dfs.get("interrupts", pd.DataFrame())
df_dispels = dfs.get("dispels", pd.DataFrame())
df_death_events = dfs.get("death_events", pd.DataFrame())

# -- Rankings with ROLE FROM WCL --
ranking_rows = []
char_role_from_wcl = {}

# Build guid lookup from player_fight_summary for wcl_char_id
_guid_lookup = {}
if len(df_fight_summary):
    for _, r in df_fight_summary.iterrows():
        name = r.get("player", "?")
        guid = r.get("guid", 0)
        if guid and name not in _guid_lookup:
            _guid_lookup[name] = guid

# Build ilvl lookup from player_fight_summary (most recent per player)
_ilvl_lookup = {}
if len(df_fight_summary):
    pfs_sorted = df_fight_summary.sort_values("date", ascending=False) if "date" in df_fight_summary.columns else df_fight_summary
    for _, r in pfs_sorted.iterrows():
        name = r.get("player", "?")
        ilvl = r.get("item_level")
        if name not in _ilvl_lookup and pd.notna(ilvl) and ilvl:
            _ilvl_lookup[name] = int(ilvl)

# Build actor_id lookup: (report_code, character_name) → actor_id (for WoWAnalyzer deep links)
_actor_id_lookup = {}
if len(df_actors):
    for _, r in df_actors.iterrows():
        key = (r.get("report_code", ""), r.get("name", ""))
        aid = r.get("actor_id")
        if aid and pd.notna(aid):
            _actor_id_lookup[key] = int(aid)

for _, r in df_perf.iterrows():
    char_name = r.get("player", "?")
    role_key = str(r.get("role", "")).strip()
    role_label = {"Tank": "Tank", "Healer": "Healer", "DPS": "DPS",
                  "tank": "Tank", "healer": "Healer", "dps": "DPS", "dp": "DPS"}.get(role_key, "")
    metric = r.get("metric", "")
    kill = bool(r.get("kill", False))

    # On wipe fights, skip characters not in the roster (filters pets/NPCs like Shooting Star)
    if not kill and char_name not in char_to_player:
        continue

    if role_label and kill:
        char_role_from_wcl[char_name] = role_label

    ranking_rows.append({
        "report_code": r.get("report_code", ""),
        "date": str(r.get("date", "")),
        "fight_id": int(r.get("fight_id", 0)),
        "boss": r.get("boss", ""),
        "kill": kill,
        "metric": metric,
        "character": char_name,
        "player": char_to_player.get(char_name, char_name),
        "main_alt": char_to_mainalt.get(char_name, "Alt"),
        "class": r.get("player_class", ""),
        "spec": r.get("player_spec", ""),
        "role": role_label,
        "parse_pct": r.get("rank_percent") if pd.notna(r.get("rank_percent")) else "",
        "bracket_pct": r.get("bracket_percent") if pd.notna(r.get("bracket_percent")) else "",
        "ilvl": _ilvl_lookup.get(char_name, ""),
        "amount": round(float(r.get("amount", 0)), 1),
        "wcl_char_id": _guid_lookup.get(char_name, ""),
    })

# ── Apply xlsx Role overrides (THE TRUTH) ──
for char_name, ov in roster_overrides.items():
    if ov.get("role"):
        char_role_from_wcl[char_name] = ov["role"]

# Patch ranking_rows to use overridden roles
for r in ranking_rows:
    ov_role = roster_overrides.get(r["character"], {}).get("role", "")
    if ov_role:
        r["role"] = ov_role

# -- Fights --
WIPE_DEATH_THRESHOLD = 4

fight_rows = []
# Precompute death counts per fight
_death_counts = {}
if len(df_deaths):
    for (rc, fid), grp in df_deaths.groupby(["report_code", "fight_id"]):
        _death_counts[(rc, int(fid))] = len(grp)

for _, f in df_fights.iterrows():
    rc = f.get("report_code", "")
    fid = int(f.get("fight_id", 0))
    kill = bool(f.get("kill", False))
    dur_s = float(f.get("duration_s", 0))
    raw_deaths = _death_counts.get((rc, fid), 0)
    deaths_count = raw_deaths if kill else min(raw_deaths, WIPE_DEATH_THRESHOLD)

    fight_rows.append({
        "report_code": rc,
        "date": str(f.get("date", "")),
        "fight_id": fid,
        "boss": f.get("boss", ""),
        "encounter_id": int(f.get("encounter_id", 0)),
        "kill": kill,
        "duration_s": round(dur_s, 1),
        "duration_fmt": f"{int(dur_s // 60)}:{int(dur_s % 60):02d}",
        "size": int(f.get("size", 0)),
        "boss_pct": f.get("boss_pct", 0) if pd.notna(f.get("boss_pct")) else 0,
        "deaths_total": deaths_count,
        "avg_ilvl": f.get("avg_ilvl", "") if pd.notna(f.get("avg_ilvl")) else "",
    })

# -- Deaths --
death_rows = []
if len(df_deaths):
    for _, d in df_deaths.iterrows():
        char_name = d.get("player", "?")
        death_rows.append({
            "report_code": d.get("report_code", ""),
            "date": str(d.get("date", "")),
            "fight_id": int(d.get("fight_id", 0)),
            "boss": d.get("boss", ""),
            "kill": bool(d.get("kill", False)),
            "character": char_name,
            "player": char_to_player.get(char_name, char_name),
            "main_alt": char_to_mainalt.get(char_name, "Alt"),
            "killing_blow": d.get("killing_blow_name", ""),
            "overkill": int(d.get("overkill", 0)) if pd.notna(d.get("overkill")) else 0,
            "death_window_ms": 0,  # not captured in new xlsx
            "timestamp_ms": int(d.get("timestamp_ms", 0)),
            "death_order": int(d.get("death_order", 0)),
        })

# ── Wipe death filter: keep all kill deaths + first 4 deaths on wipe pulls ──
pre_filter = len(death_rows)
filtered_death_rows = []
from itertools import groupby
death_rows.sort(key=lambda d: (d["report_code"], d["fight_id"], d["timestamp_ms"]))
for (rc, fid), group in groupby(death_rows, key=lambda d: (d["report_code"], d["fight_id"])):
    group_list = list(group)
    if group_list[0]["kill"]:
        filtered_death_rows.extend(group_list)
    else:
        filtered_death_rows.extend(group_list[:WIPE_DEATH_THRESHOLD])
death_rows = filtered_death_rows
print(f"Deaths: {pre_filter} total → {len(death_rows)} after wipe filter (kills + first {WIPE_DEATH_THRESHOLD} on wipes)")

# -- Roster with ROLE FROM WCL --
roster_rows = []
_char_tier_prefix = {}  # char → set of item names that are tier
for _, p in df_players.iterrows():
    char_name = p.get("player", "?")
    spec = p.get("wcl_spec", "")
    role = char_role_from_wcl.get(char_name, "")
    ilvl = _ilvl_lookup.get(char_name, "")
    if not ilvl:
        ilvl = p.get("ilvl", "") if pd.notna(p.get("ilvl")) else ""

    # Raid prog
    mc = p.get("mythic_completed")
    mt = p.get("mythic_total")
    if pd.notna(mc) and pd.notna(mt) and mc:
        raid_prog = f"{int(mc)}/{int(mt)} M"
    else:
        raid_prog = ""

    # Gear audit from player_equipment
    char_gear = df_equipment[df_equipment["player"] == char_name] if len(df_equipment) else pd.DataFrame()
    missing_ench = 0
    empty_sockets = 0
    ilvls_list = []
    _char_tier_items = set()  # will hold item names identified as tier for this char
    tier_slot_first_words = {}  # slot → first word of item name
    _tier_slot_names = {}       # slot → full item name
    for _, it in char_gear.iterrows():
        slot = it.get("slot", "")
        it_ilvl = it.get("item_ilvl")
        if pd.notna(it_ilvl) and it_ilvl:
            ilvls_list.append((int(it_ilvl), slot))
        if it.get("needs_enchant", False):
            missing_ench += 1
        es = it.get("empty_sockets", 0)
        if pd.notna(es):
            empty_sockets += int(es)
        # Collect item names in tier slots for set detection
        if slot in TIER_SLOTS:
            item_name = str(it.get("item_name", "")).strip()
            if item_name:
                _tier_slot_names[slot] = item_name
                tier_slot_first_words[slot] = item_name.split("'")[0] if "'" in item_name else item_name.split()[0]

    # Tier detection: count matching items in tier slots (HEAD/SHOULDER/CHEST/HANDS/LEGS)
    # Some sets share a first word ("Augur's X", "Charhound's Y")
    # Others share last 2 words ("X of Channeled Fury", "Y of the Lucent Battalion")
    # Take whichever pattern matches more slots
    tier_count = 0
    if tier_slot_first_words:
        # First-word matching (e.g., "Augur's Ephemeral Wide-Brim" → "Augur")
        prefix_counts = Counter(tier_slot_first_words.values())
        best_prefix_count = prefix_counts.most_common(1)[0][1] if prefix_counts else 0

        # Last-2-word matching (e.g., "Aspect of Channeled Fury" → "Channeled Fury")
        suffixes = {}
        for slot, name in _tier_slot_names.items():
            words = name.split()
            if len(words) >= 3:
                suffixes[slot] = " ".join(words[-2:])
        suffix_counts = Counter(suffixes.values()) if suffixes else Counter()
        best_suffix_count = suffix_counts.most_common(1)[0][1] if suffix_counts else 0

        tier_count = max(best_prefix_count, best_suffix_count)
        if tier_count < 2:
            tier_count = 0  # need at least 2 matching to call it a set

        # Determine which items are tier: items matching the winning pattern
        if best_prefix_count >= best_suffix_count and best_prefix_count >= 2:
            winning_prefix = prefix_counts.most_common(1)[0][0]
            for slot, fw in tier_slot_first_words.items():
                if fw == winning_prefix:
                    _char_tier_items.add(_tier_slot_names[slot])
        elif best_suffix_count >= 2:
            winning_suffix = suffix_counts.most_common(1)[0][0]
            for slot, name in _tier_slot_names.items():
                words = name.split()
                if len(words) >= 3 and " ".join(words[-2:]) == winning_suffix:
                    _char_tier_items.add(name)

    _char_tier_prefix[char_name] = _char_tier_items
    tier_count = min(tier_count, MAX_TIER)

    lowest_ilvl = ""
    lowest_slot = ""
    if ilvls_list:
        min_ilvl = min(i for i, s in ilvls_list)
        max_ilvl = max(i for i, s in ilvls_list)
        if min_ilvl == max_ilvl:
            lowest_ilvl = "—"
        else:
            lowest_slot_item = min(ilvls_list, key=lambda x: x[0])
            lowest_ilvl = lowest_slot_item[0]
            lowest_slot = lowest_slot_item[1]

    roster_rows.append({
        "character": char_name,
        "player": char_to_player.get(char_name, char_name),
        "main_alt": char_to_mainalt.get(char_name, "Alt"),
        "class": p.get("wcl_class", ""),
        "spec": spec, "role": role,
        "server": p.get("wcl_server", ""),
        "realm_slug": p.get("realm_slug", ""),
        "ilvl": ilvl, "raid_prog": raid_prog,
        "missing_ench": missing_ench, "empty_sockets": empty_sockets,
        "tier_count": tier_count,
        "lowest_ilvl": lowest_ilvl, "lowest_slot": lowest_slot,
        "has_gear_data": len(char_gear) > 0,
    })

# Sort: Mains first, then role (Tank > Healer > DPS), then alpha
role_order = {"Tank": 0, "Healer": 1, "DPS": 2, "": 3}
roster_rows.sort(key=lambda r: (0 if r["main_alt"] == "Main" else 1, role_order.get(r["role"], 3), r["character"]))

# ── Auto-create roster.json if it didn't exist ──
if roster_auto_created:
    auto_players = {}
    for rr in roster_rows:
        char = rr["character"]
        player = rr["player"]  # defaults to char name if unknown
        if player not in auto_players:
            auto_players[player] = {"mains": [], "alts": []}
        # First character seen for a player becomes main
        if not auto_players[player]["mains"]:
            auto_players[player]["mains"].append(char)
        elif char not in auto_players[player]["mains"]:
            auto_players[player]["alts"].append(char)

    roster["players"] = auto_players
    roster["meta"]["last_updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    roster["meta"]["source"] = "auto-created from raid data"
    roster["meta"]["locked"] = False

    with open(ROSTER_FILE, "w", encoding="utf-8") as f:
        json.dump(roster, f, indent=2, ensure_ascii=False)
    print(f"  Auto-created roster.json with {len(auto_players)} players")

    # Rebuild rostered sets from new data
    rostered_chars = set()
    for pdata in auto_players.values():
        for c in pdata.get("mains", []):
            rostered_chars.add(c)
        for c in pdata.get("alts", []):
            rostered_chars.add(c)
    rostered_players = set(auto_players.keys())

# Build char → class lookup for class-color fills
char_to_class = {r["character"]: r["class"] for r in roster_rows}

# -- Gear rows (for data sheet) --
gear_rows = []
for _, it in df_equipment.iterrows():
    char_name = it.get("player", "?")
    gear_rows.append({
        "character": char_name,
        "player": char_to_player.get(char_name, char_name),
        "slot": it.get("slot", ""),
        "item_name": it.get("item_name", ""),
        "ilvl": int(it["item_ilvl"]) if pd.notna(it.get("item_ilvl")) else "",
        "quality": it.get("quality", ""),
        "enchant": it.get("enchant", ""),
        "needs_enchant": bool(it.get("needs_enchant", False)),
        "gems": int(it.get("gems", 0)) if pd.notna(it.get("gems")) else 0,
        "empty_sockets": int(it.get("empty_sockets", 0)) if pd.notna(it.get("empty_sockets")) else 0,
        "is_tier": (it.get("item_name", "") in _char_tier_prefix.get(char_name, set())),
    })

# ══════════════════════════════════════════════════════════════════
#  COMPUTED AGGREGATES
# ══════════════════════════════════════════════════════════════════

all_dates = sorted(set(r["date"] for r in fight_rows), reverse=True)
this_week = all_dates[0] if all_dates else ""
last_week = all_dates[1] if len(all_dates) > 1 else ""
num_nights = len(all_dates)

# Kills per char per boss
kills_by_char_boss = defaultdict(lambda: defaultdict(int))
week_kills_by_char_boss = defaultdict(lambda: defaultdict(int))
for r in ranking_rows:
    if r["kill"] and r["metric"] == "dps":
        kills_by_char_boss[r["character"]][r["boss"]] += 1
        if r["date"] == this_week:
            week_kills_by_char_boss[r["character"]][r["boss"]] += 1

# Best parse per char per boss (DPS for non-healers, HPS for healers)
best_parse_by_char_boss = defaultdict(lambda: defaultdict(lambda: {"parse": 0, "amount": 0}))
for r in ranking_rows:
    use = False
    if r["metric"] == "dps" and r["role"] != "Healer": use = True
    if r["metric"] == "hps" and r["role"] == "Healer": use = True
    if use:
        key = best_parse_by_char_boss[r["character"]][r["boss"]]
        if r["parse_pct"] and r["parse_pct"] > key["parse"]:
            key["parse"] = r["parse_pct"]
            key["amount"] = r["amount"]

# Avg parse per char per boss (for AVG PERFORMANCE section)
avg_parse_by_char_boss = defaultdict(lambda: defaultdict(lambda: {"parses": [], "amounts": []}))
for r in ranking_rows:
    use = False
    if r["metric"] == "dps" and r["role"] != "Healer": use = True
    if r["metric"] == "hps" and r["role"] == "Healer": use = True
    if use and r["parse_pct"]:
        avg_parse_by_char_boss[r["character"]][r["boss"]]["parses"].append(r["parse_pct"])
        avg_parse_by_char_boss[r["character"]][r["boss"]]["amounts"].append(r["amount"])

# Deaths per character total + per night
deaths_by_char = defaultdict(int)
deaths_by_char_date = defaultdict(lambda: defaultdict(int))
for d in death_rows:
    deaths_by_char[d["character"]] += 1
    deaths_by_char_date[d["character"]][d["date"]] += 1

# Per-date averages for "vs Avg" / "most improved" / CV chart
char_avg_parse_by_date = defaultdict(lambda: defaultdict(list))
for r in ranking_rows:
    use = (r["metric"] == "dps" and r["role"] != "Healer") or \
          (r["metric"] == "hps" and r["role"] == "Healer")
    if use and r["parse_pct"]:
        char_avg_parse_by_date[r["character"]][r["date"]].append(r["parse_pct"])

# Raid-wide avg DPS per date (dps metric, non-healer — kills + wipes with data)
raid_avg_dps_by_date = defaultdict(list)
for r in ranking_rows:
    if r["metric"] == "dps" and r["role"] != "Healer" and r["amount"]:
        if roster_locked and r["character"] not in rostered_chars:
            continue
        raid_avg_dps_by_date[r["date"]].append(r["amount"])

# Per-char DPS amounts by date (for CV chart — use primary metric per role)
char_dps_by_date = defaultdict(lambda: defaultdict(list))
for r in ranking_rows:
    use = (r["metric"] == "dps" and r["role"] != "Healer") or \
          (r["metric"] == "hps" and r["role"] == "Healer")
    if use and r["amount"]:
        char_dps_by_date[r["character"]][r["date"]].append(r["amount"])

mains_list = [r for r in roster_rows if r["main_alt"] == "Main"]
alts_linked = [r for r in roster_rows if r["main_alt"] == "Alt" and
               char_to_player.get(r["character"], r["character"]) != r["character"]]
alts_unlinked = [r for r in roster_rows if r["main_alt"] == "Alt" and
                 char_to_player.get(r["character"], r["character"]) == r["character"]]

print(f"Data: {len(roster_rows)} chars, {len(fight_rows)} fights, "
      f"{len(ranking_rows)} rankings, {len(death_rows)} deaths, {len(gear_rows)} gear")
print(f"Dates: {all_dates}")
print(f"Roles from WCL: {dict(sorted(defaultdict(int, {v: sum(1 for x in char_role_from_wcl.values() if x==v) for v in set(char_role_from_wcl.values())}).items()))}")

# ══════════════════════════════════════════════════════════════════
#  SCORING ENGINE
# ══════════════════════════════════════════════════════════════════
import numpy as np

# Load mechanic rulesets
RULESETS_PATH = os.path.join(SCRIPT_DIR, "mechanic_rulesets.json")
if not os.path.exists(RULESETS_PATH):
    RULESETS_PATH = os.path.join(os.path.dirname(RAID_DATA), "mechanic_rulesets.json")
if os.path.exists(RULESETS_PATH):
    mech_rulesets = _read_json(RULESETS_PATH)
    print(f"Loaded mechanic_rulesets.json v{mech_rulesets.get('version','?')}")
else:
    mech_rulesets = {"ignore_abilities": [], "bosses": {},
                     "scoring_weights": {"mechanics":0.40,"deaths":0.35,
                                         "parse_performance":0.20,
                                         "consumables":0.05},
                     "grade_thresholds": {"A":90,"B":80,"C":70,"D":60,"F":0},
                     "scoring_methods": {}, "immune_classes": []}
    print("WARN: mechanic_rulesets.json not found — scoring will be limited")

# Belt-and-suspenders: fill in defaults if keys are missing
_default_weights = {"mechanics":0.40,"deaths":0.35,"parse_performance":0.20,"consumables":0.05}
_default_grades = {"A":90,"B":80,"C":70,"D":60,"F":0}
if "scoring_weights" not in mech_rulesets:
    mech_rulesets["scoring_weights"] = _default_weights
if "grade_thresholds" not in mech_rulesets:
    mech_rulesets["grade_thresholds"] = _default_grades

SCORE_WEIGHTS = mech_rulesets["scoring_weights"]
GRADE_THRESHOLDS = mech_rulesets["grade_thresholds"]
TANK_SPECS = {"Protection", "Blood", "Brewmaster", "Vengeance", "Guardian"}
IMMUNE_CLASSES = set(mech_rulesets.get("immune_classes", []))
IGNORE_ABILITIES = set(mech_rulesets.get("ignore_abilities", []))

# ── Boss order and short names — data-driven from mechanic_rulesets.json ──
# raids: {"Manaforge Omega": ["Boss1", "Boss2", ...], "Raid2": [...]}
# Each boss entry has optional "short_name" field.
_raids_cfg = mech_rulesets.get("raids", {})
_bosses_cfg = mech_rulesets.get("bosses", {})

if _raids_cfg:
    # Concatenate all raid boss lists in order
    BOSS_ORDER = []
    for raid_name_key, boss_list in _raids_cfg.items():
        for b in boss_list:
            if b not in BOSS_ORDER:
                BOSS_ORDER.append(b)
    print(f"Boss order: {len(BOSS_ORDER)} bosses from "
          f"{len(_raids_cfg)} raid(s): {list(_raids_cfg.keys())}")
else:
    # Fallback: use all boss keys from mechanic_rulesets (unordered)
    BOSS_ORDER = list(_bosses_cfg.keys())
    if BOSS_ORDER:
        print(f"Boss order: {len(BOSS_ORDER)} bosses (no 'raids' section "
              f"— using bosses dict order)")
    else:
        # Last resort: discover from data
        BOSS_ORDER = sorted(set(r.get("boss", "") for r in df_perf.to_dict("records")
                                if r.get("boss")))
        print(f"Boss order: {len(BOSS_ORDER)} bosses (discovered from data)")


def _auto_short_name(name):
    """Generate a short display name from a boss name."""
    if "," in name:
        name = name.split(",")[0].strip()
    if name.startswith("The "):
        return name[4:]
    words = name.split()
    if len(words) <= 1:
        return name
    first = words[0]
    if "-" in first:
        return first
    return first


# Build short name lookup: check boss entry first, auto-generate fallback
_short_names = {}
for b in BOSS_ORDER:
    cfg = _bosses_cfg.get(b, {})
    _short_names[b] = cfg.get("short_name", _auto_short_name(b))


def boss_short(name):
    return _short_names.get(name, _auto_short_name(name))

# Raid display name(s) from config
if _raids_cfg:
    RAID_DISPLAY_NAME = " + ".join(_raids_cfg.keys())
else:
    # Fallback: read enabled names from config.json raids list
    _cfg_raids = _config.get("raids", [])
    _enabled = [r["name"] for r in _cfg_raids if r.get("enabled")]
    RAID_DISPLAY_NAME = " + ".join(_enabled) if _enabled else "Raid"

# Build spec/class lookup for role-aware + immune detection
_char_spec = {}
_char_class = {}
if len(df_pdt):
    for _, r in df_pdt.iterrows():
        n, s = r.get("player", "?"), r.get("player_spec", "")
        cl = r.get("player_class", "")
        if s and n not in _char_spec:
            _char_spec[n] = s
        if cl and n not in _char_class:
            _char_class[n] = cl

def _is_tank(player):
    return _char_spec.get(player, "") in TANK_SPECS

def _is_healer(player):
    return char_role_from_wcl.get(player, "") == "Healer"

def _is_immune_class(player):
    return _char_class.get(player, "") in IMMUNE_CLASSES

def _get_grade(score):
    if score >= GRADE_THRESHOLDS["A"]: return "A"
    if score >= GRADE_THRESHOLDS["B"]: return "B"
    if score >= GRADE_THRESHOLDS["C"]: return "C"
    if score >= GRADE_THRESHOLDS["D"]: return "D"
    return "F"

# ══════════════════════════════════════════════════════════════════
#  UNIFIED MECHANICS SCORING ENGINE (v4)
# ══════════════════════════════════════════════════════════════════

# ── Tank swap rules per boss — loaded from mechanic_rulesets.json ──
# method "binary": debuff with huge multiplier, must swap every cast.
#   Failures = co-tank applications beyond max_safe → blamed on THIS tank.
# method "ratio": continuously stacking debuff, compare distribution.
#   If ratio between tanks > threshold → the one with fewer stacks failed.
TANK_SWAP_RULES = {}
for _boss_name, _boss_data in mech_rulesets.get("bosses", {}).items():
    _tsr = _boss_data.get("tank_swap_rules", [])
    if _tsr:
        TANK_SWAP_RULES[_boss_name] = _tsr
if TANK_SWAP_RULES:
    print(f"  Tank swap rules loaded for {len(TANK_SWAP_RULES)} bosses: {', '.join(TANK_SWAP_RULES.keys())}")

# ── Build debuff application lookup from df_debuffs ──
# _debuff_apps[(player, boss, date, debuff_name)] = total applications
_debuff_apps = defaultdict(int)
if len(df_debuffs):
    for _, r in df_debuffs.iterrows():
        key = (r["player"], r["boss"], str(r["date"]), r["debuff_name"])
        _debuff_apps[key] += int(r.get("applications", 0))

# ── Build tanks-per-fight lookup ──
# _tanks_in_fight[(boss, date)] = set of tank player names
_tanks_in_fight = defaultdict(set)
if len(df_fight_roster):
    for _, r in df_fight_roster.iterrows():
        if _is_tank(r["player"]):
            _tanks_in_fight[(r["boss"], str(r["date"]))].add(r["player"])

# ── Fights attended per player per date ──
_fights_attended = defaultdict(lambda: defaultdict(int))
if len(df_fight_roster):
    for _, r in df_fight_roster.iterrows():
        _fights_attended[r["player"]][str(r["date"])] += 1

# ── Build per-player × boss × date ability damage lookup ──
# ability_dmg[(player, boss, date, ability)] = total damage
_ability_dmg = defaultdict(float)
if len(df_pdt):
    for _, r in df_pdt.iterrows():
        key = (r["player"], r["boss"], str(r["date"]), r["ability_name"])
        _ability_dmg[key] += float(r["ability_total"])

# ── Build per-fight target swap damage lookup ──
# target_dmg[(player, boss, date)] = total target swap damage
_target_dmg = defaultdict(float)
if len(df_mtd):
    for _, r in df_mtd.iterrows():
        key = (r["player"], r["boss"], str(r["date"]))
        _target_dmg[key] += float(r["damage_done"])

# ── Compute raid medians for relative_fail abilities ──
# raid_ability_dmg[(boss, date, ability)] = [dmg_per_player...]
_raid_ability_dmg = defaultdict(list)
if len(df_pdt):
    for _, r in df_pdt.iterrows():
        key = (r["boss"], str(r["date"]), r["ability_name"])
        _raid_ability_dmg[key].append(float(r["ability_total"]))

# ── Mechanic scoring function (DPS + Healers only) ──
def _score_mechanics_for(player, boss, date):
    """Score a player×boss×date on mechanics.
    Returns (score 0-100, list_of_events).
    Each event = {"ability", "method", "display", "fix", "result", "value"}
    All roles go through normal mechanic checks (role_filter handles skips).
    Tanks additionally get tank-swap checks appended.
    """
    boss_rules = mech_rulesets.get("bosses", {}).get(boss, {})
    mechanics = boss_rules.get("mechanics", {})
    target_swaps = boss_rules.get("target_swap", {})
    bonus_mechs = boss_rules.get("bonus_mechanics", {})

    events = []
    penalties = 0   # count of binary fails
    bonus_count = 0  # count of passes/bonuses
    total_checks = 0 # total scoreable checks

    for ability, info in mechanics.items():
        method = info["method"]
        role_filter = info.get("role_filter", "")
        display = info.get("display", ability)
        fix = info.get("fix", "")

        # Role filter: skip if doesn't apply
        if role_filter == "non_tank" and _is_tank(player):
            continue  # tank — this check doesn't apply
        if role_filter == "tank_only" and not _is_tank(player):
            continue

        dmg = _ability_dmg.get((player, boss, date, ability), 0)

        if method == "binary_fail":
            total_checks += 1
            if dmg > 0:
                penalties += 1
                events.append({"ability": ability, "method": method, "display": display,
                               "fix": fix, "result": "FAIL", "value": round(dmg)})
            else:
                events.append({"ability": ability, "method": method, "display": display,
                               "fix": fix, "result": "PASS", "value": 0})

        elif method == "relative_fail":
            total_checks += 1
            raid_vals = _raid_ability_dmg.get((boss, date, ability), [])
            if dmg > 0 and raid_vals:
                median_v = float(np.median(raid_vals))
                if dmg > median_v * 1.5:
                    penalties += 1
                    events.append({"ability": ability, "method": method, "display": display,
                                   "fix": fix, "result": "FAIL", "value": round(dmg)})
                else:
                    events.append({"ability": ability, "method": method, "display": display,
                                   "fix": fix, "result": "OK", "value": round(dmg)})
            elif dmg > 0:
                events.append({"ability": ability, "method": method, "display": display,
                               "fix": fix, "result": "OK", "value": round(dmg)})
            else:
                events.append({"ability": ability, "method": method, "display": display,
                               "fix": fix, "result": "PASS", "value": 0})

        elif method == "binary_pass":
            total_checks += 1
            if dmg > 0:
                bonus_count += 1
                events.append({"ability": ability, "method": method, "display": display,
                               "fix": fix, "result": "PASS", "value": round(dmg)})
            else:
                penalties += 1
                events.append({"ability": ability, "method": method, "display": display,
                               "fix": fix, "result": "FAIL", "value": 0})

        elif method == "immune_soak":
            total_checks += 1
            if dmg > 0:
                bonus_count += 1
                events.append({"ability": ability, "method": method, "display": display,
                               "fix": fix, "result": "PASS", "value": round(dmg)})
            elif _is_immune_class(player):
                bonus_count += 1
                events.append({"ability": ability, "method": method, "display": display,
                               "fix": fix, "result": "PASS (immune)", "value": 0})
            else:
                penalties += 1
                events.append({"ability": ability, "method": method, "display": display,
                               "fix": fix, "result": "FAIL", "value": 0})

        elif method == "conditional_fail":
            cond = info.get("condition", "")
            if cond == "non_tank_if_damage_exists":
                if _is_tank(player):
                    continue
                # Check if ANYONE in raid took this damage this boss/date
                raid_vals = _raid_ability_dmg.get((boss, date, ability), [])
                if raid_vals and sum(raid_vals) > 0:
                    total_checks += 1
                    # Everyone non-tank fails since nobody soaked pylons
                    penalties += 1
                    events.append({"ability": ability, "method": method, "display": display,
                                   "fix": fix, "result": "FAIL", "value": round(dmg)})

    # Target swaps — DPS only (tanks/healers not expected to swap)
    if not _is_tank(player) and not _is_healer(player):
        for add_name, info in target_swaps.items():
            display = info.get("display", add_name)
            fix = info.get("fix", "")
            td = _target_dmg.get((player, boss, date), 0)
            total_checks += 1
            if td > 0:
                bonus_count += 1
                events.append({"ability": add_name, "method": "target_swap", "display": display,
                               "fix": fix, "result": "PASS", "value": round(td)})
            else:
                penalties += 1
                events.append({"ability": add_name, "method": "target_swap", "display": display,
                               "fix": fix, "result": "FAIL", "value": 0})

    # Bonus mechanics (e.g. Fracture ghost pickup)
    for ability, info in bonus_mechs.items():
        role_filter = info.get("role_filter", "")
        if role_filter == "non_tank" and _is_tank(player):
            continue
        display = info.get("display", ability)
        fix = info.get("fix", "")
        dmg = _ability_dmg.get((player, boss, date, ability), 0)
        # Bonus: no penalty for not doing it, but reward for doing it
        if dmg > 0:
            bonus_count += 1
            total_checks += 1
            events.append({"ability": ability, "method": "bonus", "display": display,
                           "fix": fix, "result": "BONUS", "value": round(dmg)})

    # Tank swap checks (tanks only — appended to normal mechanic score)
    if _is_tank(player):
        ts_score, ts_events = _score_tank_mechanics_for(player, boss, date)
        if ts_events:
            # Merge tank swap events and counts into the main tally
            for evt in ts_events:
                events.append(evt)
                total_checks += 1
                if evt["result"] == "FAIL":
                    penalties += 1

    # Calculate score: start at 100, lose points per penalty, gain for bonuses
    if total_checks == 0:
        return None, events  # No mechanics to check = no score (excluded from composite)

    pass_count = total_checks - penalties  # includes bonuses
    score = round(100 * (pass_count + bonus_count) / (total_checks + bonus_count)) if (total_checks + bonus_count) > 0 else 100
    score = min(100, max(0, score))
    return score, events


# ── Tank scoring function ──
def _score_tank_mechanics_for(player, boss, date):
    """Score a tank player×boss×date based on swap discipline.
    Returns (score 0-100, list_of_events).
    Failures attributed to THIS tank = co-tank's extra debuff applications
    (i.e., co-tank got hit because THIS tank didn't taunt in time).
    """
    rules = TANK_SWAP_RULES.get(boss, [])
    if not rules:
        return None, []  # No tank swap rules for this boss

    # Find co-tank(s) for this fight
    fight_tanks = _tanks_in_fight.get((boss, date), set())
    co_tanks = fight_tanks - {player}
    if not co_tanks:
        return None, []  # Solo tank or can't identify partner — skip

    events = []
    penalties = 0
    total_checks = 0

    for rule in rules:
        debuff = rule["debuff"]
        method = rule["method"]
        display = rule.get("display", debuff)
        fix = rule.get("fix", "")

        my_apps = _debuff_apps.get((player, boss, date, debuff), 0)

        if method == "binary":
            max_safe = rule.get("max_safe", 1)
            total_checks += 1
            # Co-tank's extra applications = my failures to taunt
            co_extra = 0
            for ct in co_tanks:
                ct_apps = _debuff_apps.get((ct, boss, date, debuff), 0)
                co_extra += max(0, ct_apps - max_safe)

            if co_extra > 0:
                penalties += 1
                events.append({"ability": debuff, "method": "tank_swap_binary",
                               "display": display, "fix": fix, "result": "FAIL",
                               "value": co_extra,
                               "detail": f"Co-tank got {co_extra} extra application(s)"})
            else:
                events.append({"ability": debuff, "method": "tank_swap_binary",
                               "display": display, "fix": fix, "result": "PASS",
                               "value": 0})

        elif method == "stacking_binary":
            # For continuously stacking debuffs with a known swap threshold.
            # Check in multiples of max_safe: 12/6 = clean, 13/5 = failure.
            # If co-tank total % max_safe != 0 AND co-tank had >max_safe total
            # (at least one full cycle + overflow), then I failed to taunt.
            max_safe = rule.get("max_safe", 6)
            total_checks += 1
            co_total = sum(_debuff_apps.get((ct, boss, date, debuff), 0)
                           for ct in co_tanks)
            co_remainder = co_total % max_safe

            if co_remainder > 0 and co_total > max_safe:
                # Co-tank had at least one cycle that exceeded max_safe
                penalties += 1
                events.append({"ability": debuff, "method": "tank_swap_stacking",
                               "display": display, "fix": fix, "result": "FAIL",
                               "value": co_total,
                               "detail": f"Co-tank {co_total} apps (not clean multiple of {max_safe})"})
            else:
                events.append({"ability": debuff, "method": "tank_swap_stacking",
                               "display": display, "fix": fix, "result": "PASS",
                               "value": my_apps})

        elif method == "ratio":
            threshold = rule.get("ratio_threshold", 2.0)
            total_checks += 1
            # Compare this tank vs co-tank applications directly
            co_apps = sum(_debuff_apps.get((ct, boss, date, debuff), 0)
                          for ct in co_tanks)
            total = my_apps + co_apps

            if total == 0:
                # Neither tank has the debuff — pass
                events.append({"ability": debuff, "method": "tank_swap_ratio",
                               "display": display, "fix": fix, "result": "PASS",
                               "value": 0})
            elif my_apps == 0 or co_apps == 0:
                # One tank has all the stacks — the one with zero failed
                if my_apps == 0 and co_apps > 0:
                    # I never taunted → fail
                    penalties += 1
                    events.append({"ability": debuff, "method": "tank_swap_ratio",
                                   "display": display, "fix": fix, "result": "FAIL",
                                   "value": co_apps,
                                   "detail": f"Co-tank {co_apps} vs you 0 apps"})
                else:
                    events.append({"ability": debuff, "method": "tank_swap_ratio",
                                   "display": display, "fix": fix, "result": "PASS",
                                   "value": my_apps})
            else:
                imbalance = max(my_apps, co_apps) / min(my_apps, co_apps)
                i_have_fewer = my_apps < co_apps
                if i_have_fewer and imbalance > threshold:
                    # I have fewer stacks = I didn't taunt enough = FAIL
                    penalties += 1
                    events.append({"ability": debuff, "method": "tank_swap_ratio",
                                   "display": display, "fix": fix, "result": "FAIL",
                                   "value": co_apps,
                                   "detail": f"Co-tank {co_apps} vs you {my_apps} apps (ratio {imbalance:.1f}x)"})
                else:
                    events.append({"ability": debuff, "method": "tank_swap_ratio",
                                   "display": display, "fix": fix, "result": "PASS",
                                   "value": my_apps})

    if total_checks == 0:
        return None, events

    pass_count = total_checks - penalties
    score = round(100 * pass_count / total_checks)
    score = min(100, max(0, score))
    return score, events

# ── Compute mechanics scores for all player × boss × date ──
_mech_score_cache = {}     # (player, boss, date) → score
_mech_events_cache = {}    # (player, boss, date) → [events]
_mech_score_overall = defaultdict(lambda: defaultdict(list))     # player → date → [scores]
_mech_score_boss = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

# Players by date
_players_by_date = defaultdict(set)
if len(df_fight_roster):
    for _, r in df_fight_roster.iterrows():
        _players_by_date[str(r["date"])].add(r["player"])

for date in all_dates:
    date_players = _players_by_date.get(date, set())
    date_bosses_list = sorted(set(r["boss"] for r in fight_rows if str(r["date"]) == date))
    for boss in date_bosses_list:
        for player in date_players:
            score, events = _score_mechanics_for(player, boss, date)
            _mech_events_cache[(player, boss, date)] = events
            if score is not None:
                _mech_score_cache[(player, boss, date)] = score
                _mech_score_overall[player][date].append(score)
                _mech_score_boss[player][boss][date].append(score)

# ── 2. DEATHS — kill deaths count, wipe deaths only first 4 ──
_death_penalty = defaultdict(lambda: defaultdict(float))
_death_penalty_boss = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
_death_count = defaultdict(lambda: defaultdict(int))
_death_count_boss = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
if len(df_deaths):
    for _, d in df_deaths.iterrows():
        player, date, boss = d["player"], str(d["date"]), d["boss"]
        kill = bool(d.get("kill", False))
        order = int(d.get("death_order", 99))
        # Count if: kill fight (any death counts) OR wipe + first 4 dead
        if not kill and order > 4:
            continue
        _death_penalty[player][date] += 1.0
        _death_penalty_boss[player][boss][date] += 1.0
        _death_count[player][date] += 1
        _death_count_boss[player][boss][date] += 1

def _death_score(player, date):
    n = _fights_attended[player].get(date, 1)
    blame_deaths = _death_penalty.get(player, {}).get(date, 0)
    rate = blame_deaths / n
    return max(0, round(100 * (1 - rate / 0.5)))

def _death_score_boss(player, boss, date):
    n = max(_boss_fight_count.get((player, boss, date), 0), 1)
    blame_deaths = _death_penalty_boss.get(player, {}).get(boss, {}).get(date, 0)
    rate = blame_deaths / n
    return max(0, round(100 * (1 - rate / 0.5)))

# Pre-build fight count per (player, boss, date) from fight_roster
_boss_fight_count = defaultdict(int)
if len(df_fight_roster):
    for _, fr in df_fight_roster.iterrows():
        _boss_fight_count[(fr["player"], fr["boss"], str(fr["date"]))] += 1

# ── 3. PARSE PERFORMANCE ──
_parse_scores = defaultdict(lambda: defaultdict(list))
_parse_scores_boss = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
for _, r in df_perf.iterrows():
    role = str(r.get("role", "")).strip()
    rl = {"tank":"Tank","healer":"Healer","dps":"DPS","dp":"DPS",
          "Tank":"Tank","Healer":"Healer","DPS":"DPS"}.get(role, "")
    m = r.get("metric", "")
    use = (m == "dps" and rl != "Healer") or (m == "hps" and rl == "Healer")
    if not use:
        continue
    rp = r.get("rank_percent")
    if pd.notna(rp) and rp:
        _parse_scores[r["player"]][str(r["date"])].append(float(rp))
        _parse_scores_boss[r["player"]][r["boss"]][str(r["date"])].append(float(rp))

# ── 3b. ESTIMATE WIPE PARSES via DPS→parse% interpolation from kill data ──
# PARSE SCORING PHILOSOPHY:
#   Kill parses come from WCL and are authoritative — always used.
#   Wipe parses are inherently deflated (early deaths, no burn phase, ramp-up issues).
#   Including all wipes destroys parse averages (5:1 wipe:kill ratio → avg parse ~25).
#   RULE: Wipe parses are ONLY used on PROG NIGHTS — dates where a boss was never killed.
#         In that case, only the BEST pull (lowest boss HP %) contributes a parse estimate.
#         This ensures prog nights aren't scoreless, without polluting kill-night averages.
from itertools import groupby as _groupby

def _build_parse_curve(points):
    """Given [(amount, parse%)], return (amounts_arr, parses_arr) for np.interp."""
    if len(points) < 1:
        return None
    sorted_pts = sorted(points, key=lambda x: x[0])
    # Dedupe: average parses at same DPS
    deduped = []
    for amt, grp in _groupby(sorted_pts, key=lambda x: x[0]):
        grp_list = list(grp)
        deduped.append((amt, sum(p for _, p in grp_list) / len(grp_list)))
    if len(deduped) == 1:
        # Single point: extrapolate linearly (0 DPS = 0%, scale proportionally)
        amt, pct = deduped[0]
        if amt > 0 and pct > 0:
            # Estimate DPS at 100%: amt / (pct/100)
            dps_100 = amt / (pct / 100.0)
            return (np.array([0.0, amt, dps_100]), np.array([0.0, pct, 100.0]))
        return None
    amounts = np.array([a for a, _ in deduped])
    parses = np.array([p for _, p in deduped])
    # Enforce monotonic non-decreasing (higher DPS = higher parse)
    for i in range(1, len(parses)):
        parses[i] = max(parses[i], parses[i-1])
    return (amounts, parses)

# Collect raw data points by (boss, spec, metric) and (boss, role, metric)
_spec_curve_data = defaultdict(list)
_role_curve_data = defaultdict(list)
for _, r in df_perf.iterrows():
    role = str(r.get("role", "")).strip()
    rl = {"tank":"Tank","healer":"Healer","dps":"DPS","dp":"DPS",
          "Tank":"Tank","Healer":"Healer","DPS":"DPS"}.get(role, "")
    if not rl:
        continue
    m = r.get("metric", "")
    use = (m == "dps" and rl != "Healer") or (m == "hps" and rl == "Healer")
    if not use:
        continue
    rp = r.get("rank_percent")
    amt = r.get("amount", 0)
    if pd.notna(rp) and rp and amt and amt > 0:
        spec = str(r.get("player_spec", "")).strip()
        if spec:
            _spec_curve_data[(r["boss"], spec, m)].append((float(amt), float(rp)))
        role_group = "Tank" if rl == "Tank" else ("Healer" if rl == "Healer" else "DPS")
        _role_curve_data[(r["boss"], role_group, m)].append((float(amt), float(rp)))

# Build curves
_spec_curves = {}
for key, pts in _spec_curve_data.items():
    curve = _build_parse_curve(pts)
    if curve is not None:
        _spec_curves[key] = curve

_role_curves = {}
for key, pts in _role_curve_data.items():
    if len(pts) >= 3:
        curve = _build_parse_curve(pts)
        if curve is not None:
            _role_curves[key] = curve

# Build boss_pct lookup: (report_code, fight_id) → boss_pct
_fight_boss_pct = {}
if len(df_fights):
    for _, f in df_fights.iterrows():
        _fight_boss_pct[(str(f["report_code"]), int(f["fight_id"]))] = f.get("boss_pct", 100.0)

# Identify prog nights: (boss, date) combos with NO kills
_has_kill = set()
if len(df_fights):
    for _, f in df_fights.iterrows():
        if f.get("kill"):
            _has_kill.add((f["boss"], str(f["date"])))

# Estimate wipe parses → store in temp dict, NOT in _parse_scores yet
# Key: (player, boss, date) → [(est_parse, boss_pct, report_code, fight_id)]
_wipe_estimates = defaultdict(list)
_wipe_parse_count = 0
_wipe_parse_spec_hits = 0
_wipe_parse_role_hits = 0
for _, r in df_perf.iterrows():
    if r.get("kill"):
        continue
    rp = r.get("rank_percent")
    if pd.notna(rp) and rp:
        continue
    amt = r.get("amount", 0)
    if not amt or amt <= 0:
        continue
    player = r["player"]
    boss = r["boss"]
    date_str = str(r["date"])
    m = r.get("metric", "")
    spec = str(r.get("player_spec", "")).strip()
    # Determine role from WCL role data
    rl = char_role_from_wcl.get(player, "")
    if rl == "Healer" and m != "hps":
        continue
    if rl != "Healer" and m != "dps":
        continue
    # Try spec-level curve first
    curve = _spec_curves.get((boss, spec, m))
    if curve:
        _wipe_parse_spec_hits += 1
    else:
        # Fallback to role-level
        role_group = "Tank" if rl == "Tank" else ("Healer" if rl == "Healer" else "DPS")
        curve = _role_curves.get((boss, role_group, m))
        if curve:
            _wipe_parse_role_hits += 1
    if not curve:
        continue
    amounts, parses = curve
    est_parse = float(np.clip(np.interp(float(amt), amounts, parses), 1, 99))
    est_parse = round(est_parse)
    boss_pct = _fight_boss_pct.get((str(r["report_code"]), int(r["fight_id"])), 100.0)
    _wipe_estimates[(player, boss, date_str)].append(
        (est_parse, boss_pct, str(r["report_code"]), int(r["fight_id"])))
    _wipe_parse_count += 1

# Apply prog-night rule: only add the BEST wipe pull (lowest boss_pct) for nights with no kills
_prog_parses_added = 0
_prog_best_pulls = set()  # (report_code, fight_id) of best pulls — for patching ranking_rows
_prog_parse_map = {}      # (player, boss, date) → est_parse
for (player, boss, date_str), estimates in _wipe_estimates.items():
    if (boss, date_str) in _has_kill:
        continue  # This boss was killed on this date — skip all wipe parses
    # Prog night: pick the estimate from the pull with lowest boss_pct (closest to kill)
    best_est, best_pct, best_code, best_fid = min(estimates, key=lambda x: x[1])
    _parse_scores[player][date_str].append(best_est)
    _parse_scores_boss[player][boss][date_str].append(best_est)
    _prog_best_pulls.add((best_code, best_fid))
    _prog_parse_map[(player, boss, date_str)] = best_est
    _prog_parses_added += 1

# Patch ranking_rows: inject estimated parses for best-pull wipe fights on prog nights
# This ensures Raid Performance player summary + detail panels show data for prog nights.
_patched = 0
for r in ranking_rows:
    if r["parse_pct"]:
        continue  # already has a real parse
    key = (r["character"], r["boss"], r["date"])
    est = _prog_parse_map.get(key)
    if est is None:
        continue
    if (r["report_code"], r["fight_id"]) not in _prog_best_pulls:
        continue  # not the best pull — don't patch
    # Only patch the relevant metric row (dps for non-healers, hps for healers)
    rl = char_role_from_wcl.get(r["character"], "")
    if (rl == "Healer" and r["metric"] == "hps") or (rl != "Healer" and r["metric"] == "dps"):
        r["parse_pct"] = est
        _patched += 1

print(f"Parse curves: {len(_spec_curves)} spec-level, {len(_role_curves)} role-level")
print(f"Wipe parses estimated: {_wipe_parse_count} (spec={_wipe_parse_spec_hits}, role_fallback={_wipe_parse_role_hits})")
print(f"  Prog-night best-pull parses added: {_prog_parses_added}")
print(f"  Kill-night wipe parses discarded: {_wipe_parse_count - _prog_parses_added}")
print(f"  Ranking rows patched for RP visibility: {_patched}")

# ── 4. CONSUMABLES (split: tempered pots vs health pots/stones) ──
_con_scores = defaultdict(lambda: defaultdict(list))
_con_scores_boss = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

# Build per-player-per-fight lookups by category
_tempered_lookup = defaultdict(int)
_health_lookup = defaultdict(int)
if len(df_consumables):
    for _, c in df_consumables.iterrows():
        key = (c["player"], int(c["fight_id"]), str(c["date"]))
        cat = c.get("category", "")
        casts = int(c.get("casts", c.get("pot_count", 0)))
        if cat in ("tempered_potion", "tempered"):
            _tempered_lookup[key] += casts
        elif cat in ("healing_potion", "healthstone", "health"):
            _health_lookup[key] += casts
        else:
            # Legacy data without category — count as tempered (old behavior)
            _tempered_lookup[key] += casts

# Build per-fight "anyone used tempered?" lookup
_anyone_tempered = defaultdict(bool)
for k, v in _tempered_lookup.items():
    if v > 0:
        _anyone_tempered[(k[1], k[2])] = True

# Build per-player-per-fight death lookup
_died_in_fight = defaultdict(bool)
if len(df_deaths):
    for _, d in df_deaths.iterrows():
        key = (d["player"], int(d["fight_id"]), str(d["date"]))
        _died_in_fight[key] = True

# Build killing blow classifier: was the death a DOT (survivable) or one-shot?
ONE_SHOT_DMG_THRESHOLD = 100000
ONE_SHOT_OVERKILL_THRESHOLD = 50000

_dot_death = {}  # (player, fight_id, date) → True if ANY death was DOT-type
if len(df_death_events):
    dmg_events = df_death_events[df_death_events["event_type"] == "damage"].copy()
    if len(dmg_events):
        for (player, fid, date_str, dorder), grp in dmg_events.groupby(
                ["player", "fight_id", "date", "death_order"]):
            last_hit = grp.loc[grp["timestamp_ms"].idxmax()]
            amount = abs(float(last_hit.get("amount", 0)))
            overkill = abs(float(last_hit.get("overkill", 0)))
            is_one_shot = amount > ONE_SHOT_DMG_THRESHOLD or overkill > ONE_SHOT_OVERKILL_THRESHOLD
            key = (player, int(fid), str(date_str))
            if not is_one_shot:
                _dot_death[key] = True
            elif key not in _dot_death:
                _dot_death[key] = False

# Build kill lookup: (fight_id, date_str) → bool
_fight_kill = {}
for _, f in df_fights.iterrows():
    _fight_kill[(int(f["fight_id"]), str(f["date"]))] = bool(f["kill"])

# Score every player-fight combo from fight_roster
if len(df_fight_roster):
    for _, r in df_fight_roster.iterrows():
        player = r["player"]
        fid = int(r["fight_id"])
        date_str = str(r["date"])
        boss = r["boss"]
        is_kill = _fight_kill.get((fid, date_str), bool(r["kill"]))
        key = (player, fid, date_str)

        tempered = _tempered_lookup.get(key, 0)
        health = _health_lookup.get(key, 0)
        anyone_t = _anyone_tempered.get((fid, date_str), False)
        died = _died_in_fight.get(key, False)

        # -- Tempered pot score (DPS/throughput pot) --
        if is_kill:
            temp_score = 100 if tempered >= 1 else 0
        else:
            if not anyone_t:
                temp_score = 100  # Nobody potted on wipe — strategic hold
            elif tempered >= 1:
                temp_score = 100
            else:
                temp_score = 0  # Others potted but you didn't

        # -- Health pot score (survival) --
        # Only penalize if death was a DOT (survivable) — one-shots aren't pottable
        if died and health == 0:
            was_dot_death = _dot_death.get(key, False)
            if was_dot_death:
                hp_score = 0   # Died to DOT without using a health pot/healthstone
            else:
                hp_score = 100  # One-shot — can't blame for not potting
        else:
            hp_score = 100  # Didn't die, or died but at least used one

        # Combined: average of both checks
        score = round((temp_score + hp_score) / 2)

        _con_scores[player][date_str].append(score)
        _con_scores_boss[player][boss][date_str].append(score)

# ── Build score_rows (per player × date) ──
score_rows = []
for date in all_dates:
    date_players = _players_by_date.get(date, set())

    for player in date_players:
        mapped_player = char_to_player.get(player, player)  # char→player for output key
        mech_vals = _mech_score_overall.get(player, {}).get(date, [])
        mech_avg = round(np.mean(mech_vals)) if mech_vals else None

        deaths = _death_score(player, date)

        parses = _parse_scores.get(player, {}).get(date, [])
        parse_avg = round(np.mean(parses)) if parses else None

        cons = _con_scores.get(player, {}).get(date, [])
        con_avg = round(np.mean(cons)) if cons else None

        components = {"deaths": deaths}
        if mech_avg is not None:
            components["mechanics"] = mech_avg
        if parse_avg is not None:
            components["parse_performance"] = parse_avg
        if con_avg is not None:
            components["consumables"] = con_avg

        aw = sum(SCORE_WEIGHTS[k] for k in components)
        composite = round(sum(components[k] * (SCORE_WEIGHTS[k] / aw) for k in components)) if aw > 0 else 50
        composite = min(100, max(0, composite))

        n_fights = _fights_attended[player].get(date, 0)
        raw_deaths = _death_count.get(player, {}).get(date, 0)

        # Count fails/passes for this player on this date
        total_fails = 0
        total_passes = 0
        for boss in set(r["boss"] for r in fight_rows if str(r["date"]) == date):
            evts = _mech_events_cache.get((player, boss, date), [])
            total_fails += sum(1 for e in evts if e["result"] == "FAIL")
            total_passes += sum(1 for e in evts if e["result"] in ("PASS", "PASS (immune)", "BONUS"))

        score_rows.append({
            "player": char_to_player.get(player, player), "date": date,
            "role": char_role_from_wcl.get(player, ""),
            "mech_score": mech_avg, "death_score": deaths,
            "parse_score": parse_avg, "con_score": con_avg,
            "composite": composite, "grade": _get_grade(composite),
            "deaths_raw": raw_deaths, "mech_fails": total_fails, "mech_passes": total_passes,
            "fights": n_fights,
        })

# ── Build boss_score_rows (per player × boss × date) ──
boss_score_rows = []
# Only score player×boss×date combos where the player actually fought that boss
_players_by_boss_date = defaultdict(set)  # (boss, date) → set of players
if len(df_fight_roster):
    for _, fr in df_fight_roster.iterrows():
        _players_by_boss_date[(fr["boss"], str(fr["date"]))].add(fr["player"])

for date in all_dates:
    date_bosses_set = sorted(set(r["boss"] for r in fight_rows if str(r["date"]) == date))

    for boss in date_bosses_set:
        for player in _players_by_boss_date.get((boss, date), set()):
            mapped_player = char_to_player.get(player, player)  # char→player for output key
            mech_s = _mech_score_cache.get((player, boss, date), None)
            deaths = _death_score_boss(player, boss, date)

            parses = _parse_scores_boss.get(player, {}).get(boss, {}).get(date, [])
            parse_avg = round(np.mean(parses)) if parses else None

            cons = _con_scores_boss.get(player, {}).get(boss, {}).get(date, [])
            con_avg = round(np.mean(cons)) if cons else None

            evts = _mech_events_cache.get((player, boss, date), [])
            fails = sum(1 for e in evts if e["result"] == "FAIL")
            passes = sum(1 for e in evts if e["result"] in ("PASS", "PASS (immune)", "BONUS"))

            # Compute composite inline
            bcomps = {"deaths": deaths}
            if mech_s is not None: bcomps["mechanics"] = mech_s
            if parse_avg is not None: bcomps["parse_performance"] = parse_avg
            if con_avg is not None: bcomps["consumables"] = con_avg
            baw = sum(SCORE_WEIGHTS[k] for k in bcomps)
            bcomp = round(sum(bcomps[k] * (SCORE_WEIGHTS[k] / baw) for k in bcomps)) if baw > 0 else 50
            bcomp = min(100, max(0, bcomp))

            boss_score_rows.append({
                "player": char_to_player.get(player, player), "date": date, "boss": boss,
                "mech_score": mech_s, "death_score": deaths,
                "parse_score": parse_avg, "con_score": con_avg,
                "deaths_raw": _death_count_boss.get(player, {}).get(boss, {}).get(date, 0),
                "mech_fails": fails, "mech_passes": passes,
                "composite": bcomp, "grade": _get_grade(bcomp),
            })

print(f"Scores: {len(score_rows)} player×date rows, {len(boss_score_rows)} player×boss×date rows")
if score_rows:
    from collections import Counter
    gc = Counter(r["grade"] for r in score_rows)
    print(f"  Grade distribution: {dict(sorted(gc.items()))}")

# ── Score lookups: per player×boss best/avg composite, raid averages ──
# best_score_by_player_boss[player][boss] = highest composite across all dates
# avg_score_by_player_boss[player][boss] = mean composite across all dates
best_score_by_player_boss = defaultdict(lambda: defaultdict(int))
avg_score_by_player_boss = defaultdict(lambda: defaultdict(float))
_bsr_bucket = defaultdict(list)  # (player, boss) → [composites]
for bsr in boss_score_rows:
    _bsr_bucket[(bsr["player"], bsr["boss"])].append(bsr["composite"])
for (player, boss), comps in _bsr_bucket.items():
    best_score_by_player_boss[player][boss] = max(comps)
    avg_score_by_player_boss[player][boss] = round(sum(comps) / len(comps), 1)

# Map char→player for character-level lookups
_char_player = {}
for rr in roster_rows:
    _char_player[rr["character"]] = rr["player"]

# best/avg score by character (via player mapping)
best_score_by_char_boss = defaultdict(lambda: defaultdict(int))
avg_score_by_char_boss = defaultdict(lambda: defaultdict(float))
for char, player in _char_player.items():
    for boss in BOSS_ORDER:
        best_score_by_char_boss[char][boss] = best_score_by_player_boss[player].get(boss, 0)
        avg_score_by_char_boss[char][boss] = avg_score_by_player_boss[player].get(boss, 0)

# Raid averages (mains only; when locked, only rostered players count)
_main_players = set()
for rr in roster_rows:
    if rr["main_alt"] == "Main":
        player = rr["player"]
        if roster_locked and player not in rostered_players:
            continue
        _main_players.add(player)

if roster_locked:
    print(f"  Roster LOCKED: {len(_main_players)} rostered mains count toward averages")

raid_avg_best_score_boss = {}
raid_avg_avg_score_boss = {}
for boss in BOSS_ORDER:
    bests = [best_score_by_player_boss[p].get(boss, 0) for p in _main_players
             if best_score_by_player_boss[p].get(boss, 0) > 0]
    avgs = [avg_score_by_player_boss[p].get(boss, 0) for p in _main_players
            if avg_score_by_player_boss[p].get(boss, 0) > 0]
    raid_avg_best_score_boss[boss] = round(sum(bests) / len(bests), 1) if bests else 0
    raid_avg_avg_score_boss[boss] = round(sum(avgs) / len(avgs), 1) if avgs else 0

# Raid average composite (latest date, mains)
if all_dates:
    latest_date = sorted(all_dates)[-1]
    _main_latest_comps = [s["composite"] for s in score_rows
                          if s["player"] in _main_players and s["date"] == latest_date
                          and s["composite"] is not None]
    raid_avg_composite = round(sum(_main_latest_comps) / len(_main_latest_comps), 1) if _main_latest_comps else 50
else:
    raid_avg_composite = 50

# ══════════════════════════════════════════════════════════════════
#  STYLES
# ══════════════════════════════════════════════════════════════════

FN = "Aptos Narrow"
F_TITLE = Font(name=FN, size=16, bold=True, color="FFFFFF")
F_SECTION = Font(name=FN, size=11, bold=True, color="FFFFFF")
F_HDR = Font(name=FN, size=10, bold=True, color="FFFFFF")
F_BODY = Font(name=FN, size=10)
F_BOLD = Font(name=FN, size=10, bold=True)
F_SMALL = Font(name=FN, size=9, color="666666")
F_GREEN_BOLD = Font(name=FN, size=10, bold=True, color="16A34A")
F_RED_BOLD = Font(name=FN, size=10, bold=True, color="DC2626")

def add_arrow_formatting(ws, cell_range):
    """Apply conditional formatting: ▲ = green, ▼ = red."""
    first_cell = cell_range.split(":")[0]
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("▲",{first_cell}))'],
        font=Font(name=FN, color="16A34A", bold=True)))
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("▼",{first_cell}))'],
        font=Font(name=FN, color="DC2626", bold=True)))

def add_grade_coloring(ws, cell_range):
    """Conditional formatting: color grade letters A=green, B=blue, C=amber, D=orange, F=red."""
    first_cell = cell_range.split(":")[0]
    grade_styles = [
        ("A", "2E7D32", "E8F5E9"),  # green text, light green bg
        ("B", "1565C0", "E3F2FD"),  # blue text, light blue bg
        ("C", "F57F17", "FFF8E1"),  # amber text, light amber bg
        ("D", "E65100", "FFF3E0"),  # orange text, light orange bg
        ("F", "C62828", "FFEBEE"),  # red text, light red bg
    ]
    for letter, fg, bg in grade_styles:
        ws.conditional_formatting.add(cell_range, FormulaRule(
            formula=[f'EXACT({first_cell},"{letter}")'],
            font=Font(name=FN, bold=True, color=fg),
            fill=PatternFill("solid", fgColor=bg)))

def add_score_coloring(ws, cell_range):
    """Conditional formatting: color numeric scores by grade thresholds."""
    first_cell = cell_range.split(":")[0]
    # Order matters: most specific first (highest threshold first for stopIfTrue)
    score_styles = [
        (90, "2E7D32", "E8F5E9"),  # A
        (80, "1565C0", "E3F2FD"),  # B
        (70, "F57F17", "FFF8E1"),  # C
        (60, "E65100", "FFF3E0"),  # D
    ]
    for threshold, fg, bg in score_styles:
        ws.conditional_formatting.add(cell_range, FormulaRule(
            formula=[f'AND(ISNUMBER({first_cell}),{first_cell}>={threshold})'],
            font=Font(name=FN, bold=True, color=fg),
            fill=PatternFill("solid", fgColor=bg),
            stopIfTrue=True))
    # F = below 60
    ws.conditional_formatting.add(cell_range, FormulaRule(
        formula=[f'AND(ISNUMBER({first_cell}),{first_cell}<60)'],
        font=Font(name=FN, bold=True, color="C62828"),
        fill=PatternFill("solid", fgColor="FFEBEE"),
        stopIfTrue=True))

# ── WoW class colors (pastel fills for row highlighting) ──
CLASS_FILL = {
    "DeathKnight":  PatternFill("solid", fgColor="F5D0D7"),
    "DemonHunter":  PatternFill("solid", fgColor="E5C6F0"),
    "Druid":        PatternFill("solid", fgColor="FFE0BF"),
    "Evoker":       PatternFill("solid", fgColor="C6E5DD"),
    "Hunter":       PatternFill("solid", fgColor="E5F2D0"),
    "Mage":         PatternFill("solid", fgColor="D2F0FA"),
    "Monk":         PatternFill("solid", fgColor="CCFFE5"),
    "Paladin":      PatternFill("solid", fgColor="FCE0EE"),
    "Priest":       PatternFill("solid", fgColor="E8E8F0"),
    "Rogue":        PatternFill("solid", fgColor="FFF9CC"),
    "Shaman":       PatternFill("solid", fgColor="CCE0F7"),
    "Warlock":      PatternFill("solid", fgColor="DCDDFA"),
    "Warrior":      PatternFill("solid", fgColor="F0E0D0"),
}

def apply_class_fill(ws, row, num_cols, class_name, start_col=1):
    """Apply class-colored fill to columns in a row."""
    fill = CLASS_FILL.get(class_name)
    if fill:
        for col in range(start_col, start_col + num_cols):
            ws.cell(row=row, column=col).fill = fill

def interpolate_series(values):
    """Linearly interpolate None gaps between known values.
    Leading/trailing Nones stay None (empty in chart)."""
    out = list(values)
    n = len(out)
    for i in range(n):
        if out[i] is not None:
            continue
        prev_i = next((j for j in range(i - 1, -1, -1) if out[j] is not None), None)
        next_i = next((j for j in range(i + 1, n) if out[j] is not None), None)
        if prev_i is not None and next_i is not None:
            frac = (i - prev_i) / (next_i - prev_i)
            out[i] = round(out[prev_i] + frac * (out[next_i] - out[prev_i]), 1)
    return out

X_DARK = PatternFill("solid", fgColor="1F2937")
X_HDR = PatternFill("solid", fgColor="374151")
X_SEC = PatternFill("solid", fgColor="4B5563")
X_EVEN = PatternFill("solid", fgColor="F9FAFB")
X_ODD = PatternFill("solid", fgColor="FFFFFF")
X_GREEN = PatternFill("solid", fgColor="D1FAE5")
X_YELLOW = PatternFill("solid", fgColor="FEF3C7")
X_RED = PatternFill("solid", fgColor="FEE2E2")
X_ACCENT = PatternFill("solid", fgColor="2563EB")

AC = Alignment(horizontal="center", vertical="center")
AL = Alignment(horizontal="left", vertical="center")
AR = Alignment(horizontal="right", vertical="center")
BT = Border(bottom=Side(style="thin", color="E5E7EB"))

def hdr(ws, row, headers, c0=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=c0+i, value=h)
        c.font, c.fill, c.alignment = F_HDR, X_HDR, AC

def drow(ws, row, vals, c0=1, even=False):
    fill = X_EVEN if even else X_ODD
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=c0+i, value=v)
        c.font, c.fill, c.border = F_BODY, fill, BT

def title(ws, row, text, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font, c.fill, c.alignment = F_TITLE, X_DARK, AL
    for col in range(1, span+1): ws.cell(row=row, column=col).fill = X_DARK

def section(ws, row, text, span=10, c0=1):
    ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c0+span-1)
    c = ws.cell(row=row, column=c0, value=text)
    c.font, c.fill, c.alignment = F_SECTION, X_SEC, AL
    for col in range(c0, c0+span): ws.cell(row=row, column=col).fill = X_SEC

def widths(ws, w):
    for i, v in enumerate(w, 1): ws.column_dimensions[get_column_letter(i)].width = v

wb = Workbook()

# ══════════════════════════════════════════════════════════════════
#  DATA SHEETS (hidden)
# ══════════════════════════════════════════════════════════════════
ws = wb.active; ws.title = "d_roster"
hdr(ws, 1, ["Character","Player","Main/Alt","Class","Spec","Role","Server",
            "iLvl","Raid Prog","Miss Ench","Empty Sock","Tier","Lowest iLvl","Lowest Slot"])
for i, r in enumerate(roster_rows):
    drow(ws, i+2, [r["character"],r["player"],r["main_alt"],r["class"],r["spec"],
                    r["role"],r["server"],r["ilvl"],r["raid_prog"],
                    r["missing_ench"],r["empty_sockets"],r["tier_count"],
                    r["lowest_ilvl"],r["lowest_slot"]], even=i%2==0)

ws = wb.create_sheet("d_fights")
hdr(ws, 1, ["Report","Date","Fight ID","Boss","Kill","Duration(s)","Duration","Size","Deaths","Avg iLvl"])
for i, r in enumerate(fight_rows):
    drow(ws, i+2, [r["report_code"],r["date"],r["fight_id"],r["boss"],r["kill"],
                    r["duration_s"],r["duration_fmt"],r["size"],r["deaths_total"],r["avg_ilvl"]], even=i%2==0)

ws = wb.create_sheet("d_rankings")
hdr(ws, 1, ["Report","Date","Fight ID","Boss","Kill","Metric","Character","Player",
            "Main/Alt","Class","Spec","Role","Parse%","Bracket%","Amount","WCL ID"])
for i, r in enumerate(ranking_rows):
    drow(ws, i+2, [r["report_code"],r["date"],r["fight_id"],r["boss"],r["kill"],
                    r["metric"],r["character"],r["player"],r["main_alt"],r["class"],
                    r["spec"],r["role"],r["parse_pct"],r["bracket_pct"],r["amount"],
                    r["wcl_char_id"]], even=i%2==0)

ws = wb.create_sheet("d_deaths")
hdr(ws, 1, ["Report","Date","Fight ID","Boss","Kill","Character","Player",
            "Main/Alt","Killing Blow","Overkill","Death Window(ms)","Timestamp(ms)"])
for i, r in enumerate(death_rows):
    drow(ws, i+2, [r["report_code"],r["date"],r["fight_id"],r["boss"],r["kill"],
                    r["character"],r["player"],r["main_alt"],r["killing_blow"],
                    r["overkill"],r["death_window_ms"],r["timestamp_ms"]], even=i%2==0)

ws = wb.create_sheet("d_gear")
hdr(ws, 1, ["Character","Player","Slot","Item","iLvl","Quality","Enchant",
            "Needs Ench","Gems","Empty Sock","Tier"])
for i, r in enumerate(gear_rows):
    drow(ws, i+2, [r["character"],r["player"],r["slot"],r["item_name"],r["ilvl"],
                    r["quality"],r["enchant"],r["needs_enchant"],r["gems"],
                    r["empty_sockets"],r["is_tier"]], even=i%2==0)

print("Data sheets done.")

# ── d_scores (hidden: per player × date, keyed for XLOOKUP) ──
alltime_scores = {}
for player in sorted(set(r["player"] for r in score_rows)):
    prows = [r for r in score_rows if r["player"] == player]
    if not prows:
        continue
    alltime_scores[player] = {
        "player": player, "date": "All Time",
        "role": prows[0]["role"],
        "mech_score": round(np.mean([r["mech_score"] for r in prows if r["mech_score"] is not None])) if any(r["mech_score"] is not None for r in prows) else None,
        "death_score": round(np.mean([r["death_score"] for r in prows])),
        "parse_score": round(np.mean([r["parse_score"] for r in prows if r["parse_score"] is not None])) if any(r["parse_score"] is not None for r in prows) else None,
        "con_score": round(np.mean([r["con_score"] for r in prows if r["con_score"] is not None])) if any(r["con_score"] is not None for r in prows) else None,
        "composite": round(np.mean([r["composite"] for r in prows])),
        "grade": _get_grade(round(np.mean([r["composite"] for r in prows]))),
        "deaths_raw": sum(r["deaths_raw"] for r in prows),
        "mech_fails": sum(r["mech_fails"] for r in prows),
        "mech_passes": sum(r["mech_passes"] for r in prows),
        "fights": sum(r["fights"] for r in prows),
    }

all_score_rows = score_rows + list(alltime_scores.values())

# d_scores columns:
# A=Key, B=Player, C=Date, D=Role, E=Mechanics, F=Death, G=Parse, H=Consume,
# I=Composite, J=Grade, K=Deaths Raw, L=Mech Fails, M=Mech Passes, N=Fights
ws = wb.create_sheet("d_scores")
hdr(ws, 1, ["Key","Player","Date","Role","Mechanics","Death","Parse","Consume",
            "Composite","Grade","Deaths Raw","Mech Fails","Mech Passes","Fights"])
for i, r in enumerate(all_score_rows):
    key = r["player"] + "|" + r["date"]
    drow(ws, i+2, [key, r["player"],r["date"],r["role"],
                    r["mech_score"],r["death_score"],
                    r["parse_score"],r["con_score"],
                    r["composite"],r["grade"],
                    r["deaths_raw"],r["mech_fails"],r["mech_passes"],
                    r["fights"]], even=i%2==0)

# ── d_boss_scores (hidden: per player × boss × date, keyed for XLOOKUP) ──
alltime_boss_scores = []
for player in sorted(set(r["player"] for r in boss_score_rows)):
    for boss in BOSS_ORDER:
        brows = [r for r in boss_score_rows if r["player"] == player and r["boss"] == boss]
        if not brows:
            continue
        def _avg_or_none(vals):
            filt = [v for v in vals if v is not None]
            return round(np.mean(filt)) if filt else None

        at_mech = _avg_or_none([r["mech_score"] for r in brows])
        at_death = round(np.mean([r["death_score"] for r in brows]))
        at_parse = _avg_or_none([r["parse_score"] for r in brows])
        at_con = _avg_or_none([r["con_score"] for r in brows])

        comps = {"deaths": at_death}
        if at_mech is not None: comps["mechanics"] = at_mech
        if at_parse is not None: comps["parse_performance"] = at_parse
        if at_con is not None: comps["consumables"] = at_con
        aw = sum(SCORE_WEIGHTS[k] for k in comps)
        bc = round(sum(comps[k] * (SCORE_WEIGHTS[k] / aw) for k in comps)) if aw > 0 else 50
        bc = min(100, max(0, bc))

        alltime_boss_scores.append({
            "player": player, "date": "All Time", "boss": boss,
            "mech_score": at_mech, "death_score": at_death,
            "parse_score": at_parse, "con_score": at_con,
            "deaths_raw": sum(r["deaths_raw"] for r in brows),
            "mech_fails": sum(r["mech_fails"] for r in brows),
            "composite": bc, "grade": _get_grade(bc),
        })

all_boss_score_rows = boss_score_rows + alltime_boss_scores

# d_boss_scores columns:
# A=Key, B=Player, C=Date, D=Boss, E=Mechanics, F=Death, G=Parse, H=Consume,
# I=Deaths Raw, J=Mech Fails, K=Composite, L=Grade
ws = wb.create_sheet("d_boss_scores")
hdr(ws, 1, ["Key","Player","Date","Boss","Mechanics","Death","Parse","Consume",
            "Deaths Raw","Mech Fails","Composite","Grade"])
for i, r in enumerate(all_boss_score_rows):
    key = r["player"] + "|" + r["boss"] + "|" + r["date"]
    if "composite" not in r:
        comps = {"deaths": r["death_score"]}
        if r["mech_score"] is not None: comps["mechanics"] = r["mech_score"]
        if r["parse_score"] is not None: comps["parse_performance"] = r["parse_score"]
        if r["con_score"] is not None: comps["consumables"] = r["con_score"]
        aw = sum(SCORE_WEIGHTS[k] for k in comps)
        bc = round(sum(comps[k] * (SCORE_WEIGHTS[k] / aw) for k in comps)) if aw > 0 else 50
        bc = min(100, max(0, bc))
        r["composite"] = bc
        r["grade"] = _get_grade(bc)
    drow(ws, i+2, [key, r["player"],r["date"],r["boss"],
                    r["mech_score"],r["death_score"],
                    r["parse_score"],r["con_score"],
                    r["deaths_raw"],r.get("mech_fails",0),
                    r["composite"],r["grade"]], even=i%2==0)

# ── d_mech_detail: per player × boss × date mechanic events for explainer ──
ws_md = wb.create_sheet("d_mech_detail")
hdr(ws_md, 1, ["Key","Player","Boss","Date","Ability","Result","Display","Fix","Value"])
md_row = 2
for (player, boss, date), events in _mech_events_cache.items():
    for evt in events:
        key = f"{player}|{boss}|{date}|{evt['ability']}"
        drow(ws_md, md_row, [key, player, boss, date, evt["ability"],
                              evt["result"], evt["display"], evt["fix"],
                              evt["value"]], even=md_row%2==0)
        md_row += 1

print(f"Score data sheets done ({len(all_score_rows)} + {len(all_boss_score_rows)} rows, {md_row-2} mech events).")

# ══════════════════════════════════════════════════════════════════
#  SUMMARY DASHBOARD
# ══════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Summary")
CS = 11
title(ws, 1, "SUMMARY DASHBOARD", CS)

meta = dict(df_meta.iloc[0]) if len(df_meta) else {}
# Guild rankings: build speed dict from DataFrame
_gr_speed = {}
if len(df_guild_rankings):
    speed_row = df_guild_rankings[df_guild_rankings["metric"] == "speed"]
    if len(speed_row):
        sr = speed_row.iloc[0]
        _gr_speed = {
            "serverRank": {"number": sr.get("serverRank", "—") if pd.notna(sr.get("serverRank")) else "—"},
            "regionRank": {"number": sr.get("regionRank", "—") if pd.notna(sr.get("regionRank")) else "—"},
        }
speed = _gr_speed

# Guild info block
ws.cell(row=3, column=1, value="Guild:").font = F_BOLD
_team = roster.get("team", {})
_guild_name = _team.get("guild_name", "")
_team_name_display = _team.get("team_name", "")
_server = _team.get("server", "")
_region = _team.get("region", "US")
_guild_line = " — ".join(filter(None, [_guild_name, _team_name_display])) or "Raid Team"
_realm_line = " — ".join(filter(None, [_server.title(), _region.upper()])) or ""

ws.cell(row=3, column=2, value=_guild_line).font = F_BODY
ws.cell(row=4, column=1, value="Realm:").font = F_BOLD
ws.cell(row=4, column=2, value=_realm_line).font = F_BODY
ws.cell(row=5, column=1, value="Mains:").font = F_BOLD
ws.cell(row=5, column=2, value=len(mains_list)).font = F_BODY
ws.cell(row=5, column=3, value="Alts:").font = F_BOLD
ws.cell(row=5, column=4, value=len(alts_linked)+len(alts_unlinked)).font = F_BODY

ws.cell(row=3, column=5, value="Raid:").font = F_BOLD
ws.cell(row=3, column=6, value=RAID_DISPLAY_NAME).font = F_BODY
ws.cell(row=4, column=5, value="Speed Rank:").font = F_BOLD
sr_s = (speed.get("serverRank") or {}).get("number", "—")
sr_r = (speed.get("regionRank") or {}).get("number", "—")
ws.cell(row=4, column=6, value=f"Server #{sr_s} | Region #{sr_r}").font = F_BODY
ws.cell(row=5, column=5, value="Last Refresh:").font = F_BOLD
ws.cell(row=5, column=6, value=meta.get("pull_timestamp", "")[:16]).font = F_BODY

# ── LEADERBOARDS (3 groups) ──
medals = ["🥇","🥈","🥉","4.","5."]

# GROUP 1: Best Avg Parse (mains, kills, relevant metric)
section(ws, 7, "BEST AVG PARSE", 3, c0=1)
hdr(ws, 8, ["#","Player","Avg Parse"], c0=1)

char_avg_parse_all = {}
for r in ranking_rows:
    if r["main_alt"] == "Main":
        if roster_locked and r["character"] not in rostered_chars:
            continue
        use = (r["metric"] == "dps" and r["role"] != "Healer") or (r["metric"] == "hps" and r["role"] == "Healer")
        if use and r["parse_pct"]:
            char_avg_parse_all.setdefault(r["character"], []).append(r["parse_pct"])

parse_board = sorted([(c, sum(p)/len(p)) for c, p in char_avg_parse_all.items() if p],
                     key=lambda x: -x[1])

for i, (char, avg) in enumerate(parse_board[:5]):
    r = 9 + i
    ws.cell(row=r, column=1, value=medals[i]).font = F_BODY
    ws.cell(row=r, column=2, value=char_to_player.get(char, char)).font = F_BODY
    ws.cell(row=r, column=3, value=f"{avg:.1f}%").font = F_BODY
    apply_class_fill(ws, r, 3, char_to_class.get(char, ""))

# GROUP 2: Average Score (mains, composite score)
section(ws, 7, "AVERAGE SCORE", 3, c0=5)
hdr(ws, 8, ["#","Player","Avg Score"], c0=5)

char_avg_score_all = {}
for sr_row in score_rows:
    player = sr_row["player"]
    if roster_locked and player not in rostered_players:
        continue
    # Find character for this player via roster_rows
    char = None
    for rr in roster_rows:
        if rr["player"] == player:
            char = rr["character"]
            break
    if not char:
        continue
    if char_to_mainalt.get(char, "Alt") != "Main":
        continue
    comp = sr_row.get("composite")
    if comp is not None:
        char_avg_score_all.setdefault(char, []).append(comp)

score_board = sorted([(c, sum(a)/len(a)) for c, a in char_avg_score_all.items() if a],
                     key=lambda x: -x[1])

for i, (char, avg) in enumerate(score_board[:5]):
    r = 9 + i
    ws.cell(row=r, column=5, value=medals[i]).font = F_BODY
    ws.cell(row=r, column=6, value=char_to_player.get(char, char)).font = F_BODY
    ws.cell(row=r, column=7, value=f"{avg:.0f}").font = F_BODY
    apply_class_fill(ws, r, 3, char_to_class.get(char, ""), start_col=5)

# GROUP 3: Most Improved (composite score Δ) — compare latest vs previous week
section(ws, 7, "MOST IMPROVED", 3, c0=9)
hdr(ws, 8, ["#","Player","Score Δ"], c0=9)

if this_week and last_week:
    improvement = []
    for r in mains_list:
        player = r["player"]
        if roster_locked and r["character"] not in rostered_chars:
            continue
        tw_score = next((s["composite"] for s in score_rows if s["player"] == player and s["date"] == this_week), None)
        lw_score = next((s["composite"] for s in score_rows if s["player"] == player and s["date"] == last_week), None)
        if tw_score is not None and lw_score is not None:
            delta = tw_score - lw_score
            improvement.append((r["character"], player, delta))

    improvement.sort(key=lambda x: -x[2])
    for i, (char, player, delta) in enumerate(improvement[:5]):
        r_row = 9 + i
        ws.cell(row=r_row, column=9, value=medals[i]).font = F_BODY
        ws.cell(row=r_row, column=10, value=player).font = F_BODY
        agg_str = f"▲+{delta}" if delta > 0 else (f"▼{delta}" if delta < 0 else "—")
        ws.cell(row=r_row, column=11, value=agg_str).font = F_GREEN_BOLD if delta > 0 else (F_RED_BOLD if delta < 0 else F_BODY)
        apply_class_fill(ws, r_row, 3, char_to_class.get(char, ""), start_col=9)

# ── RAID READINESS — MAINS ──
row = 15
section(ws, row, "RAID READINESS — MAINS", CS)
row += 1
hdr(ws, row, ["Player","Character","Class","iLvl","Lowest Piece",
              "Status","Miss. Ench","Empty Sock","Tier Pcs"])
row += 1
for i, r in enumerate(mains_list):
    if r["has_gear_data"]:
        issues = r["missing_ench"] + r["empty_sockets"]
        status = "✅ Ready" if issues == 0 else f"⚠️ {issues} issue{'s' if issues != 1 else ''}"
        lowest = f"{r['lowest_slot']} {r['lowest_ilvl']}" if r['lowest_slot'] and r['lowest_ilvl'] != "—" else "—"
        vals = [r["player"], r["character"], r["class"], r["ilvl"], lowest,
                status, r["missing_ench"], r["empty_sockets"], r["tier_count"]]
    else:
        vals = [r["player"], r["character"], r["class"], r["ilvl"], "—",
                "No Data", "—", "—", "—"]
        issues = -1  # sentinel for coloring
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=j+1, value=v)
        c.font = F_BODY; c.border = BT
    apply_class_fill(ws, row, 9, r["class"])
    sc = ws.cell(row=row, column=6)  # Status in col F
    if issues == -1:
        sc.fill = X_EVEN  # neutral gray for no data
    else:
        sc.fill = X_GREEN if issues == 0 else (X_YELLOW if issues <= 2 else X_RED)
    row += 1

# ALTS readiness
row += 1
section(ws, row, "RAID READINESS — ALTS", CS)
row += 1
hdr(ws, row, ["Player","Character","Class","iLvl","Lowest Piece",
              "Status","Miss. Ench","Empty Sock","Tier Pcs"])
row += 1
all_alts = alts_linked + alts_unlinked
for i, r in enumerate(all_alts):
    if r["has_gear_data"]:
        issues = r["missing_ench"] + r["empty_sockets"]
        status = "✅ Ready" if issues == 0 else f"⚠️ {issues}"
        lowest = f"{r['lowest_slot']} {r['lowest_ilvl']}" if r['lowest_slot'] and r['lowest_ilvl'] != "—" else "—"
        vals = [r["player"], r["character"], r["class"], r["ilvl"], lowest,
                status, r["missing_ench"], r["empty_sockets"], r["tier_count"]]
    else:
        vals = [r["player"], r["character"], r["class"], r["ilvl"], "—",
                "No Data", "—", "—", "—"]
    fill = X_EVEN if i % 2 == 0 else X_ODD
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=j+1, value=v)
        c.font = F_BODY; c.fill = fill; c.border = BT
    # Status coloring
    sc = ws.cell(row=row, column=6)  # Status in col F
    if r["has_gear_data"]:
        sc.fill = X_GREEN if issues == 0 else (X_YELLOW if issues <= 2 else X_RED)
    row += 1

widths(ws, [14,18,15,6,16,14,12,12,8,14,12,10])
print("Summary done.")

# ══════════════════════════════════════════════════════════════════
#  RAIDS SHEET
# ══════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Raids")
CS_R = 7 + len(BOSS_ORDER)
title(ws, 1, f"RAIDS — {RAID_DISPLAY_NAME}", CS_R)
ws.cell(row=2, column=1, value=f"Data: {', '.join(all_dates)}").font = F_SMALL
ws.cell(row=2, column=1).fill = X_DARK
for col in range(1, CS_R+1): ws.cell(row=2, column=col).fill = X_DARK

# Header row shared by all three tables
headers_k = ["Character","Class / Spec","Role","Main/Alt","Player","Score","Δ"] + [boss_short(b) for b in BOSS_ORDER]

# Precompute composite lookup: player → latest, best, avg composite + delta vs raid avg
_latest_comp = {}
_best_comp = {}
_avg_comp = {}
_comp_delta = {}
for player in set(r["player"] for r in score_rows):
    psorted = sorted([s for s in score_rows if s["player"] == player], key=lambda x: x["date"])
    if psorted:
        _latest_comp[player] = psorted[-1]["composite"]
        _best_comp[player] = max(s["composite"] for s in psorted)
        _avg_comp[player] = round(sum(s["composite"] for s in psorted) / len(psorted))
        _comp_delta[player] = psorted[-1]["composite"] - raid_avg_composite
# Ensure every roster character's player is covered
for rr in roster_rows:
    ch, p = rr["character"], rr["player"]
    if p not in _latest_comp and ch in _latest_comp:
        _latest_comp[p] = _latest_comp[ch]
        _best_comp[p] = _best_comp.get(ch, _latest_comp[ch])
        _avg_comp[p] = _avg_comp.get(ch, _latest_comp[ch])
        _comp_delta[p] = _comp_delta[ch]

# BEST PERFORMANCE — first (most important)
section(ws, 4, "BEST PERFORMANCE — Best Score vs Raid Avg (Kills)", CS_R)
hdr(ws, 5, headers_k)

row = 6
for i, r in enumerate(roster_rows):
    char = r["character"]
    player = r["player"]
    comp = _best_comp.get(player, "")
    delta = _comp_delta.get(player)
    delta_str = f"▲+{delta:.0f}" if delta and delta > 0 else (f"▼{delta:.0f}" if delta and delta < 0 else "—")

    vals = [char, f"{r['class']} / {r['spec']}", r["role"], r["main_alt"], player, comp, delta_str]
    for boss in BOSS_ORDER:
        best_s = best_score_by_char_boss[char].get(boss, 0)
        raid_avg_best = raid_avg_best_score_boss.get(boss, 0)
        if best_s > 0:
            d = round(best_s - raid_avg_best)
            delta_boss = f"▲+{d}" if d > 0 else (f"▼{d}" if d < 0 else "—")
            vals.append(f"{best_s} ({delta_boss})")
        else:
            vals.append("—")
    fill = X_EVEN if i % 2 == 0 else X_ODD
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=j+1, value=v)
        c.font = F_BODY; c.fill = fill; c.border = BT
        if j >= 7: c.alignment = AC
    apply_class_fill(ws, row, CS_R, r["class"])
    # Color the score column
    sc = ws.cell(row=row, column=6)
    if isinstance(comp, (int, float)):
        g = _get_grade(comp)
        gc_map = {"A":"2E7D32","B":"1565C0","C":"F57F17","D":"E65100","F":"C62828"}
        bg_map = {"A":"E8F5E9","B":"E3F2FD","C":"FFF8E1","D":"FFF3E0","F":"FFEBEE"}
        sc.font = Font(name=FN, size=10, bold=True, color=gc_map.get(g,"333333"))
        sc.fill = PatternFill("solid", fgColor=bg_map.get(g,"FFFFFF"))
    sc.alignment = AC
    row += 1

# Arrow formatting for delta column
add_arrow_formatting(ws, f"G6:G{row-1}")
# Arrow formatting for boss columns
for bi in range(len(BOSS_ORDER)):
    bc = get_column_letter(8 + bi)
    add_arrow_formatting(ws, f"{bc}6:{bc}{row-1}")

# AVERAGE PERFORMANCE — second
row += 1
section(ws, row, "AVERAGE PERFORMANCE — Avg Score vs Raid Avg (Kills Only)", CS_R)
row += 1
hdr(ws, row, headers_k)
row += 1

avg_start = row
for i, r in enumerate(roster_rows):
    char = r["character"]
    player = r["player"]
    comp = _avg_comp.get(player, "")
    delta = _comp_delta.get(player)
    delta_str = f"▲+{delta:.0f}" if delta and delta > 0 else (f"▼{delta:.0f}" if delta and delta < 0 else "—")

    vals = [char, f"{r['class']} / {r['spec']}", r["role"], r["main_alt"], player, comp, delta_str]
    for boss in BOSS_ORDER:
        avg_s = avg_score_by_char_boss[char].get(boss, 0)
        raid_avg_avg = raid_avg_avg_score_boss.get(boss, 0)
        if avg_s > 0:
            d = round(avg_s - raid_avg_avg)
            delta_boss = f"▲+{d}" if d > 0 else (f"▼{d}" if d < 0 else "—")
            vals.append(f"{avg_s:.0f} ({delta_boss})")
        else:
            vals.append("—")
    fill = X_EVEN if i % 2 == 0 else X_ODD
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=j+1, value=v)
        c.font = F_BODY; c.fill = fill; c.border = BT
        if j >= 7: c.alignment = AC
    apply_class_fill(ws, row, CS_R, r["class"])
    sc = ws.cell(row=row, column=6)
    if isinstance(comp, (int, float)):
        g = _get_grade(comp)
        gc_map = {"A":"2E7D32","B":"1565C0","C":"F57F17","D":"E65100","F":"C62828"}
        bg_map = {"A":"E8F5E9","B":"E3F2FD","C":"FFF8E1","D":"FFF3E0","F":"FFEBEE"}
        sc.font = Font(name=FN, size=10, bold=True, color=gc_map.get(g,"333333"))
        sc.fill = PatternFill("solid", fgColor=bg_map.get(g,"FFFFFF"))
    sc.alignment = AC
    row += 1

add_arrow_formatting(ws, f"G{avg_start}:G{row-1}")
for bi in range(len(BOSS_ORDER)):
    bc = get_column_letter(8 + bi)
    add_arrow_formatting(ws, f"{bc}{avg_start}:{bc}{row-1}")

widths(ws, [18, 22, 8, 8, 14, 8, 8] + [16]*len(BOSS_ORDER))
print("Raids done.")

# ══════════════════════════════════════════════════════════════════
#  RAID PERFORMANCE — Dropdown-driven (one night at a time)
# ══════════════════════════════════════════════════════════════════
MAX_PLAYERS = 25  # max player rows to pre-allocate on visible sheet

# ── Pre-compute data for hidden sheets ──
rp_boss_rows = []     # one row per (date, boss)
rp_player_rows = []   # one row per (date, rank)
rp_detail_rows = []   # one row per (date, boss, rank)

# Precompute overall avg parse per boss (all dates, kills, mains+alts)
_overall_boss_avg_parse = {}
for boss in BOSS_ORDER:
    all_p = [r["parse_pct"] for r in ranking_rows
             if r["boss"] == boss and r["metric"] == "dps"
             and r["role"] != "Healer" and r["parse_pct"]]
    _overall_boss_avg_parse[boss] = sum(all_p)/len(all_p) if all_p else 0

for date_idx, date in enumerate(all_dates):
    date_fights_local = [f for f in fight_rows if f["date"] == date]
    date_deaths_local = [d for d in death_rows if d["date"] == date]
    prev_date = all_dates[date_idx + 1] if date_idx + 1 < len(all_dates) else ""

    boss_map = defaultdict(list)
    for f in date_fights_local: boss_map[f["boss"]].append(f)

    # ── Boss-level data ──
    for boss in BOSS_ORDER:
        bfights = boss_map.get(boss, [])
        if not bfights:
            continue  # skip unfought bosses entirely — no row in d_rp_boss

        kills_f = [f for f in bfights if f["kill"]]
        result = "Kill" if kills_f else "Wipe"
        best = kills_f[0] if kills_f else bfights[-1]
        result_tag = "Kill ✓" if kills_f else "Wipe ✗"

        boss_dps_vals = [r["amount"] for r in ranking_rows
                         if r["date"] == date and r["boss"] == boss
                         and r["metric"] == "dps" and r["role"] != "Healer" and r["amount"]]
        boss_parse_vals = [r["parse_pct"] for r in ranking_rows
                           if r["date"] == date and r["boss"] == boss
                           and r["metric"] == "dps" and r["parse_pct"]]
        raid_dps = f"{sum(boss_dps_vals)/len(boss_dps_vals)/1000:.0f}k" if boss_dps_vals else "—"
        raid_parse = f"{sum(boss_parse_vals)/len(boss_parse_vals):.0f}%" if boss_parse_vals else "—"

        total_deaths = sum(f["deaths_total"] for f in bfights)
        kb_counts = defaultdict(int)
        for d in date_deaths_local:
            if d["boss"] == boss and d["killing_blow"]:
                kb_counts[d["killing_blow"]] += 1
        top_kb = max(kb_counts, key=kb_counts.get) if kb_counts else "—"

        # Parse vs overall average
        if boss_parse_vals:
            night_avg = sum(boss_parse_vals) / len(boss_parse_vals)
            overall_avg = _overall_boss_avg_parse.get(boss, 0)
            pva_d = round(night_avg - overall_avg)
            parse_vs_avg = f"▲+{pva_d}" if pva_d > 0 else (f"▼{pva_d}" if pva_d < 0 else "—")
        else:
            parse_vs_avg = "—"

        rp_boss_rows.append({
            "key": date + boss, "date": date, "boss": boss,
            "result": result, "result_tag": result_tag, "pulls": len(bfights),
            "avg_parse": raid_parse, "raid_dps": raid_dps,
            "parse_vs_avg": parse_vs_avg,
            "deaths": total_deaths, "top_death": top_kb})

        # ── Per-boss per-player detail ──
        player_boss_agg = defaultdict(lambda: {"parses": [], "amounts": [], "deaths": 0})
        for r in ranking_rows:
            if r["date"] == date and r["boss"] == boss:
                use = (r["metric"] == "dps" and r["role"] != "Healer") or \
                      (r["metric"] == "hps" and r["role"] == "Healer")
                if use and r["parse_pct"]:
                    pba = player_boss_agg[r["player"]]
                    pba["parses"].append(r["parse_pct"])
                    pba["amounts"].append(r["amount"])
        for d in date_deaths_local:
            if d["boss"] == boss:
                player_boss_agg[d["player"]]["deaths"] += 1

        # Compute raid avg composite for this boss on this date
        boss_date_comps = [b["composite"] for b in boss_score_rows
                           if b["boss"] == boss and b["date"] == date]
        raid_avg_boss_comp = sum(boss_date_comps) / len(boss_date_comps) if boss_date_comps else 50

        detail_list = []
        for player, pba in player_boss_agg.items():
            if pba["parses"]:
                avg_p = sum(pba["parses"]) / len(pba["parses"])
                avg_a = sum(pba["amounts"]) / len(pba["amounts"])
                # Score Δ: player's boss composite vs raid avg composite for this boss+date
                player_boss_comp = next((b["composite"] for b in boss_score_rows
                                         if b["player"] == player and b["boss"] == boss
                                         and b["date"] == date), None)
                if player_boss_comp is not None:
                    d = round(player_boss_comp - raid_avg_boss_comp)
                    vs = f"▲+{d}" if d > 0 else (f"▼{d}" if d < 0 else "—")
                else:
                    vs = "—"
                detail_list.append((player, avg_p, avg_a, pba["deaths"], player_boss_comp, vs))
        detail_list.sort(key=lambda x: -x[1])

        for rank, (player, avg_p, avg_a, deaths, score, vs) in enumerate(detail_list, 1):
            rp_detail_rows.append({
                "key": date + boss + f"{rank:02d}",
                "date": date, "boss": boss, "rank": rank,
                "player": player, "parse": f"{avg_p:.0f}%",
                "dps": f"{avg_a/1000:.0f}k", "deaths": deaths,
                "score": score if score is not None else "", "vs_prev": vs})

    # ── Player summary (aggregated across all bosses for this date) ──
    player_agg = defaultdict(lambda: {"chars": set(), "cls_spec": set(), "parses": [],
                                       "amounts": [], "deaths": 0, "top_kb": defaultdict(int)})
    for r in ranking_rows:
        if r["date"] == date:
            use = (r["metric"] == "dps" and r["role"] != "Healer") or \
                  (r["metric"] == "hps" and r["role"] == "Healer")
            if use:
                pa = player_agg[r["player"]]
                pa["chars"].add(r["character"])
                pa["cls_spec"].add(f"{r['class']}/{r['spec']}")
                if r["parse_pct"]:
                    pa["parses"].append(r["parse_pct"])
                    pa["amounts"].append(r["amount"])
    for d in date_deaths_local:
        pa = player_agg[d["player"]]
        pa["deaths"] += 1
        if d["killing_blow"]:
            pa["top_kb"][d["killing_blow"]] += 1

    plist = []
    # Compute raid average composite for this date (all players)
    date_comps = [s["composite"] for s in score_rows if s["date"] == date and s["composite"] is not None]
    raid_avg_comp_date = sum(date_comps) / len(date_comps) if date_comps else 50

    # Precompute player→composite for this date
    _date_comp_lookup = {}
    for s in score_rows:
        if s["date"] == date and s["composite"] is not None:
            _date_comp_lookup[s["player"]] = s["composite"]

    # Compute raid average parse for this date (all players)
    date_all_parses = []
    for r in ranking_rows:
        if r["date"] == date:
            use = (r["metric"] == "dps" and r["role"] != "Healer") or \
                  (r["metric"] == "hps" and r["role"] == "Healer")
            if use and r["parse_pct"]:
                date_all_parses.append(r["parse_pct"])
    raid_avg_parse_date = sum(date_all_parses) / len(date_all_parses) if date_all_parses else 0

    for player, agg in player_agg.items():
        if agg["parses"]:
            avg_p = sum(agg["parses"]) / len(agg["parses"])
            avg_a = sum(agg["amounts"]) / len(agg["amounts"])
            top_kb = max(agg["top_kb"], key=agg["top_kb"].get) if agg["top_kb"] else "—"

            # vs raid average parse for this date
            parse_delta = round(avg_p - raid_avg_parse_date)
            vs_parse = f"▲+{parse_delta}" if parse_delta > 0 else (f"▼{parse_delta}" if parse_delta < 0 else "—")

            # Score vs raid's average composite for this date
            tw_comp = _date_comp_lookup.get(player)
            if tw_comp is not None:
                comp_delta = round(tw_comp - raid_avg_comp_date)
                score_vs = f"▲+{comp_delta}" if comp_delta > 0 else (f"▼{comp_delta}" if comp_delta < 0 else "—")
            else:
                score_vs = "—"

            plist.append((player, agg, avg_p, avg_a, top_kb, vs_parse, score_vs))
    plist.sort(key=lambda x: -x[2])

    for rank, (player, agg, avg_p, avg_a, top_kb, vs_parse, score_vs) in enumerate(plist, 1):
        rp_player_rows.append({
            "key": date + f"{rank:02d}", "date": date, "rank": rank,
            "player": player, "characters": ", ".join(sorted(agg["chars"])),
            "cls_spec": ", ".join(sorted(agg["cls_spec"])),
            "avg_parse": f"{avg_p:.0f}%", "vs_avg_parse": vs_parse,
            "score_vs_avg": score_vs,
            "deaths": agg["deaths"], "top_death": top_kb})

print(f"Raid Perf data: {len(rp_boss_rows)} boss rows, {len(rp_player_rows)} player rows, {len(rp_detail_rows)} detail rows")

# ── Write hidden data sheets ──
ws_rb = wb.create_sheet("d_rp_boss")
hdr(ws_rb, 1, ["key","date","boss","result","pulls","avg_parse","raid_dps","parse_vs_avg","deaths","top_death"])
for i, r in enumerate(rp_boss_rows):
    drow(ws_rb, i+2, [r["key"],r["date"],r["boss"],r["result"],r["pulls"],
                       r["avg_parse"],r["raid_dps"],r["parse_vs_avg"],r["deaths"],r["top_death"]], even=i%2==0)

ws_rp = wb.create_sheet("d_rp_player")
hdr(ws_rp, 1, ["key","date","rank","player","characters","cls_spec","avg_parse","vs_avg_parse","score_vs_avg","deaths","top_death"])
for i, r in enumerate(rp_player_rows):
    drow(ws_rp, i+2, [r["key"],r["date"],r["rank"],r["player"],r["characters"],
                       r["cls_spec"],r["avg_parse"],r["vs_avg_parse"],r["score_vs_avg"],
                       r["deaths"],r["top_death"]], even=i%2==0)

ws_rd = wb.create_sheet("d_rp_detail")
hdr(ws_rd, 1, ["key","date","boss","rank","player","parse","dps","deaths","score","vs_prev"])
for i, r in enumerate(rp_detail_rows):
    drow(ws_rd, i+2, [r["key"],r["date"],r["boss"],r["rank"],r["player"],
                       r["parse"],r["dps"],r["deaths"],r["score"],r["vs_prev"]], even=i%2==0)

# ══════════════════════════════════════════════════════════════════
#  RAID PERFORMANCE — Visible sheet with dropdown + XLOOKUP
# ══════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Raid Performance")

# ── Row 1: Title ──
title(ws, 1, "RAID PERFORMANCE ANALYSIS", 8)

# ── Row 2: Header bar with dropdown ──
ws.cell(row=2, column=1, value="Raid Night:").font = F_BOLD
ws.cell(row=2, column=2, value=this_week).font = F_BOLD
ws.cell(row=2, column=2).fill = X_YELLOW

# Data validation dropdown for date
date_list_str = ",".join(all_dates)
dv = DataValidation(type="list", formula1=f'"{date_list_str}"', allow_blank=False)
dv.error = "Pick a raid night"
dv.errorTitle = "Invalid Date"
dv.prompt = "Select a raid night"
dv.promptTitle = "Raid Night"
ws.add_data_validation(dv)
dv.add(ws["B2"])

ws.cell(row=2, column=4, value=RAID_DISPLAY_NAME).font = F_BODY
ws.cell(row=2, column=5, value="Mythic").font = F_BODY
# Kills count formula
ws.cell(row=2, column=6).value = '=COUNTIFS(d_rp_boss!B:B,TEXT($B$2,"YYYY-MM-DD"),d_rp_boss!D:D,"Kill")&"/8 Killed"'
ws.cell(row=2, column=6).font = F_BOLD

# Wipefest link — dynamic based on dropdown
# Store date→report_code map in col Z (hidden)
date_to_code = {}
if len(df_fights):
    for _, f in df_fights.iterrows():
        d = str(f.get("date", ""))
        c = f.get("report_code", "")
        if d and c:
            date_to_code[d] = c
for di, date in enumerate(all_dates):
    ws.cell(row=1+di, column=26, value=date)   # Z = dates
    ws.cell(row=1+di, column=27, value=date_to_code.get(date, ""))  # AA = codes
ws.column_dimensions['Z'].hidden = True
ws.column_dimensions['AA'].hidden = True

# HYPERLINK formula: looks up report code for selected date
ws.cell(row=2, column=7).value = (
    '=IFERROR(HYPERLINK("https://www.wipefest.gg/report/"'
    '&XLOOKUP(TEXT($B$2,"YYYY-MM-DD"),$Z:$Z,$AA:$AA,""),'
    '"Wipefest ↗"),"")'
)
ws.cell(row=2, column=7).font = Font(size=10, color="4FC3F7", underline="single")

# WCL report link — dynamic based on dropdown (same date→code lookup)
ws.cell(row=2, column=8).value = (
    '=IFERROR(HYPERLINK("https://www.warcraftlogs.com/reports/"'
    '&XLOOKUP(TEXT($B$2,"YYYY-MM-DD"),$Z:$Z,$AA:$AA,""),'
    '"WCL ↗"),"")'
)
ws.cell(row=2, column=8).font = Font(size=10, color="4FC3F7", underline="single")

# ── Row 4: RAID OVERVIEW section ──
OVER_ROW = 4
section(ws, OVER_ROW, "RAID OVERVIEW — All Bosses Fought", 8)
hdr(ws, OVER_ROW + 1, ["Boss", "Result", "Pulls", "Avg Parse", "Raid DPS", "Parse vs Avg", "Deaths", "Top Death"])

# Boss data rows (6-13) — XLOOKUP formulas
# d_rp_boss columns: A=key, B=date, C=boss, D=result, E=pulls, F=avg_parse, G=raid_dps, H=parse_vs_avg, I=deaths, J=top_death
for i, boss in enumerate(BOSS_ORDER):
    r = OVER_ROW + 2 + i
    ws.cell(row=r, column=1, value=boss).font = F_BODY
    ws.cell(row=r, column=1).border = BT
    # Columns B-H: XLOOKUP($B$2 & boss_name, key_col, data_col, "—")
    key_expr = f'TEXT($B$2,"YYYY-MM-DD")&A{r}'
    col_map = {2: "D", 3: "E", 4: "F", 5: "G", 6: "H", 7: "I", 8: "J"}
    for col, dcol in col_map.items():
        formula = f'=XLOOKUP({key_expr},d_rp_boss!$A:$A,d_rp_boss!{dcol}:{dcol},"—")'
        c = ws.cell(row=r, column=col, value=formula)
        c.font = F_BODY; c.border = BT; c.alignment = AC
    ws.cell(row=r, column=1).fill = X_EVEN if i % 2 == 0 else X_ODD
    for col in range(2, 9):
        ws.cell(row=r, column=col).fill = X_EVEN if i % 2 == 0 else X_ODD

# Arrow formatting for Parse vs Avg column
boss_first = OVER_ROW + 2
boss_last = OVER_ROW + 2 + len(BOSS_ORDER) - 1
add_arrow_formatting(ws, f"F{boss_first}:F{boss_last}")

# ── Row 15: PLAYER SUMMARY section ──
PLAY_ROW = OVER_ROW + 2 + len(BOSS_ORDER) + 1
section(ws, PLAY_ROW, "PLAYER SUMMARY — Sorted by Avg Parse ↓", 8)
hdr(ws, PLAY_ROW + 1, ["Player", "Character(s)", "Class/Spec", "Avg Parse", "vs Avg",
                         "Score vs Avg", "Deaths", "Top Death"])

# Player data rows — XLOOKUP with rank-based key
# d_rp_player columns: A=key, B=date, C=rank, D=player, E=characters, F=cls_spec,
#                       G=avg_parse, H=vs_avg_parse, I=score_vs_avg, J=deaths, K=top_death
PLAY_DATA_START = PLAY_ROW + 2
rank_offset = PLAY_DATA_START - 1  # ROW()-offset gives rank 1,2,3...
pcol_map = {1: "D", 2: "E", 3: "F", 4: "G", 5: "H", 6: "I", 7: "J", 8: "K"}

for slot in range(MAX_PLAYERS):
    r = PLAY_DATA_START + slot
    key_expr = f'TEXT($B$2,"YYYY-MM-DD")&TEXT(ROW()-{rank_offset},"00")'
    for col, dcol in pcol_map.items():
        formula = f'=XLOOKUP({key_expr},d_rp_player!$A:$A,d_rp_player!{dcol}:{dcol},"")'
        c = ws.cell(row=r, column=col, value=formula)
        c.font = F_BODY; c.border = BT
        if col >= 4: c.alignment = AC

# Class-based conditional formatting for Player Summary rows
play_end = PLAY_DATA_START + MAX_PLAYERS - 1
play_range = f"A{PLAY_DATA_START}:H{play_end}"
for cls_name, cls_fill in CLASS_FILL.items():
    ws.conditional_formatting.add(play_range, FormulaRule(
        formula=[f'ISNUMBER(SEARCH("{cls_name}",$C{PLAY_DATA_START}))'],
        fill=cls_fill))

# ── Horizontal per-boss detail panels — dynamic boss names ──
# 8 panel slots across. Each panel is 6 data cols + 1 gap = 7 cols wide.
# Panel titles are formulas that pull the Nth boss fought that night.
# If fewer than 8 bosses were fought, remaining panels are blank.
DETAIL_START_COL = 10  # Column J
PANEL_WIDTH = 7  # 6 data cols + 1 gap
# Compute max players per boss from actual data (instead of hardcoding 20)
from collections import Counter
players_per_boss_date = Counter()
for r in rp_detail_rows:
    players_per_boss_date[(r["date"], r["boss"])] += 1
DETAIL_PLAYERS = max(players_per_boss_date.values()) if players_per_boss_date else 20
print(f"  Detail panels: {DETAIL_PLAYERS} player rows (from data)")
# d_rp_detail columns: A=key, B=date, C=boss, D=rank, E=player, F=parse, G=dps, H=deaths, I=score, J=vs_prev
DATE_EXPR = 'TEXT($B$2,"YYYY-MM-DD")'

for pi in range(8):  # 8 panel slots
    pc = DETAIL_START_COL + pi * PANEL_WIDTH  # first col of this panel
    boss_n = pi + 1  # 1-indexed boss number for this slot
    title_ref = f"${get_column_letter(pc)}${OVER_ROW}"  # e.g. $J$4

    # Row OVER_ROW: Panel title — dynamic boss name from FILTER
    # Wrap BOTH sides in TEXT() because Google Sheets auto-converts date strings to serials
    title_formula = f'=IFERROR(INDEX(FILTER(d_rp_boss!C:C,TEXT(d_rp_boss!B:B,"YYYY-MM-DD")=TEXT($B$2,"YYYY-MM-DD")),{boss_n}),"")'
    c = ws.cell(row=OVER_ROW, column=pc, value=title_formula)
    c.font = F_SECTION; c.fill = X_SEC
    for ci in range(1, 6):
        ws.cell(row=OVER_ROW, column=pc + ci).fill = X_SEC
        ws.cell(row=OVER_ROW, column=pc + ci).font = F_SECTION

    # Result tag (Kill/Wipe) — col+4
    ws.cell(row=OVER_ROW, column=pc + 4).value = \
        f'=IF({title_ref}="","",XLOOKUP({DATE_EXPR}&{title_ref},d_rp_boss!$A:$A,d_rp_boss!D:D,"—"))'
    ws.cell(row=OVER_ROW, column=pc + 4).font = F_SECTION

    # Summary stats row
    sr = OVER_ROW + 1
    ws.cell(row=sr, column=pc).value = \
        f'=IF({title_ref}="","","Parse: "&XLOOKUP({DATE_EXPR}&{title_ref},d_rp_boss!$A:$A,d_rp_boss!F:F,"—")&"  |  DPS: "&XLOOKUP({DATE_EXPR}&{title_ref},d_rp_boss!$A:$A,d_rp_boss!G:G,"—")&"  |  Deaths: "&XLOOKUP({DATE_EXPR}&{title_ref},d_rp_boss!$A:$A,d_rp_boss!I:I,"—"))'
    ws.cell(row=sr, column=pc).font = F_SMALL

    # Column headers — only show if boss exists
    hr = OVER_ROW + 2
    for hi, h in enumerate(["Player", "Parse", "DPS", "Deaths", "Score", "Score Δ"]):
        c = ws.cell(row=hr, column=pc + hi, value=h)
        c.font = F_HDR; c.fill = X_HDR; c.alignment = AC

    # Player detail rows — XLOOKUP referencing dynamic title cell
    det_start = OVER_ROW + 3
    det_offset = det_start - 1
    dcol_map = {0: "E", 1: "F", 2: "G", 3: "H", 4: "I", 5: "J"}

    for slot in range(DETAIL_PLAYERS):
        r = det_start + slot
        for ci, dcol in dcol_map.items():
            formula = f'=IF({title_ref}="","",XLOOKUP({DATE_EXPR}&{title_ref}&TEXT(ROW()-{det_offset},"00"),d_rp_detail!$A:$A,d_rp_detail!{dcol}:{dcol},""))'
            c = ws.cell(row=r, column=pc + ci, value=formula)
            c.font = F_BODY; c.border = BT
            if ci >= 1: c.alignment = AC
        for ci in range(6):
            ws.cell(row=r, column=pc + ci).fill = X_EVEN if slot % 2 == 0 else X_ODD

# ── Charts will be added after chart_data is built ──

# ── Column widths ──
widths(ws, [22, 25, 20, 10, 12, 12, 8, 20])
# Panel column widths: [Player, Parse, DPS, Deaths, Score, Score Δ, gap] × 8
for pi in range(8):
    pc = DETAIL_START_COL + pi * PANEL_WIDTH
    for ci, w in enumerate([16, 8, 10, 8, 8, 10, 2]):
        ws.column_dimensions[get_column_letter(pc + ci)].width = w

ws.freeze_panes = "A3"

# ── Arrow colors: ▲ green, ▼ red ──
# Player summary "vs Avg" column (col 5 = E)
add_arrow_formatting(ws, f"E{PLAY_DATA_START}:E{PLAY_DATA_START + MAX_PLAYERS - 1}")
# Player summary "Score vs Avg" column (col 6 = F)
add_arrow_formatting(ws, f"F{PLAY_DATA_START}:F{PLAY_DATA_START + MAX_PLAYERS - 1}")
# Detail panels "Score Δ" column (last data col of each panel)
det_start = OVER_ROW + 3
for pi in range(8):
    vs_col = get_column_letter(DETAIL_START_COL + pi * PANEL_WIDTH + 5)
    add_arrow_formatting(ws, f"{vs_col}{det_start}:{vs_col}{det_start + DETAIL_PLAYERS - 1}")

print("Raid Performance done.")

# ══════════════════════════════════════════════════════════════════
#  CHARACTER VIEW — Hidden Data Sheets
# ══════════════════════════════════════════════════════════════════
all_chars = sorted(set(r["character"] for r in roster_rows))
all_players = sorted(set(r["player"] for r in roster_rows))
default_char = mains_list[0]["character"] if mains_list else all_chars[0]
PLAYER_PREFIX = "👤 "  # prefix for player-aggregated keys

# Group chars by player
from collections import defaultdict as _dd
chars_by_player = _dd(list)
for r in roster_rows:
    chars_by_player[r["player"]].append(r)

# ── d_cv_info: one row per character + one row per player ──
ws_cvi = wb.create_sheet("d_cv_info")
hdr(ws_cvi, 1, ["character","player","class","spec","role","ilvl","prog",
                 "avg_parse","attend_pct","total_deaths","avg_deaths_night","player_chars","vs_raid_deaths","realm_slug"])
cvi_row = 2

# Precompute raid avg deaths per player per night
# For each night: total deaths / number of unique players present
raid_deaths_per_player_night = []
for date in all_dates:
    night_deaths = sum(1 for d in death_rows if d["date"] == date)
    players_present = set()
    for r in ranking_rows:
        if r["date"] == date:
            players_present.add(char_to_player.get(r["character"], r["character"]))
    if players_present:
        raid_deaths_per_player_night.append(night_deaths / len(players_present))
raid_avg_dpn = sum(raid_deaths_per_player_night) / len(raid_deaths_per_player_night) if raid_deaths_per_player_night else 0

# Character rows
for char in all_chars:
    rr = next((r for r in roster_rows if r["character"] == char), None)
    if not rr: continue
    player = char_to_player.get(char, char)
    player_chars_list = sorted(set(r["character"] for r in roster_rows if r["player"] == player))

    is_healer = rr["role"] == "Healer"
    metric = "hps" if is_healer else "dps"
    parses = [r["parse_pct"] for r in ranking_rows
              if r["character"] == char and r["metric"] == metric and r["parse_pct"]]
    avg_p = sum(parses)/len(parses) if parses else 0

    nights_present = sum(1 for date in all_dates
                         if any(r["date"] == date and r["character"] == char
                                for r in ranking_rows))

    char_d = [d for d in death_rows if d["character"] == char]
    total_d = len(char_d)
    avg_d = total_d / nights_present if nights_present > 0 else 0

    # vs raid avg: ▲ = better than avg (fewer deaths, green), ▼ = worse (more deaths, red)
    vs_raid = ""
    if nights_present > 0:
        delta = avg_d - raid_avg_dpn
        if delta > 0.05: vs_raid = f"▼+{delta:.1f}"
        elif delta < -0.05: vs_raid = f"▲{delta:.1f}"
        else: vs_raid = "—"

    drow(ws_cvi, cvi_row, [char, player, rr["class"], rr["spec"], rr["role"],
                        rr["ilvl"], rr["raid_prog"],
                        f"{avg_p:.1f}%", f"{nights_present}/{num_nights}", total_d, f"{avg_d:.1f}",
                        ", ".join(player_chars_list), vs_raid, rr.get("realm_slug", "")], even=cvi_row%2==0)
    cvi_row += 1

# Player-aggregated rows
for player in all_players:
    pchars = chars_by_player[player]
    char_names = sorted(r["character"] for r in pchars)
    pk = PLAYER_PREFIX + player

    # Class/spec/role: if all same show it, else "Multi"
    classes = set(r["class"] for r in pchars)
    specs = set(r["spec"] for r in pchars)
    roles = set(r["role"] for r in pchars)
    cls_val = list(classes)[0] if len(classes) == 1 else ", ".join(sorted(classes))
    spec_val = list(specs)[0] if len(specs) == 1 else ", ".join(sorted(specs))
    role_val = list(roles)[0] if len(roles) == 1 else ", ".join(sorted(roles))

    # Best ilvl, prog across chars
    best_ilvl = max((r["ilvl"] for r in pchars if r["ilvl"]), default="")
    best_prog = max((r["raid_prog"] for r in pchars if r["raid_prog"]), default="", key=lambda x: x or "")

    # Avg parse across ALL chars for this player
    all_parses = []
    for ch in char_names:
        chrr = next((r for r in roster_rows if r["character"] == ch), None)
        if not chrr: continue
        m = "hps" if chrr["role"] == "Healer" else "dps"
        all_parses.extend(r["parse_pct"] for r in ranking_rows
                          if r["character"] == ch and r["metric"] == m and r["parse_pct"])
    avg_p = sum(all_parses)/len(all_parses) if all_parses else 0

    # Attendance: present if ANY char present
    nights_present = sum(1 for date in all_dates
                         if any(any(r["date"] == date and r["character"] == ch
                                    for r in ranking_rows)
                                for ch in char_names))
    attend = (nights_present / num_nights * 100) if num_nights > 0 else 0

    # Deaths across all chars
    total_d = sum(len([d for d in death_rows if d["character"] == ch]) for ch in char_names)
    avg_d = total_d / nights_present if nights_present > 0 else 0

    # vs raid avg: ▲ = better (fewer deaths, green), ▼ = worse (more deaths, red)
    vs_raid = ""
    if nights_present > 0:
        delta = avg_d - raid_avg_dpn
        if delta > 0.05: vs_raid = f"▼+{delta:.1f}"
        elif delta < -0.05: vs_raid = f"▲{delta:.1f}"
        else: vs_raid = "—"

    drow(ws_cvi, cvi_row, [pk, player, cls_val, spec_val, role_val,
                            best_ilvl, best_prog,
                            f"{avg_p:.1f}%", f"{nights_present}/{num_nights}", total_d, f"{avg_d:.1f}",
                            ", ".join(char_names), vs_raid, ""], even=cvi_row%2==0)
    cvi_row += 1
ws_cvi.freeze_panes = "A2"

# ── d_cv_boss: one row per (character, boss) + player-aggregated rows ──
ws_cvb = wb.create_sheet("d_cv_boss")
hdr(ws_cvb, 1, ["key","character","boss","best_parse","best_dps","avg_parse",
                 "avg_dps","kills","this_wk","vs_roster"])

# Precompute roster avg parse per boss
roster_avg_parse_boss = {}
for boss in BOSS_ORDER:
    all_p = []
    for r in ranking_rows:
        if r["boss"] == boss:
            use = (r["metric"] == "dps" and r["role"] != "Healer") or (r["metric"] == "hps" and r["role"] == "Healer")
            if use and r["parse_pct"]: all_p.append(r["parse_pct"])
    roster_avg_parse_boss[boss] = sum(all_p)/len(all_p) if all_p else 0

cvb_row = 2

# Character boss rows
for char in all_chars:
    rr = next((r for r in roster_rows if r["character"] == char), None)
    if not rr: continue
    is_healer = rr["role"] == "Healer"
    metric = "hps" if is_healer else "dps"
    char_ranks = [r for r in ranking_rows if r["character"] == char and r["kill"] and r["metric"] == metric]

    for boss in BOSS_ORDER:
        br = [r for r in char_ranks if r["boss"] == boss]
        best_p = max((r["parse_pct"] for r in br if r["parse_pct"]), default=0)
        best_a = max((r["amount"] for r in br), default=0)
        avg_p = sum(r["parse_pct"] or 0 for r in br)/len(br) if br else 0
        avg_a = sum(r["amount"] for r in br)/len(br) if br else 0
        kills = kills_by_char_boss[char].get(boss, 0)
        wk = week_kills_by_char_boss[char].get(boss, 0)

        vs = ""
        if avg_p > 0:
            delta = avg_p - roster_avg_parse_boss.get(boss, 0)
            if delta > 0: vs = f"▲+{delta:.0f}%"
            elif delta < 0: vs = f"▼{delta:.0f}%"
            else: vs = "—"

        if kills > 0:
            drow(ws_cvb, cvb_row, [char + boss, char, boss, f"{best_p}%", f"{best_a/1000:.0f}k",
                                    f"{avg_p:.0f}%", f"{avg_a/1000:.0f}k", kills,
                                    "✓" if wk > 0 else "—", vs], even=cvb_row%2==0)
        else:
            drow(ws_cvb, cvb_row, [char + boss, char, boss, "—","—","—","—",0,"—","—"], even=cvb_row%2==0)
        cvb_row += 1

# Player-aggregated boss rows
for player in all_players:
    pk = PLAYER_PREFIX + player
    char_names = sorted(r["character"] for r in chars_by_player[player])

    for boss in BOSS_ORDER:
        # Collect rankings from ALL chars for this player on this boss
        all_br = []
        for ch in char_names:
            chrr = next((r for r in roster_rows if r["character"] == ch), None)
            if not chrr: continue
            m = "hps" if chrr["role"] == "Healer" else "dps"
            all_br.extend(r for r in ranking_rows
                          if r["character"] == ch and r["boss"] == boss and r["kill"] and r["metric"] == m)

        best_p = max((r["parse_pct"] for r in all_br if r["parse_pct"]), default=0)
        best_a = max((r["amount"] for r in all_br), default=0)
        avg_p = sum(r["parse_pct"] or 0 for r in all_br)/len(all_br) if all_br else 0
        avg_a = sum(r["amount"] for r in all_br)/len(all_br) if all_br else 0
        kills = sum(kills_by_char_boss[ch].get(boss, 0) for ch in char_names)
        wk = sum(week_kills_by_char_boss[ch].get(boss, 0) for ch in char_names)

        vs = ""
        if avg_p > 0:
            delta = avg_p - roster_avg_parse_boss.get(boss, 0)
            if delta > 0: vs = f"▲+{delta:.0f}%"
            elif delta < 0: vs = f"▼{delta:.0f}%"
            else: vs = "—"

        if kills > 0:
            drow(ws_cvb, cvb_row, [pk + boss, pk, boss, f"{best_p}%", f"{best_a/1000:.0f}k",
                                    f"{avg_p:.0f}%", f"{avg_a/1000:.0f}k", kills,
                                    "✓" if wk > 0 else "—", vs], even=cvb_row%2==0)
        else:
            drow(ws_cvb, cvb_row, [pk + boss, pk, boss, "—","—","—","—",0,"—","—"], even=cvb_row%2==0)
        cvb_row += 1
ws_cvb.freeze_panes = "A2"

# ── d_cv_deaths: character + player-aggregated ──
ws_cvd = wb.create_sheet("d_cv_deaths")
hdr(ws_cvd, 1, ["key","character","date","boss","kill","killing_blow","overkill"])
cvd_row = 2
deaths_by_char = _dd(list)
for d in death_rows:
    deaths_by_char[d["character"]].append(d)

# Character death rows
for char in all_chars:
    for rank, d in enumerate(deaths_by_char.get(char, []), 1):
        key = f"{char}{rank:02d}"
        drow(ws_cvd, cvd_row, [key, char, d["date"], d["boss"],
                                "Kill" if d["kill"] else "Wipe",
                                d["killing_blow"], d["overkill"]], even=cvd_row%2==0)
        cvd_row += 1

# Player-aggregated death rows (merge all chars, sorted by date desc)
max_deaths_per_entity = max((len(v) for v in deaths_by_char.values()), default=0)
for player in all_players:
    pk = PLAYER_PREFIX + player
    char_names = sorted(r["character"] for r in chars_by_player[player])
    player_deaths = []
    for ch in char_names:
        player_deaths.extend(deaths_by_char.get(ch, []))
    player_deaths.sort(key=lambda d: d["date"], reverse=True)
    max_deaths_per_entity = max(max_deaths_per_entity, len(player_deaths))
    for rank, d in enumerate(player_deaths, 1):
        key = f"{pk}{rank:02d}"
        drow(ws_cvd, cvd_row, [key, pk, d["date"], d["boss"],
                                "Kill" if d["kill"] else "Wipe",
                                d["killing_blow"], d["overkill"]], even=cvd_row%2==0)
        cvd_row += 1
ws_cvd.freeze_panes = "A2"

# ── d_cv_topkb: top killing blows per character/player (last 8 raid nights) ──
ws_cvkb = wb.create_sheet("d_cv_topkb")
hdr(ws_cvkb, 1, ["key","entity","rank","killing_blow","count"])
cvkb_row = 2
MAX_TOPKB = 10
TOPKB_NIGHTS = 8
recent_dates = set(all_dates[:TOPKB_NIGHTS])  # all_dates is sorted newest-first

# Character top killing blows
for char in all_chars:
    kb_counts = defaultdict(int)
    for d in deaths_by_char.get(char, []):
        if d["date"] in recent_dates and d["killing_blow"]:
            kb_counts[d["killing_blow"]] += 1
    top_kbs = sorted(kb_counts.items(), key=lambda x: -x[1])[:MAX_TOPKB]
    for rank, (kb, count) in enumerate(top_kbs, 1):
        key = f"{char}|{rank:02d}"
        drow(ws_cvkb, cvkb_row, [key, char, rank, kb, count], even=cvkb_row%2==0)
        cvkb_row += 1

# Player-aggregated top killing blows
for player in all_players:
    pk = PLAYER_PREFIX + player
    char_names = sorted(r["character"] for r in chars_by_player[player])
    kb_counts = defaultdict(int)
    for ch in char_names:
        for d in deaths_by_char.get(ch, []):
            if d["date"] in recent_dates and d["killing_blow"]:
                kb_counts[d["killing_blow"]] += 1
    top_kbs = sorted(kb_counts.items(), key=lambda x: -x[1])[:MAX_TOPKB]
    for rank, (kb, count) in enumerate(top_kbs, 1):
        key = f"{pk}|{rank:02d}"
        drow(ws_cvkb, cvkb_row, [key, pk, rank, kb, count], even=cvkb_row%2==0)
        cvkb_row += 1

ws_cvkb.freeze_panes = "A2"
print(f"  d_cv_topkb: {cvkb_row-2} rows (last {TOPKB_NIGHTS} nights)")

# ── d_cv_attend: character + player-aggregated ──
ws_cva = wb.create_sheet("d_cv_attend")
hdr(ws_cva, 1, ["key","character","date","present","deaths"])
cva_row = 2

# Pre-build attendance set from fight_roster (includes wipes)
_char_date_present = set()
if len(df_fight_roster):
    for _, fr in df_fight_roster.iterrows():
        _char_date_present.add((fr["player"], str(fr["date"])))

# Character attendance rows
for char in all_chars:
    player_for_char = char_to_player.get(char, char)
    for rank, date in enumerate(all_dates, 1):
        present = (player_for_char, date) in _char_date_present
        d_count = deaths_by_char_date[char].get(date, 0)
        key = f"{char}{rank:02d}"
        drow(ws_cva, cva_row, [key, char, date,
                                "✓" if present else "✗", d_count], even=cva_row%2==0)
        cva_row += 1

# Player-aggregated attendance rows
for player in all_players:
    pk = PLAYER_PREFIX + player
    char_names = sorted(r["character"] for r in chars_by_player[player])
    for rank, date in enumerate(all_dates, 1):
        present = (player, date) in _char_date_present
        d_count = sum(deaths_by_char_date[ch].get(date, 0) for ch in char_names)
        key = f"{pk}{rank:02d}"
        drow(ws_cva, cva_row, [key, pk, date,
                                "✓" if present else "✗", d_count], even=cva_row%2==0)
        cva_row += 1
ws_cva.freeze_panes = "A2"

print(f"CV data: {len(all_chars)} chars + {len(all_players)} players, {cvb_row-2} boss rows, {cvd_row-2} death rows, {cva_row-2} attend rows")

# ── d_cv_analyzer: best-parse fight per character × date (for WoWAnalyzer deep links) ──
ws_cvan = wb.create_sheet("d_cv_analyzer")
hdr(ws_cvan, 1, ["key","character","date","report_code","fight_id","actor_id","boss","parse_pct"])
cvan_row = 2

for char in all_chars:
    rr = next((r for r in roster_rows if r["character"] == char), None)
    if not rr:
        continue
    is_healer = rr["role"] == "Healer"
    metric = "hps" if is_healer else "dps"

    # Per-date best parse (KILLS ONLY — WoWAnalyzer links are for kill fight analysis)
    for date in all_dates:
        kill_fights_with_parse = [r for r in ranking_rows
                 if r["character"] == char and r["date"] == date
                 and r["kill"] and r["metric"] == metric and r["parse_pct"]]
        if not kill_fights_with_parse:
            continue
        best = max(kill_fights_with_parse, key=lambda r: r["parse_pct"])
        actor_id = _actor_id_lookup.get((best["report_code"], char), "")
        key = f"{char}|{date}"
        drow(ws_cvan, cvan_row, [key, char, date, best["report_code"],
                                  best["fight_id"], actor_id, best["boss"],
                                  best["parse_pct"]], even=cvan_row%2==0)
        cvan_row += 1

    # "All Time" best parse (kills only)
    all_kill_fights = [r for r in ranking_rows
                 if r["character"] == char and r["kill"]
                 and r["metric"] == metric and r["parse_pct"]]
    if all_kill_fights:
        best = max(all_kill_fights, key=lambda r: r["parse_pct"])
        actor_id = _actor_id_lookup.get((best["report_code"], char), "")
        key = f"{char}|All Time"
        drow(ws_cvan, cvan_row, [key, char, "All Time", best["report_code"],
                                  best["fight_id"], actor_id, best["boss"],
                                  best["parse_pct"]], even=cvan_row%2==0)
        cvan_row += 1

ws_cvan.freeze_panes = "A2"
print(f"  d_cv_analyzer: {cvan_row-2} rows")

# ── d_cv_mechfails: mechanic failures per character × date (for dynamic failures section) ──
ws_cvmf = wb.create_sheet("d_cv_mechfails")
hdr(ws_cvmf, 1, ["key","character","date","rank","boss","ability","result","display","fix","value"])
cvmf_row = 2

# Build character→player lookup
_char_to_player_cv = {}
for rr in roster_rows:
    _char_to_player_cv[rr["character"]] = rr["player"]

MAX_MECHFAIL_ROWS = 20

for char in all_chars:
    player = _char_to_player_cv.get(char, char)

    # Per-date failures
    for date in all_dates:
        fail_events = []
        for boss in BOSS_ORDER:
            evts = _mech_events_cache.get((char, boss, date), [])
            for evt in evts:
                if evt["result"] == "FAIL":
                    fail_events.append({
                        "boss": boss, "ability": evt["ability"],
                        "display": evt.get("display", ""), "fix": evt.get("fix", ""),
                        "value": evt.get("value", 0),
                    })

        # Deduplicate: boss+ability → sum value, count
        from collections import OrderedDict as _OD
        deduped = _OD()
        for fe in fail_events:
            k = (fe["boss"], fe["ability"])
            if k not in deduped:
                deduped[k] = dict(fe, count=1)
            else:
                deduped[k]["value"] += fe["value"]
                deduped[k]["count"] += 1

        sorted_fails = sorted(deduped.values(), key=lambda x: x["count"], reverse=True)
        for rank, fe in enumerate(sorted_fails[:MAX_MECHFAIL_ROWS], 1):
            key = f"{char}|{date}|{rank:02d}"
            result_text = f"✗ FAIL ×{fe['count']}" if fe["count"] > 1 else "✗ FAIL"
            drow(ws_cvmf, cvmf_row, [key, char, date, rank,
                                      boss_short(fe["boss"]), fe["ability"], result_text,
                                      fe["display"], fe["fix"],
                                      fe["value"] if fe["value"] > 0 else "—"],
                 even=cvmf_row%2==0)
            cvmf_row += 1

    # "All Time" aggregated failures
    all_fail_events = []
    for boss in BOSS_ORDER:
        for date in all_dates:
            evts = _mech_events_cache.get((char, boss, date), [])
            for evt in evts:
                if evt["result"] == "FAIL":
                    all_fail_events.append({
                        "boss": boss, "ability": evt["ability"],
                        "display": evt.get("display", ""), "fix": evt.get("fix", ""),
                        "value": evt.get("value", 0),
                    })

    deduped_all = _OD()
    for fe in all_fail_events:
        k = (fe["boss"], fe["ability"])
        if k not in deduped_all:
            deduped_all[k] = dict(fe, count=1)
        else:
            deduped_all[k]["value"] += fe["value"]
            deduped_all[k]["count"] += 1

    sorted_all = sorted(deduped_all.values(), key=lambda x: x["count"], reverse=True)
    for rank, fe in enumerate(sorted_all[:MAX_MECHFAIL_ROWS], 1):
        key = f"{char}|All Time|{rank:02d}"
        result_text = f"✗ FAIL ×{fe['count']}" if fe["count"] > 1 else "✗ FAIL"
        drow(ws_cvmf, cvmf_row, [key, char, "All Time", rank,
                                  boss_short(fe["boss"]), fe["ability"], result_text,
                                  fe["display"], fe["fix"],
                                  fe["value"] if fe["value"] > 0 else "—"],
             even=cvmf_row%2==0)
        cvmf_row += 1

    # Player-aggregated key (for "👤 Player" dropdown selections)
    # Only emit once per player (when we reach the first char of this player)
    pk = PLAYER_PREFIX + player
    player_chars = sorted(r["character"] for r in roster_rows if r["player"] == player)
    if char == player_chars[0]:
        # Per-date: gather events from ALL characters of this player
        for date in all_dates:
            pk_fail_events = []
            for pch in player_chars:
                for boss in BOSS_ORDER:
                    evts = _mech_events_cache.get((pch, boss, date), [])
                    for evt in evts:
                        if evt["result"] == "FAIL":
                            pk_fail_events.append({
                                "boss": boss, "ability": evt["ability"],
                                "display": evt.get("display", ""), "fix": evt.get("fix", ""),
                                "value": evt.get("value", 0),
                            })
            pk_deduped = _OD()
            for fe in pk_fail_events:
                k = (fe["boss"], fe["ability"])
                if k not in pk_deduped:
                    pk_deduped[k] = dict(fe, count=1)
                else:
                    pk_deduped[k]["value"] += fe["value"]
                    pk_deduped[k]["count"] += 1
            pk_sorted = sorted(pk_deduped.values(), key=lambda x: x["count"], reverse=True)
            for rank, fe in enumerate(pk_sorted[:MAX_MECHFAIL_ROWS], 1):
                key = f"{pk}|{date}|{rank:02d}"
                result_text = f"✗ FAIL ×{fe['count']}" if fe["count"] > 1 else "✗ FAIL"
                drow(ws_cvmf, cvmf_row, [key, pk, date, rank,
                                          boss_short(fe["boss"]), fe["ability"], result_text,
                                          fe["display"], fe["fix"],
                                          fe["value"] if fe["value"] > 0 else "—"],
                     even=cvmf_row%2==0)
                cvmf_row += 1

        # All Time: gather from all chars of this player
        pk_all_events = []
        for pch in player_chars:
            for boss in BOSS_ORDER:
                for date in all_dates:
                    evts = _mech_events_cache.get((pch, boss, date), [])
                    for evt in evts:
                        if evt["result"] == "FAIL":
                            pk_all_events.append({
                                "boss": boss, "ability": evt["ability"],
                                "display": evt.get("display", ""), "fix": evt.get("fix", ""),
                                "value": evt.get("value", 0),
                            })
        pk_all_deduped = _OD()
        for fe in pk_all_events:
            k = (fe["boss"], fe["ability"])
            if k not in pk_all_deduped:
                pk_all_deduped[k] = dict(fe, count=1)
            else:
                pk_all_deduped[k]["value"] += fe["value"]
                pk_all_deduped[k]["count"] += 1
        pk_sorted_all = sorted(pk_all_deduped.values(), key=lambda x: x["count"], reverse=True)
        for rank, fe in enumerate(pk_sorted_all[:MAX_MECHFAIL_ROWS], 1):
            key = f"{pk}|All Time|{rank:02d}"
            result_text = f"✗ FAIL ×{fe['count']}" if fe["count"] > 1 else "✗ FAIL"
            drow(ws_cvmf, cvmf_row, [key, pk, "All Time", rank,
                                      boss_short(fe["boss"]), fe["ability"], result_text,
                                      fe["display"], fe["fix"],
                                      fe["value"] if fe["value"] > 0 else "—"],
                 even=cvmf_row%2==0)
            cvmf_row += 1

ws_cvmf.freeze_panes = "A2"
print(f"  d_cv_mechfails: {cvmf_row-2} rows")

# Build the dropdown string for later XML injection
dd_items = [PLAYER_PREFIX + p for p in all_players] + ["──────────"] + all_chars
CV_DROPDOWN_STRING = ",".join(dd_items)

# ══════════════════════════════════════════════════════════════════
#  CHARACTER VIEW — Combined Display (merged CV + Character Detail)
# ══════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Character View")

# ── Row 1: Title ──
title(ws, 1, "CHARACTER PERFORMANCE", 12)

# ── Row 2: Dropdowns + hidden helper ──
ws.cell(row=2, column=1, value="Select:").font = F_BOLD
ws.cell(row=2, column=2, value=default_char).font = F_BOLD
ws.cell(row=2, column=2).fill = X_YELLOW

# Character/Player dropdown (placeholder for XML injection)
dv_name = DataValidation(type="list", formula1='"PLACEHOLDER"', allow_blank=False)
dv_name.error = "Pick a character or player"
ws.add_data_validation(dv_name)
dv_name.add(ws["B2"])

ws.cell(row=2, column=4, value="Raid Night:").font = F_BOLD
ws.cell(row=2, column=5, value="All Time").font = F_BOLD
ws.cell(row=2, column=5).fill = X_YELLOW

date_options = ["All Time"] + sorted(all_dates)
date_str_cv = ",".join(date_options)
dv_date_cv = DataValidation(type="list", formula1=f'"{date_str_cv}"', allow_blank=False)
dv_date_cv.prompt = "Select a raid night or All Time"
ws.add_data_validation(dv_date_cv)
dv_date_cv.add(ws["E2"])

# Hidden helper: Z2 = derived player name (for score lookups)
ws.cell(row=2, column=26,
        value='=IFERROR(XLOOKUP(B2,d_cv_info!$A:$A,d_cv_info!$B:$B,B2),B2)').font = F_BODY
ws.column_dimensions['Z'].hidden = True

CK = '$B$2'   # Character key for CV info/boss lookups
PK = '$Z$2'   # Player key for score lookups
DK = '$E$2'   # Date key

# ── Row 3-4: CHARACTER INFO ──
section(ws, 3, "CHARACTER INFO", 10)
info4 = {1: ("Player:", "B"), 3: ("Class:", "C"), 5: ("Spec:", "D"), 7: ("Role:", "E"), 9: ("iLvl:", "F")}
for col, (label, dcol) in info4.items():
    ws.cell(row=4, column=col, value=label).font = F_BOLD
    ws.cell(row=4, column=col+1,
            value=f'=XLOOKUP({CK},d_cv_info!$A:$A,d_cv_info!{dcol}:{dcol},"—")').font = F_BODY

info5 = {1: ("Prog:", "G"), 3: ("Attendance:", "I"), 5: ("Avg Parse:", "H"), 7: ("Characters:", "L")}
for col, (label, dcol) in info5.items():
    ws.cell(row=5, column=col, value=label).font = F_BOLD
    ws.cell(row=5, column=col+1,
            value=f'=XLOOKUP({CK},d_cv_info!$A:$A,d_cv_info!{dcol}:{dcol},"—")').font = F_BODY

# ── Row 6: External links (WCL, WoWAnalyzer, Wipefest) ──
# Hidden date→code mapping for Wipefest link (AB/AC columns, AD=most recent code)
for di, date in enumerate(all_dates):
    ws.cell(row=1+di, column=28, value=date)                             # AB = dates
    ws.cell(row=1+di, column=29, value=date_to_code.get(date, ""))       # AC = codes
ws.column_dimensions[get_column_letter(28)].hidden = True   # AB
ws.column_dimensions[get_column_letter(29)].hidden = True   # AC
most_recent_code = date_to_code.get(sorted(all_dates)[-1], "") if all_dates else ""
ws.cell(row=1, column=30, value=most_recent_code)
ws.column_dimensions[get_column_letter(30)].hidden = True   # AD

link_font = Font(name=FN, size=10, color="4FC3F7", underline="single")
# Realm slug is d_cv_info column N — empty for player-aggregated rows (links hidden)
realm_lookup = f'XLOOKUP({CK},d_cv_info!$A:$A,d_cv_info!N:N,"")'

# WCL report link (col 1-2) — links to the report for the selected date
ws.cell(row=6, column=1, value="Links:").font = F_BOLD
wcl_code = f'IF({DK}="All Time",$AD$1,XLOOKUP(TEXT({DK},"YYYY-MM-DD"),$AB:$AB,$AC:$AC,""))'
ws.cell(row=6, column=2).value = (
    f'=IFERROR(IF({wcl_code}="","",HYPERLINK("https://www.warcraftlogs.com/reports/"'
    f'&{wcl_code},"WCL Report ↗")),"")'
)
ws.cell(row=6, column=2).font = link_font

# WoWAnalyzer deep link (col 3-4) — links to best-parse fight for selected date
# d_cv_analyzer: A=key, D=report_code, E=fight_id, F=actor_id
# URL: wowanalyzer.com/report/{code}/{fightId}/{sourceId}/standard
ana_key = f'{CK}&"|"&{DK}'
ana_code = f'XLOOKUP({ana_key},d_cv_analyzer!$A:$A,d_cv_analyzer!$D:$D,"")'
ana_fight = f'XLOOKUP({ana_key},d_cv_analyzer!$A:$A,d_cv_analyzer!$E:$E,"")'
ana_actor = f'XLOOKUP({ana_key},d_cv_analyzer!$A:$A,d_cv_analyzer!$F:$F,"")'
ws.cell(row=6, column=4).value = (
    f'=IFERROR(IF(OR({realm_lookup}="",{ana_code}=""),"",HYPERLINK("https://wowanalyzer.com/report/"'
    f'&{ana_code}&"/"&{ana_fight}&"/"&{ana_actor}&"/standard","WoWAnalyzer ↗")),"")'
)
ws.cell(row=6, column=4).font = link_font

# Wipefest report link (col 5-6) — uses selected date or most recent
ws.cell(row=6, column=6).value = (
    f'=IFERROR(IF({realm_lookup}="","",HYPERLINK("https://www.wipefest.gg/report/"'
    f'&IF({DK}="All Time",$AD$1,XLOOKUP(TEXT({DK},"YYYY-MM-DD"),$AB:$AB,$AC:$AC,"")),'
    f'"Wipefest ↗")),"")'
)
ws.cell(row=6, column=6).font = link_font

# ── Row 7-9: OVERALL SCORE ──
section(ws, 7, "OVERALL PERFORMANCE SCORE", 10)
# d_scores: A=Key, B=Player, C=Date, D=Role, E=Mechanics, F=Death, G=Parse,
#           H=Consume, I=Composite, J=Grade, K=DeathsRaw, L=MechFails, M=MechPasses, N=Fights
score_labels_map = [
    ("Composite", "I"), ("Grade", "J"), ("Mechanics", "E"), ("Deaths", "F"),
    ("Parse", "G"), ("Pots", "H"), ("Fights", "N"),
]
for j, (label, _) in enumerate(score_labels_map):
    c = ws.cell(row=8, column=1+j, value=label)
    c.font = Font(name=FN, size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="4472C4")
    c.alignment = AC
for j, (_, dcol) in enumerate(score_labels_map):
    ws.cell(row=9, column=1+j,
            value=f'=XLOOKUP({PK}&"|"&{DK},d_scores!$A:$A,d_scores!{dcol}:{dcol},"—")')
    ws.cell(row=9, column=1+j).font = Font(name=FN, size=13, bold=True)
    ws.cell(row=9, column=1+j).alignment = AC
    ws.cell(row=9, column=1+j).border = BT

# Row 10: Scoring weights
weight_labels = ["", "", "40%", "35%", "20%", "5%", ""]
for j, wl in enumerate(weight_labels):
    c = ws.cell(row=10, column=1+j, value=wl)
    c.font = Font(name=FN, size=9, italic=True, color="666666")
    c.alignment = AC

# ── Row 11-19: BOSS PERFORMANCE ──
section(ws, 11, "BOSS PERFORMANCE", 11)
boss_hdrs = ["Boss","Best Parse","Avg Parse","Kills","This Wk","│","Grade","Mechs","Death","Parse","Pots"]
hdr(ws, 12, boss_hdrs)

# d_cv_boss: A=key, D=best_parse, F=avg_parse, H=kills, I=this_wk
# d_boss_scores: A=Key, E=Mechanics, F=Death, G=Parse, H=Consume, J=MechFails, K=Composite, L=Grade
for bi, boss in enumerate(BOSS_ORDER):
    r = 13 + bi
    ws.cell(row=r, column=1, value=boss_short(boss)).font = F_BODY
    # Parse side (from d_cv_boss)
    for col, dcol in [(2,"D"),(3,"F"),(4,"H"),(5,"I")]:
        ws.cell(row=r, column=col,
                value=f'=XLOOKUP({CK}&"{boss}",d_cv_boss!$A:$A,d_cv_boss!{dcol}:{dcol},"—")').font = F_BODY
    ws.cell(row=r, column=6, value="│").font = Font(name=FN, size=10, color="999999")
    # Score side (from d_boss_scores)
    for col, dcol in [(7,"L"),(8,"E"),(9,"F"),(10,"G"),(11,"H")]:
        ws.cell(row=r, column=col,
                value=f'=XLOOKUP({PK}&"|"&"{boss}"&"|"&{DK},d_boss_scores!$A:$A,d_boss_scores!{dcol}:{dcol},"—")').font = F_BODY
    fill = X_EVEN if bi % 2 == 0 else X_ODD
    for col in range(1, 12):
        c = ws.cell(row=r, column=col)
        c.fill = fill; c.border = BT; c.alignment = AC

# ── WEEK-OVER-WEEK TREND ──
TREND_ROW = 13 + len(BOSS_ORDER) + 1  # 22
section(ws, TREND_ROW, "▼ WEEK-OVER-WEEK TREND", 12)
trend_hdrs = ["Date","Grade","Score","Mechs","Deaths","Parse","Pots","Fights","Δ Score"]
hdr(ws, TREND_ROW + 1, trend_hdrs)

sorted_dates_cv = sorted(all_dates)
for di, date in enumerate(sorted_dates_cv):
    r = TREND_ROW + 2 + di
    ws.cell(row=r, column=1, value=date).font = F_BODY
    # d_scores col map: J=Grade, I=Composite, E=Mechanics, F=Death, G=Parse, H=Consume, L=MechFails, N=Fights
    for col, dcol in [(2,"J"),(3,"I"),(4,"E"),(5,"F"),(6,"G"),(7,"H"),(8,"N")]:
        ws.cell(row=r, column=col,
                value=f'=XLOOKUP({PK}&"|"&$A{r},d_scores!$A:$A,d_scores!{dcol}:{dcol},"")').font = F_BODY
    # Delta: difference from previous week composite
    if di > 0:
        ws.cell(row=r, column=9,
                value=f'=IF(AND(ISNUMBER(C{r}),ISNUMBER(C{r-1})),IF(C{r}-C{r-1}>0,"▲+"&(C{r}-C{r-1}),IF(C{r}-C{r-1}<0,"▼"&(C{r}-C{r-1}),"▲0")),"")').font = F_BODY
    for col in range(1, 11):
        c = ws.cell(row=r, column=col)
        c.border = BT; c.alignment = AC
        if di % 2 == 0: c.fill = X_EVEN

# ── BOSS-BY-BOSS IMPROVEMENT GRID ──
num_wk = len(sorted_dates_cv)
BIMP_ROW = TREND_ROW + 2 + num_wk + 1
section(ws, BIMP_ROW, "BOSS - WEEKLY METRICS", 1 + num_wk)

# Header: Boss + date columns
c = ws.cell(row=BIMP_ROW+1, column=1, value="Boss")
c.font = F_HDR; c.fill = X_DARK
for di, date in enumerate(sorted_dates_cv):
    c = ws.cell(row=BIMP_ROW+1, column=2+di, value=date[-5:])
    c.font = F_HDR; c.fill = X_DARK; c.alignment = AC

for bi, boss in enumerate(BOSS_ORDER):
    r = BIMP_ROW + 2 + bi
    ws.cell(row=r, column=1, value=boss_short(boss)).font = F_BODY
    ws.cell(row=r, column=1).border = BT
    for di, date in enumerate(sorted_dates_cv):
        ws.cell(row=r, column=2+di,
                value=f'=XLOOKUP({PK}&"|"&"{boss}"&"|"&"{date}",d_boss_scores!$A:$A,d_boss_scores!$K:$K,"")').font = Font(name=FN, size=9)
        c = ws.cell(row=r, column=2+di)
        c.border = BT; c.alignment = AC
        if bi % 2 == 0: c.fill = X_EVEN

# ── SURVIVABILITY ──
SURV_ROW = BIMP_ROW + 2 + len(BOSS_ORDER) + 1
section(ws, SURV_ROW, "SURVIVABILITY", 10)
sr = SURV_ROW + 1
ws.cell(row=sr, column=1, value="Total Deaths:").font = F_BOLD
ws.cell(row=sr, column=2,
        value=f'=XLOOKUP({CK},d_cv_info!$A:$A,d_cv_info!J:J,"—")').font = F_BODY
ws.cell(row=sr, column=3, value="Avg/Night:").font = F_BOLD
ws.cell(row=sr, column=4,
        value=f'=XLOOKUP({CK},d_cv_info!$A:$A,d_cv_info!K:K,"—")').font = F_BODY
ws.cell(row=sr, column=5, value="vs Raid:").font = F_BOLD
ws.cell(row=sr, column=6,
        value=f'=XLOOKUP({CK},d_cv_info!$A:$A,d_cv_info!M:M,"—")').font = F_BODY

add_arrow_formatting(ws, f"F{sr}:F{sr}")

# Top Killing Blows (last 8 raid nights, capped at 10)
sr += 2
hdr(ws, sr, ["#","Killing Blow","Count","",""])
sr += 1
KB_DISPLAY_ROWS = 10
kb_offset = sr
for slot in range(KB_DISPLAY_ROWS):
    r = sr + slot
    kb_key = f'{CK}&"|"&TEXT(ROW()-{kb_offset - 1},"00")'
    ws.cell(row=r, column=1,
            value=f'=IFERROR(XLOOKUP({kb_key},d_cv_topkb!$A:$A,d_cv_topkb!$C:$C,""),"")').font = F_BODY
    ws.cell(row=r, column=2,
            value=f'=IFERROR(XLOOKUP({kb_key},d_cv_topkb!$A:$A,d_cv_topkb!$D:$D,""),"")').font = F_BODY
    ws.cell(row=r, column=3,
            value=f'=IFERROR(XLOOKUP({kb_key},d_cv_topkb!$A:$A,d_cv_topkb!$E:$E,""),"")').font = F_BODY
    for col in range(1, 4):
        ws.cell(row=r, column=col).border = BT
    if slot % 2 == 0:
        for col in range(1, 4): ws.cell(row=r, column=col).fill = X_EVEN

sr = sr + KB_DISPLAY_ROWS + 1
ws.cell(row=sr, column=1, value="Attendance:").font = F_BOLD
ws.cell(row=sr, column=2,
        value=f'=XLOOKUP({CK},d_cv_info!$A:$A,d_cv_info!I:I,"—")').font = F_BODY

# ── MECHANIC FAILURES EXPLAINER (dynamic — driven by d_cv_mechfails) ──
AVOID_ROW = sr + 2
section(ws, AVOID_ROW, "MECHANIC FAILURES & HOW TO IMPROVE", 12)
# Headers: Boss(1), Ability(2), Result(3), What Happened(4-6 merged), How to Improve(7-10 merged), Dmg(11)
mex_r = AVOID_ROW + 1
for col, lbl, w in [(1,"Boss",None),(2,"Ability",None),(3,"Result",None),(4,"What Happened",None),
                     (7,"How to Improve",None),(11,"Dmg",None)]:
    c = ws.cell(row=mex_r, column=col, value=lbl)
    c.font = F_HDR; c.fill = X_DARK; c.alignment = AC
ws.merge_cells(start_row=mex_r, start_column=4, end_row=mex_r, end_column=6)
ws.merge_cells(start_row=mex_r, start_column=7, end_row=mex_r, end_column=10)
for col in range(4, 11):
    c = ws.cell(row=mex_r, column=col)
    c.fill = X_DARK; c.font = F_HDR

WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
FAIL_FONT = Font(name=FN, size=10, bold=True, color="C62828")
FIX_FONT = Font(name=FN, size=10, italic=True, color="1565C0")
NO_FAIL_FONT = Font(name=FN, size=10, italic=True, color="2E7D32")

# d_cv_mechfails columns: A=key, E=boss, F=ability, G=result, H=display, I=fix, J=value
# Key format: character|date|rank (e.g. "Constånce|All Time|01")
MF_KEY = f'{CK}&"|"&{DK}&"|"&TEXT(ROW()-{AVOID_ROW + 1},"00")'
MF_SHEET = "d_cv_mechfails"

MECHFAIL_DISPLAY_ROWS = 20
for slot in range(MECHFAIL_DISPLAY_ROWS):
    r = AVOID_ROW + 2 + slot
    fill = X_EVEN if slot % 2 == 0 else X_ODD

    # Col 1: Boss
    c = ws.cell(row=r, column=1,
                value=f'=IFERROR(XLOOKUP({MF_KEY},{MF_SHEET}!$A:$A,{MF_SHEET}!$E:$E,""),"")')
    c.font = F_BODY; c.border = BT; c.fill = fill
    # Col 2: Ability
    c = ws.cell(row=r, column=2,
                value=f'=IFERROR(XLOOKUP({MF_KEY},{MF_SHEET}!$A:$A,{MF_SHEET}!$F:$F,""),"")')
    c.font = F_BODY; c.border = BT; c.fill = fill
    # Col 3: Result
    c = ws.cell(row=r, column=3,
                value=f'=IFERROR(XLOOKUP({MF_KEY},{MF_SHEET}!$A:$A,{MF_SHEET}!$G:$G,""),"")')
    c.font = FAIL_FONT; c.border = BT; c.fill = fill
    # Col 4-6 merged: What Happened (display)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    c = ws.cell(row=r, column=4,
                value=f'=IFERROR(XLOOKUP({MF_KEY},{MF_SHEET}!$A:$A,{MF_SHEET}!$H:$H,""),"")')
    c.font = F_BODY; c.border = BT; c.alignment = WRAP_ALIGN; c.fill = fill
    for cc in range(5, 7):
        ws.cell(row=r, column=cc).fill = fill; ws.cell(row=r, column=cc).border = BT
    # Col 7-10 merged: How to Improve (fix)
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=10)
    c = ws.cell(row=r, column=7,
                value=f'=IFERROR(XLOOKUP({MF_KEY},{MF_SHEET}!$A:$A,{MF_SHEET}!$I:$I,""),"")')
    c.font = FIX_FONT; c.border = BT; c.alignment = WRAP_ALIGN; c.fill = fill
    for cc in range(8, 11):
        ws.cell(row=r, column=cc).fill = fill; ws.cell(row=r, column=cc).border = BT
    # Col 11: Damage
    c = ws.cell(row=r, column=11,
                value=f'=IFERROR(XLOOKUP({MF_KEY},{MF_SHEET}!$A:$A,{MF_SHEET}!$J:$J,""),"")')
    c.font = F_BODY; c.border = BT; c.fill = fill

# ── Conditional formatting: grade coloring ──
# Overall score grade (row 9, col 2)
add_grade_coloring(ws, "B9:B9")
# Overall score composite (row 9, col 1)
add_score_coloring(ws, "A9:A9")
# Overall score sub-scores (row 9, cols 3-6: Mechs, Deaths, Parse, Pots)
add_score_coloring(ws, f"C9:F9")

# Boss performance grades (rows 13-20, col 7 = G — Grade column)
boss_end = 12 + len(BOSS_ORDER)
add_grade_coloring(ws, f"G13:G{boss_end}")
# Boss performance sub-scores (cols 8-11 = H:K — Mechs, Death, Parse, Pots)
add_score_coloring(ws, f"H13:K{boss_end}")

# Week-over-week trend grades (col 2)
trend_data_start = TREND_ROW + 2
trend_data_end = trend_data_start + len(sorted_dates_cv) - 1
add_grade_coloring(ws, f"B{trend_data_start}:B{trend_data_end}")
# Trend Score column (col 3) — numeric score, apply grade-threshold coloring
add_score_coloring(ws, f"C{trend_data_start}:C{trend_data_end}")
# Trend sub-scores (cols 4-7 only: Mechs, Deaths, Parse, Pots — NOT Fights)
add_score_coloring(ws, f"D{trend_data_start}:G{trend_data_end}")
# Trend Δ Overall (col 9 = I) - green/red arrows
add_arrow_formatting(ws, f"I{trend_data_start}:I{trend_data_end}")

# Boss weekly metrics — arrow formatting for delta text
bimp_data_start = BIMP_ROW + 2
bimp_data_end = bimp_data_start + len(BOSS_ORDER) - 1
last_date_col = get_column_letter(1 + num_wk)
add_arrow_formatting(ws, f"B{bimp_data_start}:{last_date_col}{bimp_data_end}")

# Column widths and freeze
widths(ws, [16, 14, 14, 12, 16, 12, 12, 12, 12, 12, 12])
ws.freeze_panes = "A3"

# ── CV chart: DPS/HPS & Deaths over weeks ──
ws_cvcd = wb.create_sheet("d_cv_chart")
num_dates_cv = len(sorted_dates_cv)

ws_cvcd.cell(row=1, column=1, value="Entity").font = F_BOLD
for di, date in enumerate(sorted_dates_cv):
    ws_cvcd.cell(row=1, column=2+di, value=date).font = F_BODY

cv_chart_entities = []
for rr in roster_rows:
    cv_chart_entities.append(rr["character"])
player_chars_map = defaultdict(list)
for rr in roster_rows:
    p = char_to_player.get(rr["character"], rr["character"])
    if rr["character"] not in player_chars_map[p]:
        player_chars_map[p].append(rr["character"])
for player in sorted(player_chars_map.keys()):
    cv_chart_entities.append(f"👤 {player}")

cv_row = 2
for entity in cv_chart_entities:
    ws_cvcd.cell(row=cv_row, column=1, value=f"{entity} Parse").font = F_BODY
    ws_cvcd.cell(row=cv_row+1, column=1, value=f"{entity} Deaths").font = F_BODY
    ws_cvcd.cell(row=cv_row+2, column=1, value=f"{entity} Score").font = F_BODY

    if entity.startswith("👤 "):
        player_name = entity[2:]
        chars = player_chars_map.get(player_name, [])
    else:
        chars = [entity]

    # Derive player name for composite lookups
    if entity.startswith("👤 "):
        entity_player = entity[2:]
    else:
        entity_player = char_to_player.get(entity, entity)

    raw_entity_parse = []
    raw_entity_deaths = []
    raw_entity_comp = []
    for di, date in enumerate(sorted_dates_cv):
        parse_vals = []
        for ch in chars:
            parse_vals.extend(char_avg_parse_by_date[ch].get(date, []))
        raw_entity_parse.append(round(sum(parse_vals)/len(parse_vals), 1) if parse_vals else None)
        d_count = sum(deaths_by_char_date[ch].get(date, 0) for ch in chars)
        present = any(
            (char_to_player.get(ch, ch), date) in _char_date_present
            for ch in chars
        )
        raw_entity_deaths.append(d_count if present else None)
        # Composite from score_rows
        sr_match = next((s for s in score_rows if s["player"] == entity_player and s["date"] == date), None)
        raw_entity_comp.append(sr_match["composite"] if sr_match else None)

    for di, v in enumerate(interpolate_series(raw_entity_parse)):
        if v is not None: ws_cvcd.cell(row=cv_row, column=2+di, value=v)
    for di, v in enumerate(interpolate_series(raw_entity_deaths)):
        if v is not None: ws_cvcd.cell(row=cv_row+1, column=2+di, value=v)
    for di, v in enumerate(interpolate_series(raw_entity_comp)):
        if v is not None: ws_cvcd.cell(row=cv_row+2, column=2+di, value=v)
    cv_row += 3

# Dynamic helper rows — 6 series
DYN_ROW = cv_row + 2
CV_DYN_PARSE_ROW = DYN_ROW
CV_DYN_DEATHS_ROW = DYN_ROW + 1
CV_DYN_COMP_ROW = DYN_ROW + 2
CV_DYN_RAID_PARSE_ROW = DYN_ROW + 3
CV_DYN_RAID_DEATHS_ROW = DYN_ROW + 4
CV_DYN_RAID_COMP_ROW = DYN_ROW + 5

ws_cvcd.cell(row=CV_DYN_PARSE_ROW, column=1, value="Parse %").font = F_BOLD
ws_cvcd.cell(row=CV_DYN_DEATHS_ROW, column=1, value="Deaths").font = F_BOLD
ws_cvcd.cell(row=CV_DYN_COMP_ROW, column=1, value="Score").font = F_BOLD
ws_cvcd.cell(row=CV_DYN_RAID_PARSE_ROW, column=1, value="Raid Avg Parse %").font = F_BOLD
ws_cvcd.cell(row=CV_DYN_RAID_DEATHS_ROW, column=1, value="Raid Avg Deaths").font = F_BOLD
ws_cvcd.cell(row=CV_DYN_RAID_COMP_ROW, column=1, value="Raid Avg Score").font = F_BOLD

raid_avg_parse_by_date = defaultdict(list)
raid_deaths_per_date = defaultdict(int)
for r in ranking_rows:
    if r["parse_pct"]:
        use = (r["metric"] == "dps" and r["role"] != "Healer") or \
              (r["metric"] == "hps" and r["role"] == "Healer")
        if use: raid_avg_parse_by_date[r["date"]].append(r["parse_pct"])
for d in death_rows:
    raid_deaths_per_date[d["date"]] += 1

# Precompute raid avg composite per date
raid_avg_comp_by_date = defaultdict(list)
for sr_r in score_rows:
    raid_avg_comp_by_date[sr_r["date"]].append(sr_r["composite"])

for di in range(num_dates_cv):
    date_col = get_column_letter(2 + di)
    ws_cvcd.cell(row=CV_DYN_PARSE_ROW, column=2+di,
        value=f'=IFERROR(INDEX({date_col}:{date_col},MATCH(\'Character View\'!$B$2&" Parse",$A:$A,0)),"")')
    ws_cvcd.cell(row=CV_DYN_DEATHS_ROW, column=2+di,
        value=f'=IFERROR(INDEX({date_col}:{date_col},MATCH(\'Character View\'!$B$2&" Deaths",$A:$A,0)),"")')
    ws_cvcd.cell(row=CV_DYN_COMP_ROW, column=2+di,
        value=f'=IFERROR(INDEX({date_col}:{date_col},MATCH(\'Character View\'!$B$2&" Score",$A:$A,0)),"")')

raw_raid_parse = []
raw_raid_deaths = []
raw_raid_comp = []
for date in sorted_dates_cv:
    rp_vals = raid_avg_parse_by_date.get(date, [])
    raw_raid_parse.append(round(sum(rp_vals)/len(rp_vals), 1) if rp_vals else None)
    rd = raid_deaths_per_date.get(date, 0)
    pn = set()
    for r in ranking_rows:
        if r["date"] == date:
            pn.add(char_to_player.get(r["character"], r["character"]))
    raw_raid_deaths.append(round(rd / len(pn), 1) if pn else None)
    rc_vals = raid_avg_comp_by_date.get(date, [])
    raw_raid_comp.append(round(sum(rc_vals)/len(rc_vals), 1) if rc_vals else None)

for di, v in enumerate(interpolate_series(raw_raid_parse)):
    if v is not None: ws_cvcd.cell(row=CV_DYN_RAID_PARSE_ROW, column=2+di, value=v)
for di, v in enumerate(interpolate_series(raw_raid_deaths)):
    if v is not None: ws_cvcd.cell(row=CV_DYN_RAID_DEATHS_ROW, column=2+di, value=v)
for di, v in enumerate(interpolate_series(raw_raid_comp)):
    if v is not None: ws_cvcd.cell(row=CV_DYN_RAID_COMP_ROW, column=2+di, value=v)

# Add chart to Character View
ws_cv = wb["Character View"]
if num_dates_cv >= 2:
    from openpyxl.chart.series import DataPoint, SeriesLabel
    from openpyxl.chart.marker import Marker

    cv_ch = LineChart()
    cv_ch.title = "Performance Over Weeks"
    cv_ch.style = 10
    cv_ch.display_blanks = "span"

    cats = Reference(ws_cvcd, min_col=2, max_col=1+num_dates_cv, min_row=1)
    parse_ref = Reference(ws_cvcd, min_col=2, max_col=1+num_dates_cv, min_row=CV_DYN_PARSE_ROW)
    deaths_ref = Reference(ws_cvcd, min_col=2, max_col=1+num_dates_cv, min_row=CV_DYN_DEATHS_ROW)
    comp_ref = Reference(ws_cvcd, min_col=2, max_col=1+num_dates_cv, min_row=CV_DYN_COMP_ROW)
    raid_parse_ref = Reference(ws_cvcd, min_col=2, max_col=1+num_dates_cv, min_row=CV_DYN_RAID_PARSE_ROW)
    raid_deaths_ref = Reference(ws_cvcd, min_col=2, max_col=1+num_dates_cv, min_row=CV_DYN_RAID_DEATHS_ROW)
    raid_comp_ref = Reference(ws_cvcd, min_col=2, max_col=1+num_dates_cv, min_row=CV_DYN_RAID_COMP_ROW)

    cv_ch.add_data(parse_ref, from_rows=True)
    cv_ch.add_data(deaths_ref, from_rows=True)
    cv_ch.add_data(comp_ref, from_rows=True)
    cv_ch.add_data(raid_parse_ref, from_rows=True)
    cv_ch.add_data(raid_deaths_ref, from_rows=True)
    cv_ch.add_data(raid_comp_ref, from_rows=True)
    cv_ch.set_categories(cats)

    # Series 0: Parse % — blue solid with markers (this week)
    s0 = cv_ch.series[0]
    s0.graphicalProperties.line.width = 25000
    s0.graphicalProperties.line.solidFill = "4472C4"
    s0.marker = Marker(symbol="circle", size=6)
    s0.title = SeriesLabel(v="Parse %")

    # Series 1: Deaths — red solid with markers (this week)
    s1 = cv_ch.series[1]
    s1.graphicalProperties.line.width = 25000
    s1.graphicalProperties.line.solidFill = "C0504D"
    s1.marker = Marker(symbol="diamond", size=6)
    s1.title = SeriesLabel(v="Deaths")

    # Series 2: Score — green solid bold with markers (this week)
    s2 = cv_ch.series[2]
    s2.graphicalProperties.line.width = 30000
    s2.graphicalProperties.line.solidFill = "2E7D32"
    s2.marker = Marker(symbol="square", size=7)
    s2.title = SeriesLabel(v="Score")

    # Series 3: Raid Avg Parse % — blue dashed (overall)
    s3 = cv_ch.series[3]
    s3.graphicalProperties.line.width = 15000
    s3.graphicalProperties.line.solidFill = "4472C4"
    s3.graphicalProperties.line.dashStyle = "dash"
    s3.title = SeriesLabel(v="Raid Avg Parse %")

    # Series 4: Raid Avg Deaths — red dashed (overall)
    s4 = cv_ch.series[4]
    s4.graphicalProperties.line.width = 15000
    s4.graphicalProperties.line.solidFill = "C0504D"
    s4.graphicalProperties.line.dashStyle = "dash"
    s4.title = SeriesLabel(v="Raid Avg Deaths")

    # Series 5: Raid Avg Score — green dashed (overall)
    s5 = cv_ch.series[5]
    s5.graphicalProperties.line.width = 15000
    s5.graphicalProperties.line.solidFill = "2E7D32"
    s5.graphicalProperties.line.dashStyle = "dash"
    s5.title = SeriesLabel(v="Raid Avg Score")

    cv_ch.legend.position = "b"
    cv_ch.width = 20; cv_ch.height = 13
    ws_cv.add_chart(cv_ch, "N3")

print("Character View done.")

# ══════════════════════════════════════════════════════════════════
#  ROSTER SHEET
# ══════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Roster")
title(ws, 1, "GUILD ROSTER", 22)
ws.cell(row=2, column=1, value="Input: B-G  |  Auto: H-M  |  Sort: Mains→Alts, Tank→Healer→DPS").font = F_SMALL
ws.cell(row=2, column=1).fill = X_DARK
for col in range(1, 23): ws.cell(row=2, column=col).fill = X_DARK

# MAINS
row = 4
section(ws, row, "MAINS", 12)
row += 1
hdr(ws, row, ["#","Player","Character","Realm","Role","Main/Alt","Notes",
              "Class","Spec","Role (WCL)","iLvl","Raid Prog"])
row += 1
num = 1
for i, r in enumerate(mains_list):
    ov = roster_overrides.get(r["character"], {})
    role_ov = ov.get("role", "")
    notes = ov.get("notes", "")
    vals = [num, r["player"], r["character"], r["server"], role_ov, r["main_alt"], notes,
            r["class"], r["spec"], r["role"], r["ilvl"], r["raid_prog"]]
    fill = X_EVEN if i % 2 == 0 else X_ODD
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=j+1, value=v)
        c.font = F_BODY; c.fill = fill; c.border = BT
    num += 1; row += 1

# ALTS
row += 1
section(ws, row, "ALTS", 12)
row += 1
hdr(ws, row, ["#","Player","Character","Realm","Role","Main/Alt","Notes",
              "Class","Spec","Role (WCL)","iLvl","Raid Prog"])
row += 1
for i, r in enumerate(all_alts):
    ov = roster_overrides.get(r["character"], {})
    role_ov = ov.get("role", "")
    notes_default = "pug/trial" if r["player"] == r["character"] else "linked"
    notes = ov.get("notes", "") or notes_default
    vals = [num, r["player"], r["character"], r["server"], role_ov, r["main_alt"], notes,
            r["class"], r["spec"], r["role"], r["ilvl"], r["raid_prog"]]
    fill = X_EVEN if i % 2 == 0 else X_ODD
    for j, v in enumerate(vals):
        c = ws.cell(row=row, column=j+1, value=v)
        c.font = F_BODY; c.fill = fill; c.border = BT
    num += 1; row += 1

# COMPOSITION sidebar (col O-R)
comp_row = 4
section(ws, comp_row, "COMPOSITION", 4, c0=15)
comp_row += 1
hdr(ws, comp_row, ["Class","Mains","Alts","Total"], c0=15)
comp_row += 1

class_counts = defaultdict(lambda: {"mains": 0, "alts": 0})
for r in roster_rows:
    if r["main_alt"] == "Main":
        class_counts[r["class"]]["mains"] += 1
    else:
        class_counts[r["class"]]["alts"] += 1

for cls in sorted(class_counts.keys()):
    cc = class_counts[cls]
    for j, v in enumerate([cls, cc["mains"], cc["alts"], cc["mains"]+cc["alts"]]):
        c = ws.cell(row=comp_row, column=15+j, value=v)
        c.font = F_BODY; c.border = BT
    comp_row += 1

# Totals
for j, v in enumerate(["Total", len(mains_list), len(all_alts), len(roster_rows)]):
    c = ws.cell(row=comp_row, column=15+j, value=v)
    c.font = F_BOLD; c.border = BT

# BUFF TRACKER sidebar (col T-V)
buff_row = 4
section(ws, buff_row, "RAID BUFF TRACKER", 3, c0=20)
buff_row += 1
hdr(ws, buff_row, ["Buff","Description","Covered?"], c0=20)
buff_row += 1

buffs = [
    ("DeathKnight", "Anti-Magic Zone (raid CD)"),
    ("DemonHunter", "Chaos Brand (+5% Magic Dmg)"),
    ("Druid", "Mark of the Wild (+2% Vers)"),
    ("Evoker", "Blessing of the Bronze (+movement)"),
    ("Hunter", "Hunter's Mark (+5% target dmg)"),
    ("Mage", "Arcane Intellect (+5% Int)"),
    ("Monk", "Mystic Touch (+5% Phys Dmg)"),
    ("Paladin", "Devotion Aura (+3% DR)"),
    ("Priest", "Power Word: Fortitude (+5% Stam)"),
    ("Rogue", "Atrophic Poison (-3% enemy dmg)"),
    ("Shaman", "Windfury Totem (melee haste)"),
    ("Warlock", "Healthstones / Summons / Soulstone"),
    ("Warrior", "Battle Shout (+5% AP)"),
]

present_classes = set(r["class"] for r in roster_rows)
for cls, desc in buffs:
    covered = "✓" if cls in present_classes else "✗"
    ws.cell(row=buff_row, column=20, value=cls).font = F_BODY
    ws.cell(row=buff_row, column=21, value=desc).font = F_BODY
    ws.cell(row=buff_row, column=22, value=covered).font = F_GREEN_BOLD if covered == "✓" else F_RED_BOLD
    buff_row += 1

widths(ws, [4,14,18,15,8,8,12,15,15,10,6,8,10,2, 15,6,6,6,2, 15,30,8])
print("Roster done.")

# ══════════════════════════════════════════════════════════════════
#  CHART DATA
# ══════════════════════════════════════════════════════════════════

ws = wb.create_sheet("chart_data")
# Row 1: Week dates
ws.cell(row=1, column=1, value="Week").font = F_BOLD
for i, date in enumerate(sorted(all_dates)):
    ws.cell(row=1, column=2+i, value=date).font = F_BODY

# Row 2: Raid avg DPS — interpolate gaps (no-kill nights)
ws.cell(row=2, column=1, value="Raid Avg DPS (k)").font = F_BODY
raw_dps = []
for date in sorted(all_dates):
    vals = raid_avg_dps_by_date.get(date, [])
    raw_dps.append(round(sum(vals)/len(vals)/1000) if vals else None)
for i, v in enumerate(interpolate_series(raw_dps)):
    ws.cell(row=2, column=2+i, value=v if v is not None else "").font = F_BODY

# Row 3: Raid deaths (always present — no interpolation needed)
ws.cell(row=3, column=1, value="Raid Deaths").font = F_BODY
for i, date in enumerate(sorted(all_dates)):
    total = sum(1 for d in death_rows if d["date"] == date
                and (not roster_locked or d["character"] in rostered_chars))
    ws.cell(row=3, column=2+i, value=total).font = F_BODY

# Row 4: Raid Avg Score
ws.cell(row=4, column=1, value="Raid Avg Score").font = F_BODY
raid_comp_cd = defaultdict(list)
for sr_r in score_rows:
    if roster_locked and sr_r["player"] not in rostered_players:
        continue
    raid_comp_cd[sr_r["date"]].append(sr_r["composite"])
raw_raid_comp_cd = []
for date in sorted(all_dates):
    vals = raid_comp_cd.get(date, [])
    raw_raid_comp_cd.append(round(sum(vals)/len(vals), 1) if vals else None)
for i, v in enumerate(interpolate_series(raw_raid_comp_cd)):
    ws.cell(row=4, column=2+i, value=v if v is not None else "").font = F_BODY

# Per-boss DPS and deaths — interpolate gaps
r = 6
for boss in BOSS_ORDER:
    ws.cell(row=r, column=1, value=f"{boss} DPS").font = F_BODY
    ws.cell(row=r+1, column=1, value=f"{boss} Deaths").font = F_BODY

    raw_boss_dps = []
    raw_boss_deaths = []
    for date in sorted(all_dates):
        dps_vals = [rr["amount"] for rr in ranking_rows
                    if rr["date"] == date and rr["boss"] == boss
                    and rr["metric"] == "dps" and rr["role"] != "Healer" and rr["amount"]
                    and (not roster_locked or rr["character"] in rostered_chars)]
        boss_fought = any(fr["boss"] == boss and fr["date"] == date for fr in fight_rows)
        if boss_fought and dps_vals:
            raw_boss_dps.append(round(sum(dps_vals)/len(dps_vals)/1000))
        else:
            raw_boss_dps.append(None)
        if boss_fought:
            boss_deaths = [d for d in death_rows
                           if d["date"] == date and d["boss"] == boss
                           and (not roster_locked or d["character"] in rostered_chars)]
            raw_boss_deaths.append(len(boss_deaths))
        else:
            raw_boss_deaths.append(None)

    for i, v in enumerate(interpolate_series(raw_boss_dps)):
        ws.cell(row=r, column=2+i, value=v if v is not None else "").font = F_BODY
    for i, v in enumerate(interpolate_series(raw_boss_deaths)):
        ws.cell(row=r+1, column=2+i, value=v if v is not None else "").font = F_BODY
    r += 3

print("chart_data done.")

# ══════════════════════════════════════════════════════════════════
#  DYNAMIC CHART HELPER — d_chart_dyn sheet
# ══════════════════════════════════════════════════════════════════
# Charts can't reference formulas that change dynamically, but they CAN
# reference cells whose VALUES change via formulas. So we build a helper
# sheet (d_chart_dyn) with INDEX/MATCH formulas that pull the right boss's
# trend data based on which boss the Raid Performance panels are showing.
# Charts then reference d_chart_dyn, and when the dropdown changes, the
# formulas recalculate and the charts update.

num_dates = len(sorted(all_dates))
ws_rp = wb["Raid Performance"]

ws_dyn = wb.create_sheet("d_chart_dyn")

# Row 1: dates (same as chart_data)
ws_dyn.cell(row=1, column=1, value="Date").font = F_BOLD
for i in range(num_dates):
    date_col = get_column_letter(2 + i)
    ws_dyn.cell(row=1, column=2+i, value=f"=chart_data!{date_col}1")

# Row 2: Raid Avg DPS (static — doesn't change with dropdown)
ws_dyn.cell(row=2, column=1, value="Raid Avg DPS (k)")
for i in range(num_dates):
    date_col = get_column_letter(2 + i)
    ws_dyn.cell(row=2, column=2+i, value=f"=chart_data!{date_col}2")

# Row 3: Raid Deaths (static)
ws_dyn.cell(row=3, column=1, value="Raid Deaths")
for i in range(num_dates):
    date_col = get_column_letter(2 + i)
    ws_dyn.cell(row=3, column=2+i, value=f"=chart_data!{date_col}3")

# Row 4: Raid Avg Score (static)
ws_dyn.cell(row=4, column=1, value="Raid Avg Score")
for i in range(num_dates):
    date_col = get_column_letter(2 + i)
    ws_dyn.cell(row=4, column=2+i, value=f"=chart_data!{date_col}4")

# Rows 6+: Per-slot dynamic boss data (8 slots × 2 rows each, 3-row spacing)
for si in range(8):
    dps_row = 6 + si * 3
    deaths_row = dps_row + 1
    # Panel title cell reference
    panel_col = DETAIL_START_COL + si * PANEL_WIDTH
    title_ref = f"'Raid Performance'!${get_column_letter(panel_col)}${OVER_ROW}"

    # DPS label + values
    ws_dyn.cell(row=dps_row, column=1,
        value=f'=IF({title_ref}="","",{title_ref}&" DPS")')
    # Deaths label + values
    ws_dyn.cell(row=deaths_row, column=1,
        value=f'=IF({title_ref}="","",{title_ref}&" Deaths")')

    for i in range(num_dates):
        date_col = get_column_letter(2 + i)
        # DPS: look up boss name in chart_data column A, return value from same date column
        ws_dyn.cell(row=dps_row, column=2+i,
            value=f'=IFERROR(INDEX(chart_data!{date_col}:{date_col},MATCH({title_ref}&" DPS",chart_data!$A:$A,0)),"")')
        # Deaths: same approach
        ws_dyn.cell(row=deaths_row, column=2+i,
            value=f'=IFERROR(INDEX(chart_data!{date_col}:{date_col},MATCH({title_ref}&" Deaths",chart_data!$A:$A,0)),"")')

print("d_chart_dyn done.")

# ══════════════════════════════════════════════════════════════════
#  CHARTS — reference d_chart_dyn for dynamic data
# ══════════════════════════════════════════════════════════════════
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.line import LineProperties

def kill_legend(chart):
    """Remove legend and series names so Google Sheets can't auto-generate one."""
    chart.legend = None
    for s in chart.series:
        s.title = None

def hide_axes(chart):
    """Hide x and y axis labels/ticks."""
    chart.x_axis.delete = True
    chart.y_axis.delete = True

# Charts anchor right after the last detail player row
CHART_TOP_ROW = OVER_ROW + 3 + DETAIL_PLAYERS  # row after last player data

if num_dates >= 2:
    # ── Raid-wide chart (cols A-H) ──
    ch = LineChart()
    ch.title = "Raid-Wide: DPS, Deaths & Score"
    ch.style = 10
    ch.display_blanks = "span"

    cats = Reference(ws_dyn, min_col=2, max_col=1+num_dates, min_row=1)
    dps_ref = Reference(ws_dyn, min_col=2, max_col=1+num_dates, min_row=2)
    deaths_ref = Reference(ws_dyn, min_col=2, max_col=1+num_dates, min_row=3)
    comp_ref = Reference(ws_dyn, min_col=2, max_col=1+num_dates, min_row=4)

    ch.add_data(dps_ref, from_rows=True)
    ch.add_data(deaths_ref, from_rows=True)
    ch.add_data(comp_ref, from_rows=True)
    ch.set_categories(cats)

    # Blue solid = DPS
    ch.series[0].graphicalProperties.line.width = 25000
    ch.series[0].graphicalProperties.line.solidFill = "4472C4"
    # Red solid = Deaths
    if len(ch.series) > 1:
        ch.series[1].graphicalProperties.line.width = 25000
        ch.series[1].graphicalProperties.line.solidFill = "C0504D"
    # Green solid = Score
    if len(ch.series) > 2:
        ch.series[2].graphicalProperties.line.width = 25000
        ch.series[2].graphicalProperties.line.solidFill = "2E7D32"

    kill_legend(ch)
    hide_axes(ch)
    ch.width = 22
    ch.height = 10
    # Anchor below the player summary (left side), not at panel row
    RAIDCHART_ROW = PLAY_DATA_START + MAX_PLAYERS + 1
    ws_rp.add_chart(ch, f"A{RAIDCHART_ROW}")

    # ── Per-boss charts — 8 slots ──
    for si in range(8):
        dps_row = 6 + si * 3
        deaths_row = dps_row + 1

        bch = LineChart()
        bch.title = None
        bch.style = 10
        bch.display_blanks = "span"

        cats = Reference(ws_dyn, min_col=2, max_col=1+num_dates, min_row=1)
        dps_ref = Reference(ws_dyn, min_col=2, max_col=1+num_dates, min_row=dps_row)
        deaths_ref = Reference(ws_dyn, min_col=2, max_col=1+num_dates, min_row=deaths_row)

        bch.add_data(dps_ref, from_rows=True)
        bch.add_data(deaths_ref, from_rows=True)
        bch.set_categories(cats)

        # Blue solid = DPS
        bch.series[0].graphicalProperties.line.width = 25000
        bch.series[0].graphicalProperties.line.solidFill = "4472C4"
        # Red solid = Deaths
        if len(bch.series) >= 2:
            bch.series[1].graphicalProperties.line.width = 25000
            bch.series[1].graphicalProperties.line.solidFill = "C0504D"

        kill_legend(bch)
        hide_axes(bch)

        # Width: fit within 5 panel columns
        bch.width = 8
        bch.height = 7

        pc = DETAIL_START_COL + si * PANEL_WIDTH
        anchor_cell = f"{get_column_letter(pc)}{CHART_TOP_ROW}"
        ws_rp.add_chart(bch, anchor_cell)

    # Legend note row beneath all charts
    legend_row = CHART_TOP_ROW + 16
    c = ws_rp.cell(row=legend_row, column=DETAIL_START_COL,
                    value="Blue = DPS  |  Red = Deaths  |  Green = Score")
    c.font = F_SMALL

    print(f"Charts: 9 charts at row {CHART_TOP_ROW} (dynamic via d_chart_dyn)")
else:
    print("Charts: skipped (need 2+ dates)")

# ══════════════════════════════════════════════════════════════════
#  SCORECARD — Visible sheet with date dropdown
# ══════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Scorecard")
SC_COLS = 14  # A through N

title(ws, 1, "PLAYER SCORECARD", SC_COLS)

# ── Row 2: dropdown ──
ws.cell(row=2, column=1, value="Raid Night:").font = F_BOLD
# this_week already defined as all_dates[0] (newest date, reverse-sorted)
ws.cell(row=2, column=2, value=this_week).font = F_BOLD
ws.cell(row=2, column=2).fill = PatternFill("solid", fgColor="FFFF00")

sc_date_str = ",".join(all_dates)
sc_dv = DataValidation(type="list", formula1=f'"{sc_date_str}"', allow_blank=False)
sc_dv.prompt = "Select a raid night"
sc_dv.promptTitle = "Raid Night"
ws.add_data_validation(sc_dv)
sc_dv.add(ws["B2"])

ws.cell(row=2, column=4, value="Weights:").font = Font(name=FN, size=9, italic=True)
ws.cell(row=2, column=5, value=f"Mechs {int(SCORE_WEIGHTS['mechanics']*100)}%  Deaths {int(SCORE_WEIGHTS['deaths']*100)}%  Parse {int(SCORE_WEIGHTS['parse_performance']*100)}%  Pots {int(SCORE_WEIGHTS['consumables']*100)}%").font = Font(name=FN, size=9, italic=True)

# ── Row 3: Grade legend ──
grade_colors = {"A": "2E7D32", "B": "1565C0", "C": "F57F17", "D": "E65100", "F": "C62828"}
grade_names = {"A": "Excellent", "B": "Good", "C": "Average", "D": "Below Avg", "F": "Needs Work"}
col = 1
for g in ["A", "B", "C", "D", "F"]:
    c = ws.cell(row=3, column=col, value=f"{g}: {grade_names[g]}")
    c.font = Font(name=FN, size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=grade_colors[g])
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    col += 3

# ── Row 4: Headers ──
SC_HDRS = ["#", "Player", "Role", "Grade", "Score",
           "Mechs", "Deaths", "Parse", "Pots",
           "Fails", "Deaths #", "Fights", "Trend"]
section(ws, 4, "OVERALL RANKINGS", SC_COLS)
hdr(ws, 5, SC_HDRS)

# ── Build lookup: d_scores row index for XLOOKUP ──
# Data rows in d_scores start at row 2 (row 1 = header)
# Use INDEX/MATCH: =INDEX(d_scores!I:I,MATCH(1,(d_scores!A:A=<player>)*(d_scores!B:B=$B$2),0))
# But that's CSE. Use SUMPRODUCT or hardcode from Python.

# Since XLOOKUP is not reliable in all Excel versions, just build static rows
# that we re-sort by composite. The dropdown is a visual reference; the data
# is static for the last date. We write ALL dates to d_scores for manual use.

# Get latest date scores, sorted by composite descending
latest_scores = [r for r in score_rows if r["date"] == this_week]
if roster_locked:
    latest_scores = [r for r in latest_scores if r["player"] in rostered_players]
latest_scores.sort(key=lambda x: x["composite"], reverse=True)

if not latest_scores:
    latest_scores = sorted(score_rows, key=lambda x: x["composite"], reverse=True)
    if roster_locked:
        latest_scores = [r for r in latest_scores if r["player"] in rostered_players]

for i, r in enumerate(latest_scores):
    row_num = 6 + i
    vals = [
        i + 1, r["player"], r["role"], r["grade"], r["composite"],
        r["mech_score"] if r["mech_score"] is not None else "",
        r["death_score"],
        r["parse_score"] if r["parse_score"] is not None else "",
        r["con_score"] if r["con_score"] is not None else "",
        r["mech_fails"], r["deaths_raw"], r["fights"],
        "",  # Trend placeholder
    ]
    for j, v in enumerate(vals):
        c = ws.cell(row=row_num, column=1+j, value=v)
        c.font = F_BODY
        c.border = BT
        if i % 2 == 0:
            c.fill = X_EVEN

    # Color the grade cell
    gc = ws.cell(row=row_num, column=4)
    g = r["grade"]
    if g in grade_colors:
        gc.fill = PatternFill("solid", fgColor=grade_colors[g])
        gc.font = Font(name=FN, size=10, bold=True, color="FFFFFF")

    # Color composite score
    cc = ws.cell(row=row_num, column=5)
    if r["composite"] >= 80:
        cc.font = Font(name=FN, size=10, bold=True, color="2E7D32")
    elif r["composite"] >= 60:
        cc.font = Font(name=FN, size=10, color="1565C0")
    elif r["composite"] < 40:
        cc.font = Font(name=FN, size=10, color="C62828")

    # Trend: compare to previous date
    prev_dates = [d for d in all_dates if d < this_week]
    if prev_dates:
        prev = prev_dates[-1]
        prev_row = next((s for s in score_rows if s["player"] == r["player"] and s["date"] == prev), None)
        if prev_row:
            diff = r["composite"] - prev_row["composite"]
            trend_c = ws.cell(row=row_num, column=13)
            if diff > 0:
                trend_c.value = f"▲ +{diff}"
                trend_c.font = Font(name=FN, size=10, color="2E7D32")
            elif diff < 0:
                trend_c.value = f"▼ {diff}"
                trend_c.font = Font(name=FN, size=10, color="C62828")
            else:
                trend_c.value = "─"

# ── Per-boss grades section ──
boss_section_row = 6 + len(latest_scores) + 2
section(ws, boss_section_row, "PER-BOSS BREAKDOWN (latest night)", SC_COLS)
boss_section_row += 1

date_bosses = sorted(set(r["boss"] for r in fight_rows if str(r["date"]) == this_week),
                     key=lambda b: BOSS_ORDER.index(b) if b in BOSS_ORDER else 99)

if date_bosses:
    # Header row: Player + boss short names
    boss_hdrs = ["#", "Player"] + [boss_short(b) for b in date_bosses]
    hdr(ws, boss_section_row, boss_hdrs)
    boss_section_row += 1

    for i, sr in enumerate(latest_scores):
        row_num = boss_section_row + i
        ws.cell(row=row_num, column=1, value=i+1).font = F_BODY
        ws.cell(row=row_num, column=2, value=sr["player"]).font = F_BODY

        for j, boss in enumerate(date_bosses):
            br = next((b for b in boss_score_rows
                       if b["player"] == sr["player"] and b["boss"] == boss and b["date"] == this_week), None)
            if br:
                # Mini composite for this boss
                comps = {"deaths": br["death_score"]}
                if br["mech_score"] is not None:
                    comps["mechanics"] = br["mech_score"]
                if br["parse_score"] is not None:
                    comps["parse_performance"] = br["parse_score"]
                if br["con_score"] is not None:
                    comps["consumables"] = br["con_score"]

                aw = sum(SCORE_WEIGHTS[k] for k in comps)
                boss_composite = round(sum(comps[k] * (SCORE_WEIGHTS[k] / aw) for k in comps)) if aw > 0 else 50
                boss_composite = min(100, max(0, boss_composite))
                boss_grade = _get_grade(boss_composite)

                c = ws.cell(row=row_num, column=3+j, value=f"{boss_grade} {boss_composite}")
                if boss_grade in grade_colors:
                    c.fill = PatternFill("solid", fgColor=grade_colors[boss_grade])
                    c.font = Font(name=FN, size=9, bold=True, color="FFFFFF")
            else:
                ws.cell(row=row_num, column=3+j, value="—").font = Font(name=FN, size=9, color="999999")

            c = ws.cell(row=row_num, column=3+j)
            c.border = BT
            if i % 2 == 0 and not c.fill.fgColor or c.fill.fgColor == "00000000":
                c.fill = X_EVEN

        ws.cell(row=row_num, column=1).border = BT
        ws.cell(row=row_num, column=2).border = BT
        if i % 2 == 0:
            ws.cell(row=row_num, column=1).fill = X_EVEN
            ws.cell(row=row_num, column=2).fill = X_EVEN

widths(ws, [4, 16, 6, 7, 10, 11, 7, 7, 6, 7, 9, 7, 10])
ws.freeze_panes = "C6"
print("Scorecard done.")

# ══════════════════════════════════════════════════════════════════
#  FINALIZE — Single output file
# ══════════════════════════════════════════════════════════════════

# Ensure output directory exists
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Remove raw data sheets — not needed in display file
for name in ["d_roster", "d_fights", "d_rankings", "d_deaths", "d_gear"]:
    if name in wb.sheetnames:
        del wb[name]

# Order sheets: display first, then hidden support
display_sheets = ["Summary", "Raids", "Raid Performance",
                  "Character View", "Roster", "Scorecard"]
formula_support = ["d_rp_boss", "d_rp_player", "d_rp_detail", "chart_data", "d_chart_dyn",
                   "d_cv_info", "d_cv_boss", "d_cv_deaths", "d_cv_attend", "d_cv_analyzer", "d_cv_mechfails", "d_cv_chart",
                   "d_scores", "d_boss_scores", "d_mech_detail", "d_cv_topkb"]

desired_order = display_sheets + formula_support
for i, name in enumerate(desired_order):
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=i - idx)

# Hide formula-support tabs
for name in formula_support:
    if name in wb.sheetnames:
        wb[name].sheet_state = "hidden"
        wb[name].freeze_panes = "A2"

# Build output filename: "TeamName [date_slug].xlsx"
_team_name = roster.get("team", {}).get("team_name", "Raid Tracker")
_date_slug = this_week if this_week else datetime.now(timezone.utc).strftime("%Y-%m-%d")
DISPLAY_PATH = os.path.join(OUTPUT_DIR, f"{_team_name} [{_date_slug}].xlsx")
wb.save(DISPLAY_PATH)

# ── XML post-processing: inject full dropdown list into Character View ──
# openpyxl caps inline lists at 255 chars; Google Sheets handles longer strings fine.
# We crack open the xlsx ZIP and replace the PLACEHOLDER with the real list.
import zipfile, shutil, tempfile

# Find which sheet XML corresponds to "Character View"
cv_sheet_idx = wb.sheetnames.index("Character View") + 1
cv_xml_path = f"xl/worksheets/sheet{cv_sheet_idx}.xml"

tmp_path = DISPLAY_PATH + ".tmp"
with zipfile.ZipFile(DISPLAY_PATH, 'r') as zin:
    with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == cv_xml_path:
                xml_str = data.decode("utf-8")
                # Replace the placeholder with real comma-separated list
                # openpyxl writes literal quotes: "PLACEHOLDER"
                escaped_dd = CV_DROPDOWN_STRING.replace('&', '&amp;').replace('<', '&lt;')
                xml_str = xml_str.replace(
                    '"PLACEHOLDER"',
                    '"' + escaped_dd + '"'
                )
                data = xml_str.encode("utf-8")
                print(f"  Injected dropdown ({len(CV_DROPDOWN_STRING)} chars) into {cv_xml_path}")
            zout.writestr(item, data)

shutil.move(tmp_path, DISPLAY_PATH)

d_size = os.path.getsize(DISPLAY_PATH) / (1024 * 1024)
print(f"\nOutput: {DISPLAY_PATH} ({d_size:.1f} MB)")
print(f"  Sheets: {wb.sheetnames}")
print("\nDone.")
